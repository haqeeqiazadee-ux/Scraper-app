#!/usr/bin/env python3
"""
Generate the hard-coded Apify actor catalog from the mapping CSV.

Usage:
    python scripts/generate_actor_catalog.py

Reads:  docs/apify_catalog_implementation_mapping.csv
Writes: packages/core/actor_catalog/generated/apify_actor_catalog.json
        apps/web/src/data/apifyActors.generated.json  (compact frontend copy)

Deterministic and idempotent — same CSV always produces same JSON output.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = REPO_ROOT / "docs" / "apify_catalog_implementation_mapping.csv"
STORE_XLSX_PATH = REPO_ROOT / "apify_store_catalog.xlsx"
BACKEND_OUT = REPO_ROOT / "packages" / "core" / "actor_catalog" / "generated" / "apify_actor_catalog.json"
FRONTEND_OUT = REPO_ROOT / "apps" / "web" / "src" / "data" / "apifyActors.generated.json"
PUBLIC_CHUNKS_DIR = REPO_ROOT / "apps" / "web" / "public" / "data" / "actors"
CHUNK_SIZE = 1000

ROUTE_STRATEGY_MAP: dict[str, str] = {
    "Natively Supported by Pipeline": "native_pipeline",
    "yt-dlp integration": "yt_dlp",
    "Job Board Pydantic schemas": "job_board_schema",
    "Real Estate Pydantic schemas": "real_estate_schema",
}


def classify_route_strategy(missing_components: str) -> str:
    """Determine the route strategy from the Missing/Required Components field."""
    mc = missing_components.strip()
    if not mc or mc == "Missing/Required Components":
        return "unsupported"
    # Exact match first
    if mc in ROUTE_STRATEGY_MAP:
        return ROUTE_STRATEGY_MAP[mc]
    # Mixed families — pick the first known family, mark as mixed
    parts = [p.strip() for p in mc.split("|")]
    strategies = []
    for part in parts:
        if part in ROUTE_STRATEGY_MAP:
            strategies.append(ROUTE_STRATEGY_MAP[part])
    if strategies:
        return strategies[0]  # primary strategy
    return "unsupported"


def classify_runnable(strategy: str) -> str:
    """Determine if an actor is runnable now."""
    if strategy == "native_pipeline":
        return "runnable"
    if strategy in ("yt_dlp", "job_board_schema", "real_estate_schema"):
        return "runnable_with_schema"
    if strategy == "apify_api":
        return "runnable_via_apify"
    return "blocked"


def parse_categories(raw: str) -> list[str]:
    """Parse the categories field, handling quoted CSV values."""
    raw = raw.strip().strip('"')
    return [c.strip() for c in raw.split(",") if c.strip()]


def extract_username(name: str) -> str:
    """Extract username/owner from actor name if it contains a slash."""
    if "/" in name:
        return name.split("/")[0].strip()
    return ""


def generate_initials(title: str) -> str:
    """Generate 2-letter initials from actor title."""
    words = re.findall(r'[A-Za-z]+', title)
    if len(words) >= 2:
        return (words[0][0] + words[1][0]).upper()
    if words:
        return words[0][:2].upper()
    return "??"


def _safe_int(value: object) -> int:
    if value in (None, ""):
        return 0
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return 0


def _safe_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        return round(float(str(value)), 2)
    except (TypeError, ValueError):
        return None


def load_store_metadata() -> dict[str, dict]:
    """Load richer Apify Store metadata keyed by Actor ID."""
    if not STORE_XLSX_PATH.exists():
        return {}
    try:
        import openpyxl
    except ImportError:
        print("WARNING: openpyxl unavailable; continuing without store metadata", file=sys.stderr)
        return {}

    workbook = openpyxl.load_workbook(STORE_XLSX_PATH, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    metadata: dict[str, dict] = {}

    for row in rows:
        record = dict(zip(headers, row))
        actor_id = str(record.get("Actor ID") or "").strip()
        if not actor_id:
            continue
        metadata[actor_id] = {
            "developer": str(record.get("Developer") or "").strip(),
            "url": str(record.get("URL") or "").strip(),
            "total_runs": _safe_int(record.get("Total Runs")),
            "total_users": _safe_int(record.get("Total Users")),
            "rating": _safe_float(record.get("Rating")),
            "review_count": _safe_int(record.get("Review Count")),
            "bookmarks": _safe_int(record.get("Bookmarks")),
            "created_at": str(record.get("Created At") or "").strip(),
            "last_run_started_at": str(record.get("Last Run Started At") or "").strip(),
            "description": str(record.get("Description") or "").strip(),
        }

    workbook.close()
    return metadata


def main() -> int:
    if not CSV_PATH.exists():
        print(f"ERROR: CSV not found at {CSV_PATH}", file=sys.stderr)
        return 1

    # Read CSV
    with open(CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
        raw = f.read().replace("\r\n", "\n").replace("\r", "\n")

    reader = csv.DictReader(raw.splitlines())
    actors: list[dict] = []
    seen_ids: set[str] = set()
    skipped = 0
    store_metadata = load_store_metadata()

    for row in reader:
        actor_id = (row.get("Actor ID") or "").strip()
        if not actor_id or actor_id == "Actor ID":
            skipped += 1
            continue
        if actor_id in seen_ids:
            skipped += 1
            continue
        seen_ids.add(actor_id)

        name = (row.get("Name") or "").strip()
        title = (row.get("Title") or name).strip()
        categories_raw = row.get("Categories") or ""
        required_code = (row.get("Required Code/Logic Used") or "").strip()
        missing_components = (row.get("Missing/Required Components") or "").strip()

        strategy = classify_route_strategy(missing_components)
        categories = parse_categories(categories_raw)
        metadata = store_metadata.get(actor_id, {})
        developer = metadata.get("developer") or extract_username(name)
        description = metadata.get("description") or f"{title} - Apify actor"

        actor = {
            "actor_id": actor_id,
            "name": name,
            "title": title,
            "username": developer,
            "developer": developer,
            "url": metadata.get("url", ""),
            "description": f"{title} — Apify actor",
            "categories": categories,
            "description": description,
            "pricing_model": "unknown",
            "total_runs": metadata.get("total_runs", 0),
            "total_users": metadata.get("total_users", 0),
            "rating": metadata.get("rating"),
            "review_count": metadata.get("review_count", 0),
            "bookmarks": metadata.get("bookmarks", 0),
            "created_at": metadata.get("created_at", ""),
            "last_run_started_at": metadata.get("last_run_started_at", ""),
            "implementation_status": "mapped",
            "native_support_reason": required_code if strategy == "native_pipeline" else "",
            "missing_components": missing_components if strategy != "native_pipeline" else "",
            "route_strategy": strategy,
            "source": "apify_catalog_csv",
            "initials": generate_initials(title),
            "runnable_status": classify_runnable(strategy),
        }
        actors.append(actor)

    # Sort by actor_id for deterministic output
    actors.sort(key=lambda a: a["actor_id"])

    # Compute source checksum
    csv_hash = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]

    catalog = {
        "version": "1.0.0",
        "generated_at": "deterministic-from-source",
        "source_file": "docs/apify_catalog_implementation_mapping.csv",
        "source_checksum": csv_hash,
        "total_actors": len(actors),
        "skipped_rows": skipped,
        "actors": actors,
    }

    # Write backend catalog
    BACKEND_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(BACKEND_OUT, "w", encoding="utf-8") as f:
        json.dump(catalog, f, ensure_ascii=False, separators=(",", ":"))

    # Write frontend catalog (compact: only fields the UI needs)
    frontend_actors = []
    for a in actors:
        frontend_actors.append({
            "id": a["actor_id"],
            "name": a["name"],
            "title": a["title"],
            "username": a["username"],
            "developer": a["developer"],
            "url": a["url"],
            "description": a["description"],
            "categories": a["categories"],
            "pricing_model": a["pricing_model"],
            "total_runs": a["total_runs"],
            "total_users": a["total_users"],
            "rating": a["rating"],
            "review_count": a["review_count"],
            "bookmarks": a["bookmarks"],
            "initials": a["initials"],
            "route_strategy": a["route_strategy"],
            "runnable_status": a["runnable_status"],
            "missing_components": a["missing_components"],
        })

    frontend_catalog = {
        "version": "1.0.0",
        "total": len(frontend_actors),
        "actors": frontend_actors,
    }
    FRONTEND_OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(FRONTEND_OUT, "w", encoding="utf-8") as f:
        json.dump(frontend_catalog, f, ensure_ascii=False, separators=(",", ":"))

    # Stats
    strategy_counts: dict[str, int] = {}
    category_counts: dict[str, int] = {}
    runnable_counts: dict[str, int] = {}
    developer_counts: dict[str, int] = {}
    pricing_counts: dict[str, int] = {}
    for a in actors:
        s = a["route_strategy"]
        strategy_counts[s] = strategy_counts.get(s, 0) + 1
        r = a["runnable_status"]
        runnable_counts[r] = runnable_counts.get(r, 0) + 1
        developer = a.get("developer") or "unknown"
        developer_counts[developer] = developer_counts.get(developer, 0) + 1
        pricing = a.get("pricing_model") or "unknown"
        pricing_counts[pricing] = pricing_counts.get(pricing, 0) + 1
        for c in a["categories"]:
            category_counts[c] = category_counts.get(c, 0) + 1

    # Write public frontend chunks used by /actors without importing the full
    # catalog into the Vite bundle.
    PUBLIC_CHUNKS_DIR.mkdir(parents=True, exist_ok=True)
    for stale in PUBLIC_CHUNKS_DIR.glob("chunk-*.json"):
        stale.unlink()

    chunks = []
    for chunk_index, start in enumerate(range(0, len(frontend_actors), CHUNK_SIZE)):
        chunk = frontend_actors[start: start + CHUNK_SIZE]
        chunk_path = PUBLIC_CHUNKS_DIR / f"chunk-{chunk_index}.json"
        with open(chunk_path, "w", encoding="utf-8") as f:
            json.dump(chunk, f, ensure_ascii=False, separators=(",", ":"))
        chunks.append({
            "index": chunk_index,
            "start": start,
            "count": len(chunk),
        })

    chunk_index_payload = {
        "version": "1.0.0",
        "total": len(frontend_actors),
        "chunk_size": CHUNK_SIZE,
        "chunk_count": len(chunks),
        "chunks": chunks,
    }
    with open(PUBLIC_CHUNKS_DIR / "index.json", "w", encoding="utf-8") as f:
        json.dump(chunk_index_payload, f, ensure_ascii=False, separators=(",", ":"))

    stats_payload = {
        "version": "1.0.0",
        "total": len(frontend_actors),
        "by_strategy": strategy_counts,
        "by_category": dict(sorted(category_counts.items(), key=lambda x: -x[1])),
        "by_runnable": runnable_counts,
        "by_developer": dict(sorted(developer_counts.items(), key=lambda x: (-x[1], x[0]))),
        "by_pricing_model": dict(sorted(pricing_counts.items(), key=lambda x: (-x[1], x[0]))),
        "categories": sorted(category_counts),
        "developers": sorted(developer_counts),
        "pricing_models": sorted(pricing_counts),
    }
    with open(PUBLIC_CHUNKS_DIR / "stats.json", "w", encoding="utf-8") as f:
        json.dump(stats_payload, f, ensure_ascii=False, separators=(",", ":"))

    print(f"Generated {len(actors)} actors ({skipped} rows skipped)")
    print(f"Backend:  {BACKEND_OUT} ({BACKEND_OUT.stat().st_size:,} bytes)")
    print(f"Frontend: {FRONTEND_OUT} ({FRONTEND_OUT.stat().st_size:,} bytes)")
    print(f"Chunks:   {PUBLIC_CHUNKS_DIR} ({len(chunks)} chunks)")
    print(f"Store metadata matched: {sum(1 for a in actors if a.get('url'))}")
    print(f"\nRoute strategies:")
    for s, c in sorted(strategy_counts.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}")
    print(f"\nRunnable status:")
    for r, c in sorted(runnable_counts.items(), key=lambda x: -x[1]):
        print(f"  {r}: {c}")
    print(f"\nTop categories:")
    for cat, c in sorted(category_counts.items(), key=lambda x: -x[1])[:15]:
        print(f"  {cat}: {c}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
