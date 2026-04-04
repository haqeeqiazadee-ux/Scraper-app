#!/usr/bin/env python3
"""Generate E2E test results tracking Excel spreadsheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Test Results"

headers = ["#", "Test File", "Test Name", "Category", "Workflow Tested", "Expected Result", "Actual Result", "Status", "Failure Reason", "Notes"]
hf = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", fgColor="2563EB")
bd = Border(left=Side("thin", color="D1D5DB"), right=Side("thin", color="D1D5DB"), top=Side("thin", color="D1D5DB"), bottom=Side("thin", color="D1D5DB"))

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center")
    c.border = bd

widths = [5, 28, 42, 14, 40, 30, 15, 10, 35, 12]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.auto_filter.ref = "A1:J1"
ws.freeze_panes = "A2"

fe = PatternFill("solid", fgColor="F3F4F6")
fo = PatternFill("solid", fgColor="FFFFFF")
pf = PatternFill("solid", fgColor="FEF3C7")

tests = [
    ("test_navigation.py", "test_app_opens_to_dashboard", "Navigation", "App startup redirect", "Dashboard loads"),
    ("test_navigation.py", "test_no_login_screen", "Navigation", "Login bypass", "Redirects to dashboard"),
    ("test_navigation.py", "test_sidebar_core", "Navigation", "Sidebar CORE group", "Group visible"),
    ("test_navigation.py", "test_sidebar_tools", "Navigation", "Sidebar TOOLS group", "HYDRA items visible"),
    ("test_navigation.py", "test_sidebar_integration", "Navigation", "INTEGRATION group", "MCP Server visible"),
    ("test_navigation.py", "test_sidebar_monitoring", "Navigation", "MONITORING group", "Group visible"),
    ("test_navigation.py", "test_nav_to_tasks", "Navigation", "Sidebar->Tasks", "URL /tasks"),
    ("test_navigation.py", "test_nav_to_crawl", "Navigation", "Sidebar->Crawl", "URL /crawl"),
    ("test_navigation.py", "test_nav_to_mcp", "Navigation", "Sidebar->MCP", "URL /mcp"),
    ("test_navigation.py", "test_404_redirects", "Navigation", "404 handling", "Redirects"),
    ("test_navigation.py", "test_active_highlight", "Navigation", "Active styling", "CSS active"),
    ("test_navigation.py", "test_nav_to_search", "Navigation", "Sidebar->Search", "URL /search"),
    ("test_dashboard.py", "test_title", "Dashboard", "Page load", "Title visible"),
    ("test_dashboard.py", "test_stat_cards", "Dashboard", "Stats display", "4 cards"),
    ("test_dashboard.py", "test_quick_actions", "Dashboard", "Quick actions", "4 original"),
    ("test_dashboard.py", "test_hydra_actions", "Dashboard", "HYDRA actions", "Crawl/Search/MCP"),
    ("test_dashboard.py", "test_action_crawl", "Dashboard", "Action->Crawl", "URL /crawl"),
    ("test_dashboard.py", "test_action_search", "Dashboard", "Action->Search", "URL /search"),
    ("test_dashboard.py", "test_action_mcp", "Dashboard", "Action->MCP", "URL /mcp"),
    ("test_dashboard.py", "test_health_badge", "Dashboard", "Health badge", "Badge visible"),
    ("test_dashboard.py", "test_recent_tasks", "Dashboard", "Tasks area", "Table or empty"),
    ("test_dashboard.py", "test_last_refreshed", "Dashboard", "Refresh time", "Timestamp"),
    ("test_tasks_e2e.py", "test_page_renders", "Tasks", "Page load", "Title visible"),
    ("test_tasks_e2e.py", "test_create_btn", "Tasks", "Create button", "Button visible"),
    ("test_tasks_e2e.py", "test_filters", "Tasks", "Filter buttons", "Filters visible"),
    ("test_tasks_e2e.py", "test_create_form", "Tasks", "Create form", "Form opens"),
    ("test_tasks_e2e.py", "test_empty_state", "Tasks", "Empty state", "Content present"),
    ("test_tasks_e2e.py", "test_filter_click", "Tasks", "Filter click", "Click responds"),
    ("test_tasks_e2e.py", "test_results_page", "Tasks", "Results page", "Content visible"),
    ("test_tasks_e2e.py", "test_policies_page", "Tasks", "Policies page", "Content visible"),
    ("test_crawl_e2e.py", "test_page_renders", "Crawl", "Page load", "Title visible"),
    ("test_crawl_e2e.py", "test_form_inputs", "Crawl", "Form inputs", "All inputs"),
    ("test_crawl_e2e.py", "test_start_btn", "Crawl", "Start button", "Button visible"),
    ("test_crawl_e2e.py", "test_url_input", "Crawl", "URL input", "Can type"),
    ("test_crawl_e2e.py", "test_depth_default", "Crawl", "Default depth", "Value 3"),
    ("test_crawl_e2e.py", "test_pages_default", "Crawl", "Default pages", "Value 100"),
    ("test_crawl_e2e.py", "test_format_opts", "Crawl", "Format dropdown", "json/md/html"),
    ("test_crawl_e2e.py", "test_empty_url", "Crawl", "Empty URL", "Error shown"),
    ("test_crawl_e2e.py", "test_submit", "Crawl", "Submit form", "Loading state"),
    ("test_crawl_e2e.py", "test_sidebar", "Crawl", "Sidebar nav", "Navigates"),
    ("test_search_e2e.py", "test_page_renders", "Search", "Page load", "Title visible"),
    ("test_search_e2e.py", "test_query_input", "Search", "Query input", "Input visible"),
    ("test_search_e2e.py", "test_max_results", "Search", "Max results", "Dropdown"),
    ("test_search_e2e.py", "test_button", "Search", "Search button", "Button visible"),
    ("test_search_e2e.py", "test_empty_state", "Search", "Empty state", "Prompt shown"),
    ("test_search_e2e.py", "test_type_query", "Search", "Type query", "Can type"),
    ("test_search_e2e.py", "test_submit", "Search", "Submit search", "Loading"),
    ("test_search_e2e.py", "test_sidebar", "Search", "Sidebar nav", "Navigates"),
    ("test_extract_e2e.py", "test_page_renders", "Extract", "Page load", "Title visible"),
    ("test_extract_e2e.py", "test_url_input", "Extract", "URL input", "Input visible"),
    ("test_extract_e2e.py", "test_schema_area", "Extract", "Schema textarea", "Textarea visible"),
    ("test_extract_e2e.py", "test_button", "Extract", "Extract button", "Button visible"),
    ("test_extract_e2e.py", "test_placeholder", "Extract", "Placeholder", "Example JSON"),
    ("test_extract_e2e.py", "test_type_url", "Extract", "Type URL", "Can type"),
    ("test_extract_e2e.py", "test_type_schema", "Extract", "Type JSON", "Can type"),
    ("test_extract_e2e.py", "test_sidebar", "Extract", "Sidebar nav", "Navigates"),
    ("test_changes_e2e.py", "test_page_renders", "Changes", "Page load", "Title visible"),
    ("test_changes_e2e.py", "test_textareas", "Changes", "Textareas", "Old+New visible"),
    ("test_changes_e2e.py", "test_compare_btn", "Changes", "Compare button", "Button visible"),
    ("test_changes_e2e.py", "test_threshold", "Changes", "Threshold", "Input visible"),
    ("test_changes_e2e.py", "test_empty_state", "Changes", "Empty state", "No results"),
    ("test_changes_e2e.py", "test_invalid_json", "Changes", "Invalid JSON", "Error shown"),
    ("test_changes_e2e.py", "test_valid_compare", "Changes", "Valid diff", "Results shown"),
    ("test_changes_e2e.py", "test_added_items", "Changes", "Added items", "Added badge"),
    ("test_changes_e2e.py", "test_price_change", "Changes", "Price change", "Delta shown"),
    ("test_changes_e2e.py", "test_identical", "Changes", "Same data", "No changes"),
    ("test_mcp_e2e.py", "test_page_renders", "MCP", "Page load", "Title visible"),
    ("test_mcp_e2e.py", "test_quick_start", "MCP", "Quick start", "Instructions"),
    ("test_mcp_e2e.py", "test_tool_cards", "MCP", "Tool cards", "5 cards"),
    ("test_mcp_e2e.py", "test_tool_names", "MCP", "Tool names", "All 5 names"),
    ("test_mcp_e2e.py", "test_config", "MCP", "Config sections", "Claude/VS Code"),
    ("test_mcp_e2e.py", "test_code_blocks", "MCP", "Code blocks", "Pre visible"),
    ("test_mcp_e2e.py", "test_copy_btns", "MCP", "Copy buttons", "Buttons visible"),
    ("test_mcp_e2e.py", "test_no_api", "MCP", "No API calls", "Info only"),
    ("test_monitoring_e2e.py", "test_templates", "Monitoring", "Templates", "Content loads"),
    ("test_monitoring_e2e.py", "test_route_tester", "Monitoring", "Route tester", "Content loads"),
    ("test_monitoring_e2e.py", "test_scrape_tester", "Monitoring", "Scrape tester", "Page loads"),
    ("test_monitoring_e2e.py", "test_amazon", "Monitoring", "Amazon page", "Content loads"),
    ("test_monitoring_e2e.py", "test_amazon_types", "Monitoring", "Amazon types", "ASIN visible"),
    ("test_monitoring_e2e.py", "test_maps", "Monitoring", "Maps page", "Content loads"),
    ("test_monitoring_e2e.py", "test_sessions", "Monitoring", "Sessions", "Page loads"),
    ("test_monitoring_e2e.py", "test_proxies", "Monitoring", "Proxies", "Page loads"),
    ("test_monitoring_e2e.py", "test_webhooks", "Monitoring", "Webhooks", "Page loads"),
    ("test_monitoring_e2e.py", "test_billing", "Monitoring", "Billing", "Content loads"),
    ("test_api_full_e2e.py", "test_health", "API", "GET /health", "200 healthy"),
    ("test_api_full_e2e.py", "test_ready", "API", "GET /ready", "200"),
    ("test_api_full_e2e.py", "test_create_task", "API", "POST /tasks", "201 + id"),
    ("test_api_full_e2e.py", "test_list_tasks", "API", "GET /tasks", "200 + items"),
    ("test_api_full_e2e.py", "test_get_task", "API", "GET /tasks/{id}", "200"),
    ("test_api_full_e2e.py", "test_create_policy", "API", "POST /policies", "201"),
    ("test_api_full_e2e.py", "test_list_policies", "API", "GET /policies", "200"),
    ("test_api_full_e2e.py", "test_list_results", "API", "GET /results", "200"),
    ("test_api_full_e2e.py", "test_templates", "API", "GET /templates", "200"),
    ("test_api_full_e2e.py", "test_categories", "API", "GET /templates/cat", "200"),
    ("test_api_full_e2e.py", "test_crawl_start", "API", "POST /crawl", "201"),
    ("test_api_full_e2e.py", "test_crawl_404", "API", "GET /crawl/bad", "404"),
    ("test_api_full_e2e.py", "test_search", "API", "POST /search", "200 or 503"),
    ("test_api_full_e2e.py", "test_extract", "API", "POST /extract", "200 or 422"),
    ("test_api_full_e2e.py", "test_schedules", "API", "GET /schedules", "200"),
    ("test_api_full_e2e.py", "test_metrics", "API", "GET /metrics", "200"),
    ("test_api_full_e2e.py", "test_sessions", "API", "GET /sessions", "200"),
    ("test_api_full_e2e.py", "test_webhooks_api", "API", "GET /webhooks", "200"),
    ("test_api_full_e2e.py", "test_404", "API", "GET /nonexistent", "404"),
    ("test_api_full_e2e.py", "test_cors", "API", "OPTIONS /health", "CORS headers"),
]

for i, (f, name, cat, wf, exp) in enumerate(tests, 1):
    r = i + 1
    fill = fe if i % 2 == 0 else fo
    for col, val in enumerate([i, f, name, cat, wf, exp, "Pending", "Pending", "", "Sandbox"], 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.border = bd
        c.alignment = Alignment(vertical="center")
        c.fill = pf if col == 8 else fill

# Summary sheet
ws2 = wb.create_sheet("Summary")
for col, h in enumerate(["Category", "Total", "Passed", "Failed", "Pending", "Pass Rate"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center")
    c.border = bd

cats = [("Navigation", 12), ("Dashboard", 10), ("Tasks", 8), ("Crawl", 10),
        ("Search", 8), ("Extract", 8), ("Changes", 10), ("MCP", 8),
        ("Monitoring", 10), ("API Backend", 20)]
for i, (cat, t) in enumerate(cats, 2):
    for col, val in enumerate([cat, t, 0, 0, t, "0%"], 1):
        c = ws2.cell(row=i, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = fe if i % 2 == 0 else fo
        c.border = bd

tr = len(cats) + 2
for col, val in enumerate(["TOTAL", len(tests), 0, 0, len(tests), "0%"], 1):
    c = ws2.cell(row=tr, column=col, value=val)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1E40AF")
    c.border = bd

for i in range(1, 7):
    ws2.column_dimensions[get_column_letter(i)].width = [18, 14, 10, 10, 10, 12][i - 1]

# Live tests sheet
ws3 = wb.create_sheet("Live Tests (Phase 2)")
for col, h in enumerate(["Service", "Test Action", "API Key", "Cost", "Status", "Notes"], 1):
    c = ws3.cell(row=1, column=col, value=h)
    c.font = hf
    c.fill = PatternFill("solid", fgColor="059669")
    c.alignment = Alignment(horizontal="center")
    c.border = bd

live = [
    ("Keepa API", "ASIN lookup", "KEEPA_API_KEY", "~1 token", "Pending", "Configured"),
    ("Google Maps", "Business search", "SERVICE_ACCOUNT", "USD 0.032", "Pending", "Check billing"),
    ("Brave Search", "Web search", "BRAVE_SEARCH_API_KEY", "Free", "Pending", "Sign up"),
    ("HTTP Scrape", "books.toscrape.com", "None", "Free", "Pending", "No key"),
    ("Browser Scrape", "JS page", "None", "Free", "Pending", "Playwright"),
    ("OpenAI", "AI normalize", "OPENAI_API_KEY", "~USD 0.01", "Pending", "Check credit"),
]
for i, (s, t, k, co, st, n) in enumerate(live, 2):
    for col, val in enumerate([s, t, k, co, st, n], 1):
        c = ws3.cell(row=i, column=col, value=val)
        c.font = Font(name="Arial", size=10)
        c.fill = fe if i % 2 == 0 else fo
        c.border = bd

for i in range(1, 7):
    ws3.column_dimensions[get_column_letter(i)].width = [18, 25, 22, 15, 12, 20][i - 1]

wb.save("docs/E2E_TEST_RESULTS.xlsx")
print(f"Created docs/E2E_TEST_RESULTS.xlsx with {len(tests)} test cases across 3 sheets")
