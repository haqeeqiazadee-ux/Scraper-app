"""Create WORKFLOW_FIX_LOG.xlsx — tracks every fix with challenge verification."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Workflow Fix Log"

headers = ["#", "Workflow", "Challenge", "Test URL / Action", "Expected Result",
           "Actual Result", "Status", "Fix Applied", "Verified On", "Evidence", "Blocker"]
hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hfill = PatternFill("solid", fgColor="1E40AF")
bd = Border(left=Side("thin", color="D1D5DB"), right=Side("thin", color="D1D5DB"),
            top=Side("thin", color="D1D5DB"), bottom=Side("thin", color="D1D5DB"))

for col, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = hf
    c.fill = hfill
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    c.border = bd

widths = [4, 16, 38, 38, 30, 30, 10, 38, 14, 25, 25]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

rows = [
    (1, "Quick Scrape", "Scrape yousell.online -> 20+ data points",
     "https://yousell.online", "Title, headings, plans, features, CTAs",
     "", "PENDING", "", "", "", ""),
    (2, "Quick Scrape", "Extraction mode selector (Everything/Products/Content)",
     "UI: ScrapeTestPage.tsx", "Dropdown with modes",
     "", "PENDING", "", "", "", ""),
    (3, "Quick Scrape", "Save results to Results page after scrape",
     "POST /results after extraction", "Result in Results & Export",
     "", "PENDING", "", "", "", ""),
    (4, "Results", "Quick Scrape results appear in Results within 5s",
     "myscraper.netlify.app/results", "New result row",
     "", "PENDING", "", "", "", ""),
    (5, "Results", "Export CSV downloads real file",
     "Click Export Results", "CSV file with data",
     "", "PENDING", "", "", "", ""),
    (6, "Results", "Export JSON downloads real file",
     "Click Export Results", "JSON file with data",
     "", "PENDING", "", "", "", ""),
    (7, "Changes", "Widget $19.99->$14.99 + new Gadget",
     "myscraper.netlify.app/changes", "1 added, 1 price change (-25%)",
     "", "PENDING", "", "", "", ""),
    (8, "MCP Server", "Copy Claude Desktop config = valid JSON",
     "Click Copy button", "Valid JSON pasted",
     "", "PENDING", "", "", "", ""),
    (9, "Web Search", "Search gaming laptops -> 3+ results",
     "myscraper.netlify.app/search", "3+ results with titles",
     "", "PENDING", "", "", "", ""),
    (10, "Web Search", "Serper API works on deployed site",
     "POST /api/v1/search live", "HTTP 200",
     "", "PENDING", "", "", "", ""),
    (11, "Extract", "Extract 4 fields from book page",
     "books.toscrape.com/catalogue/...", "title, price, desc, availability",
     "", "PENDING", "", "", "", ""),
    (12, "Amazon", "ASIN B09V3KXJPB -> iPad with price",
     "myscraper.netlify.app/amazon", "Product card with iPad",
     "", "PENDING", "", "", "", ""),
    (13, "Amazon", "Keepa API works on deployed site",
     "POST /api/v1/keepa/query live", "HTTP 200 with product",
     "", "PENDING", "", "", "", ""),
    (14, "Google Maps", "Restaurants in Dubai -> 5+ businesses",
     "myscraper.netlify.app/google-maps", "5+ names with ratings",
     "", "PENDING", "", "", "", ""),
    (15, "Google Maps", "Serper Places works on deployed site",
     "POST /api/v1/maps/search live", "HTTP 200 with businesses",
     "", "PENDING", "", "", "", ""),
    (16, "FB Groups", "Shows requirements message in UI",
     "myscraper.netlify.app/facebook-groups", "Note: needs self-hosted backend",
     "", "PENDING", "", "", "", "Needs server-side browser"),
    (17, "Templates", "Trustpilot template -> Start -> results",
     "trustpilot.com/review/amazon.com", "Task with extracted reviews",
     "", "PENDING", "", "", "", ""),
    (18, "Templates", "eBay template -> Start -> results",
     "ebay.co.uk/sch/i.html?_nkw=gpu", "Task with extracted items",
     "", "PENDING", "", "", "", ""),
    (19, "Schedules", "Create -> list -> delete works",
     "myscraper.netlify.app/schedules", "Schedule CRUD works",
     "", "PENDING", "", "", "", ""),
    (20, "Web Crawl", "Crawl books.toscrape depth=1 -> 3+ pages",
     "https://books.toscrape.com/", "3+ pages with data",
     "", "PENDING", "", "", "", "May need self-hosted"),
    (21, "End-to-End", "Scrape -> Results -> Export CSV",
     "Any URL -> Results -> Export", "CSV with scraped data",
     "", "PENDING", "", "", "", ""),
    (22, "End-to-End", "Template -> Task -> Execute -> Results",
     "Any template -> Start Scrape", "Task with extracted data",
     "", "PENDING", "", "", "", ""),
]

pending_fill = PatternFill("solid", fgColor="FEF3C7")

for row_data in rows:
    r = row_data[0] + 1
    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=r, column=col, value=val)
        c.font = Font(name="Arial", size=9)
        c.border = bd
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if col == 7:
            c.fill = pending_fill
            c.alignment = Alignment(horizontal="center", vertical="top")

# Summary
ws2 = wb.create_sheet("Summary")
sh = ["Metric", "Value"]
for col, h in enumerate(sh, 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font = hf
    c.fill = hfill
    c.border = bd

summary = [
    ("Total tests", len(rows)),
    ("Passed", 0),
    ("Failed", 0),
    ("Pending", len(rows)),
    ("Blocked", 0),
    ("Pass rate", "0%"),
    ("", ""),
    ("Workflows covered", 12),
    ("Quick Scrape tests", 3),
    ("Results tests", 3),
    ("Change Detection tests", 1),
    ("MCP Server tests", 1),
    ("Web Search tests", 2),
    ("Structured Extract tests", 1),
    ("Amazon tests", 2),
    ("Google Maps tests", 2),
    ("Facebook Groups tests", 1),
    ("Templates tests", 2),
    ("Schedules tests", 1),
    ("Web Crawl tests", 1),
    ("End-to-End tests", 2),
]

for i, (metric, val) in enumerate(summary, 2):
    ws2.cell(row=i, column=1, value=metric).font = Font(name="Arial", size=10, bold=bool(metric))
    ws2.cell(row=i, column=2, value=val).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=1).border = bd
    ws2.cell(row=i, column=2).border = bd

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 12

path = "docs/WORKFLOW_FIX_LOG.xlsx"
wb.save(path)
print(f"Created: {path}")
print(f"Tests: {len(rows)} (ALL PENDING)")
