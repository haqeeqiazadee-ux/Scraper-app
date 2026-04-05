"""
Facebook Group Scraper -- extracts all posts from a Facebook group feed.

Pipeline: authenticate via cookies -> navigate to group URL -> scroll to load all posts ->
capture posts into JS memory (Facebook virtualizes DOM!) -> export to Excel with dynamic columns.

Key design decisions driven by live-testing on Facebook:
  1. Uses textContent everywhere (innerText returns "" due to Facebook CSS tricks).
  2. Scrolls the overflow container, not window (Facebook uses custom scroll containers).
  3. Captures posts into JS memory during scrolling (DOM virtualization removes old posts).
  4. Maps EditThisCookie's sameSite:"no_restriction" -> "None" and expirationDate -> expires.
  5. Supports CDP connection as fallback when chromium.launch() fails in sandboxes.

Usage:
    scraper = FacebookGroupScraper()
    posts = await scraper.scrape_group(
        url="https://www.facebook.com/groups/367202228711807/",
        cookies=[{"name": "c_user", "value": "...", "domain": ".facebook.com"}, ...],
    )
    scraper.export_to_excel(posts, "group_posts.xlsx")
"""

from __future__ import annotations

import asyncio
import json
import logging
import re
from datetime import datetime, timedelta, timezone
from typing import Any, Optional
from urllib.parse import urlparse

from bs4 import BeautifulSoup, Tag

from packages.core.ai_providers.social.base import (
    deep_get,
    extract_json_after_key,
    find_key_recursive,
    parse_count,
    parse_timestamp,
)
from packages.core.device_profiles import DeviceProfile

logger = logging.getLogger(__name__)

__all__ = ["FacebookGroupScraper"]

# ---------------------------------------------------------------------------
# Camoufox availability check
# ---------------------------------------------------------------------------

_HAS_CAMOUFOX = False
try:
    import camoufox  # noqa: F401
    _HAS_CAMOUFOX = True
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Currency / price patterns
# ---------------------------------------------------------------------------

_PRICE_PATTERN = re.compile(
    r"(?P<currency>[£$€¥₹])\s*(?P<price>[\d,]+(?:\.\d{1,2})?)"
    r"|(?P<price2>[\d,]+(?:\.\d{1,2})?)\s*(?P<currency2>[£$€¥₹])",
)

_CURRENCY_SYMBOLS: dict[str, str] = {
    "£": "GBP",
    "$": "USD",
    "€": "EUR",
    "¥": "JPY",
    "₹": "INR",
}

_CONDITION_PATTERNS = re.compile(
    r"\b(brand\s*new|like\s*new|new|used|refurbished|excellent|good\s*condition"
    r"|fair\s*condition|for\s*parts)\b",
    re.IGNORECASE,
)

# Priority order for Excel columns
_COLUMN_PRIORITY = [
    "author",
    "author_name",
    "listing_title",
    "description",
    "price",
    "location",
    "post_type",
    "image_count",
    "timestamp",
    "post_id",
    "currency",
    "condition",
    "like_count",
    "comment_count",
    "share_count",
    "image_urls",
    "post_url",
    "author_profile_url",
    "text",  # raw text last (debug)
]

# ---------------------------------------------------------------------------
# JavaScript constants for in-browser post collection
#
# CRITICAL: Facebook virtualizes the DOM -- posts disappear when scrolled
# past. We inject JS that captures posts into window.__fbPosts as they
# appear, BEFORE they get recycled out of DOM.
#
# CRITICAL: Facebook uses CSS tricks that make innerText return "".
# We use textContent everywhere.
#
# CRITICAL: Facebook uses a custom overflow scroll container (div with
# overflow-y: auto). We find this container and scrollBy on IT, not window.
# ---------------------------------------------------------------------------

COLLECTOR_INIT_JS = """
() => {
    window.__fbPosts = [];
    window.__fbSeen = new Set();
}
"""

CAPTURE_JS = """
() => {
    const feed = document.querySelector('div[role="feed"]');
    if (!feed) return { newPosts: 0, total: 0 };

    // Find content divs using textContent (NOT innerText -- Facebook makes innerText empty)
    const allDivs = feed.querySelectorAll('div');
    let n = 0;

    allDivs.forEach(d => {
        const t = d.textContent;
        if (!t || t.length < 40 || t.length > 3000) return;
        // Skip containers that have too many children (parent wrappers)
        if (d.childElementCount > 15) return;

        // Clean Facebook repeated text
        const clean = t.replace(/Facebook/g, '').trim();
        if (clean.length < 30) return;

        // Dedup by content signature
        const sig = clean.substring(0, 100);
        if (window.__fbSeen.has(sig)) return;
        window.__fbSeen.add(sig);

        // Extract structured fields
        const post = {};

        // Author: first multi-word capitalized name
        const authorMatch = clean.match(/^([A-Z][a-z]+(?:\\s+[A-Z][a-z]+)+)/);
        if (authorMatch) post.author = authorMatch[1];

        // Price
        const priceMatch = clean.match(/[\\u00a3$\\u20ac]\\s*[\\d,]+(?:\\.\\d{2})?/);
        if (priceMatch) {
            post.price = priceMatch[0];
            post.post_type = 'sale_listing';
        } else {
            post.post_type = 'discussion';
        }

        // Location: City, County/Region pattern
        const locMatch = clean.match(/([A-Z][a-z]+(?:\\s+[A-Z][a-z]+)*,\\s*[A-Z][a-z]+(?:\\s+[A-Z][a-z]+)*)/);
        if (locMatch) post.location = locMatch[1];

        // Images in this area
        const imgs = d.querySelectorAll('img[src*="scontent"]');
        post.image_count = imgs.length;
        if (imgs.length > 0) {
            post.image_urls = Array.from(imgs).map(i => i.src);
        }

        // --- SMART TEXT PARSING ---
        // Raw textContent is a dump: Author + timestamp junk + "Shared with Public group" + POST CONTENT + price + location + title + "Message" + UI junk
        // Strategy: strip noise, extract clean description and listing title

        let body = clean;

        // 1. Strip everything before "Shared with Public group" or "Shared with"
        const sharedIdx = body.indexOf('Shared with');
        if (sharedIdx > 0) {
            const afterShared = body.indexOf('group', sharedIdx);
            body = afterShared > 0 ? body.substring(afterShared + 5).trim() : body.substring(sharedIdx + 11).trim();
        }

        // 2. Strip everything after "Message" (FB UI buttons)
        const msgIdx = body.indexOf('Message');
        if (msgIdx > 0) body = body.substring(0, msgIdx).trim();

        // 3. Remove "See more" markers
        body = body.replace(/See more/g, '').trim();

        // 4. Remove "Like", "Comment", "Share" buttons text
        body = body.replace(/\\b(Like|Comment|Share|Reply|Send)\\b/g, '').trim();

        // 5. Split remaining text: description is before the price, listing title is after location
        let description = body;
        let listing_title = '';

        if (post.price) {
            const priceIdx = body.indexOf(post.price);
            if (priceIdx > 0) {
                description = body.substring(0, priceIdx).trim();
                // After price+location comes the listing title
                let afterPrice = body.substring(priceIdx + post.price.length).trim();
                // Remove location prefix if present
                if (post.location) {
                    const locIdx = afterPrice.indexOf(post.location);
                    if (locIdx >= 0) {
                        afterPrice = afterPrice.substring(locIdx + post.location.length).trim();
                    }
                }
                // Clean up dots, dashes from separator
                afterPrice = afterPrice.replace(/^[\\s·.\\-]+/, '').trim();
                // Listing title is the next meaningful chunk (before any junk)
                const titleMatch = afterPrice.match(/^([A-Za-z0-9][^\\n]{5,100})/);
                if (titleMatch) listing_title = titleMatch[1].trim();
            }
        }

        // 6. Clean up description: collapse whitespace, trim
        description = description.replace(/\\s{2,}/g, ' ').trim();
        if (description.length > 500) description = description.substring(0, 500) + '...';

        post.description = description;
        if (listing_title) post.listing_title = listing_title;

        // Keep raw text for debugging but don't show in Excel by default
        // post.raw_text = clean.substring(0, 500);

        // Only keep posts that have author OR price (skip UI chrome)
        if (post.author || post.price) {
            window.__fbPosts.push(post);
            n++;
        }
    });

    return { newPosts: n, total: window.__fbPosts.length };
}
"""

SCROLL_JS = """
() => {
    // Facebook uses a custom scroll container, not window scroll
    const feed = document.querySelector('div[role="feed"]');
    let el = feed;
    while (el && el !== document.body) {
        const s = getComputedStyle(el);
        if (s.overflowY === 'auto' || s.overflowY === 'scroll' || s.overflow === 'auto') {
            el.scrollBy(0, 1200);
            return 'container';
        }
        el = el.parentElement;
    }
    // Fallback to window scroll
    window.scrollBy(0, 1200);
    return 'window';
}
"""


class FacebookGroupScraper:
    """Scrape all posts from a Facebook group and export to Excel.

    Supports three browser modes:
      1. CDP -- connect to an existing Chrome via --remote-debugging-port
      2. Camoufox -- best stealth (if installed)
      3. Playwright Chromium -- fallback
    """

    def __init__(self, headless: bool = True, cdp_url: str | None = None) -> None:
        self._headless = headless
        self._cdp_url = cdp_url  # e.g. "http://127.0.0.1:9222" for CDP mode
        self._pw = None  # Playwright instance (kept alive for CDP cleanup)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def scrape_group(
        self,
        url: str,
        cookies: list[dict[str, Any]],
        max_posts: int = 0,
        output_path: Optional[str] = None,
    ) -> list[dict[str, Any]]:
        """Full pipeline: browser setup -> navigate -> scroll+capture -> export.

        Args:
            url: Facebook group URL.
            cookies: List of cookie dicts (must include c_user, xs at minimum).
            max_posts: Stop after collecting this many posts. 0 = unlimited.
            output_path: If provided, auto-export to this Excel path.

        Returns:
            List of post dicts.
        """
        browser, context, page = await self._setup_browser(cookies)
        try:
            posts_json = await self._scroll_and_collect(page, url, max_posts)
            posts = json.loads(posts_json) if isinstance(posts_json, str) else posts_json

            # Deduplicate by author+price+text[:50]
            posts = self._deduplicate(posts)

            logger.info(
                "Extraction complete: %d posts from %s", len(posts), url,
            )
            if output_path:
                self.export_to_excel(posts, output_path)
            return posts
        finally:
            await page.close()
            # Don't close browser if CDP mode (we didn't create it)
            if not self._cdp_url:
                await context.close()
                await browser.close()
            if self._pw:
                await self._pw.stop()
                self._pw = None

    # ------------------------------------------------------------------
    # Browser setup -- 3 modes
    # ------------------------------------------------------------------

    async def _setup_browser(
        self,
        cookies: list[dict[str, Any]],
    ) -> tuple[Any, Any, Any]:
        """Try in order: CDP connect -> Camoufox launch -> Playwright launch.

        Returns (browser, context, page).
        """
        # Mode 1: CDP (connect to existing Chrome)
        if self._cdp_url:
            return await self._connect_cdp(cookies)

        # Mode 2: Camoufox (best stealth)
        if _HAS_CAMOUFOX:
            try:
                return await self._launch_camoufox(cookies)
            except Exception as exc:
                logger.warning("Camoufox launch failed, falling back: %s", exc)

        # Mode 3: Playwright Chromium (fallback)
        try:
            return await self._launch_playwright(cookies)
        except Exception as e:
            if "spawn" in str(e).lower():
                raise RuntimeError(
                    "Cannot launch browser. Either:\n"
                    "1. Start Chrome with: chrome.exe --remote-debugging-port=9222\n"
                    "2. Pass cdp_url='http://127.0.0.1:9222' to FacebookGroupScraper\n"
                    f"Original error: {e}"
                )
            raise

    async def _connect_cdp(
        self,
        cookies: list[dict[str, Any]],
    ) -> tuple[Any, Any, Any]:
        """Connect to existing Chrome via CDP."""
        from playwright.async_api import async_playwright

        self._pw = await async_playwright().start()
        browser = await self._pw.chromium.connect_over_cdp(self._cdp_url)
        context = browser.contexts[0] if browser.contexts else await browser.new_context()

        prepared = self._prepare_cookies(cookies)
        await context.add_cookies(prepared)
        logger.info("CDP: injected %d cookies via %s", len(prepared), self._cdp_url)

        page = await context.new_page()
        return browser, context, page

    async def _launch_camoufox(
        self,
        cookies: list[dict[str, Any]],
    ) -> tuple[Any, Any, Any]:
        """Launch Camoufox with C++-level stealth."""
        from camoufox.async_api import AsyncCamoufox

        profile = DeviceProfile.random()
        cm = AsyncCamoufox(headless=self._headless, geoip=True)
        browser = await cm.__aenter__()
        # Store the context manager so we can close properly
        self._camoufox_cm = cm

        context = await browser.new_context(
            viewport=profile.viewport,
            locale=profile.locale,
            timezone_id=profile.timezone,
            ignore_https_errors=True,
            screen=profile.screen,
            color_scheme="light",
        )

        # Inject cookies before any navigation
        prepared = self._prepare_cookies(cookies)
        await context.add_cookies(prepared)
        logger.info("Camoufox: injected %d cookies (headless=%s)", len(prepared), self._headless)

        page = await context.new_page()
        return browser, context, page

    async def _launch_playwright(
        self,
        cookies: list[dict[str, Any]],
    ) -> tuple[Any, Any, Any]:
        """Fallback: launch Playwright Chromium with anti-detection flags."""
        from playwright.async_api import async_playwright

        profile = DeviceProfile.random()
        self._pw = await async_playwright().start()
        browser = await self._pw.chromium.launch(
            headless=self._headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
                "--disable-infobars",
                f"--window-size={profile.viewport['width']},{profile.viewport['height']}",
            ],
        )
        context = await browser.new_context(
            user_agent=profile.user_agent,
            viewport=profile.viewport,
            locale=profile.locale,
            timezone_id=profile.timezone,
            ignore_https_errors=True,
            screen=profile.screen,
            color_scheme="light",
        )

        # Inject cookies before any navigation
        prepared = self._prepare_cookies(cookies)
        await context.add_cookies(prepared)
        logger.info("Playwright: injected %d cookies (headless=%s)", len(prepared), self._headless)

        page = await context.new_page()

        # Basic stealth patches for Playwright
        await page.add_init_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
        )
        return browser, context, page

    @staticmethod
    def _prepare_cookies(cookies: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Ensure each cookie dict has the required fields for Playwright.

        EditThisCookie uses ``"no_restriction"`` for SameSite but Playwright
        expects ``"None"``, ``"Lax"``, or ``"Strict"``.
        Also maps ``expirationDate`` -> ``expires``.
        """
        SAME_SITE_MAP = {
            "no_restriction": "None",
            "unspecified": "Lax",
            "lax": "Lax",
            "strict": "Strict",
            "none": "None",
        }

        prepared: list[dict[str, Any]] = []
        for c in cookies:
            entry: dict[str, Any] = {
                "name": c["name"],
                "value": c["value"],
                "domain": c.get("domain", ".facebook.com"),
                "path": c.get("path", "/"),
            }
            # Map expirationDate -> expires (EditThisCookie format)
            if "expirationDate" in c:
                entry["expires"] = c["expirationDate"]
            elif "expires" in c:
                entry["expires"] = c["expires"]
            if "httpOnly" in c:
                entry["httpOnly"] = c["httpOnly"]
            entry["secure"] = c.get("secure", True)
            # Map sameSite values (no_restriction -> None)
            raw_ss = str(c.get("sameSite", "None")).lower()
            entry["sameSite"] = SAME_SITE_MAP.get(raw_ss, "None")
            prepared.append(entry)
        return prepared

    # ------------------------------------------------------------------
    # Scroll and collect -- captures posts into JS memory during scrolling
    # ------------------------------------------------------------------

    async def _scroll_and_collect(
        self,
        page: Any,
        url: str,
        max_posts: int,
    ) -> str:
        """Navigate, inject post collector, scroll, return collected posts as JSON.

        Key insight: Facebook virtualizes DOM -- posts disappear when scrolled past.
        We inject JS that captures posts into window.__fbPosts as they appear,
        BEFORE they get recycled out of DOM.

        Uses textContent (NOT innerText -- Facebook CSS makes innerText empty).
        Scrolls the overflow container (NOT window -- Facebook uses custom scroll).
        """
        logger.info("Navigating to %s", url)

        # Navigate and wait for initial content
        await page.goto(url, wait_until="domcontentloaded", timeout=60000)

        # Smart wait: try networkidle, fall back to a fixed wait
        try:
            await page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            await page.wait_for_timeout(8000)

        # Wait for the feed container to appear
        try:
            await page.wait_for_selector('div[role="feed"]', timeout=15000)
        except Exception:
            logger.warning("No feed container found after initial load")

        # Inject persistent post collector using textContent (NOT innerText)
        await page.evaluate(COLLECTOR_INIT_JS)

        # Click any "See more" buttons to expand truncated post text
        async def _click_see_more():
            try:
                see_more_buttons = await page.query_selector_all(
                    'div[role="button"]:has-text("See more"), '
                    'span[role="button"]:has-text("See more")'
                )
                for btn in see_more_buttons[:10]:
                    try:
                        await btn.click(timeout=1000)
                        await page.wait_for_timeout(300)
                    except Exception:
                        pass
            except Exception:
                pass

        # Scroll loop: scroll the overflow container, capture after each scroll
        stale = 0
        for scroll_num in range(1, 100):
            # Try expanding truncated posts
            await _click_see_more()

            # Capture visible posts into JS memory
            result = await page.evaluate(CAPTURE_JS)
            new_posts = result.get("newPosts", 0)
            total = result.get("total", 0)

            if new_posts == 0:
                stale += 1
            else:
                stale = 0

            logger.info(
                "Scroll %d: +%d new, %d total, stale=%d",
                scroll_num, new_posts, total, stale,
            )

            if stale >= 4:
                logger.info("Stopping: %d consecutive stale scrolls", stale)
                break
            if max_posts > 0 and total >= max_posts:
                logger.info("Reached max_posts limit (%d), stopping scroll", max_posts)
                break

            # Scroll the overflow container (NOT window)
            scroll_target = await page.evaluate(SCROLL_JS)
            logger.debug("Scrolled via: %s", scroll_target)
            await page.wait_for_timeout(2500)

        # Final capture pass
        await page.evaluate(CAPTURE_JS)

        # Get all collected posts from JS memory
        posts_json = await page.evaluate("JSON.stringify(window.__fbPosts)")
        total_collected = await page.evaluate("window.__fbPosts.length")
        logger.info("Scrolling complete -- %d posts collected from JS memory", total_collected)

        return posts_json  # Return JSON string, not HTML

    # ------------------------------------------------------------------
    # Deduplication
    # ------------------------------------------------------------------

    @staticmethod
    def _deduplicate(posts: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Deduplicate posts by author + price + text[:50]."""
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for post in posts:
            key_parts = [
                post.get("author", post.get("author_name", "")),
                post.get("price", ""),
                post.get("text", "")[:50],
            ]
            key = "|".join(str(p) for p in key_parts)
            if key in seen:
                continue
            seen.add(key)
            unique.append(post)
        return unique

    # ------------------------------------------------------------------
    # Post extraction (for non-browser / API fallback)
    # ------------------------------------------------------------------

    def extract_posts(self, html: str, url: str) -> list[dict[str, Any]]:
        """Extract structured post data from the group page HTML.

        Strategy:
        1. Try embedded JSON (ScheduledServerJS / Relay) for rich data.
        2. Fall back to DOM-based extraction from div[role="article"].
        3. Deduplicate by post_id.

        NOTE: This is the fallback path for when we have raw HTML (e.g. from
        an API response). The primary path uses JS-injected collection via
        _scroll_and_collect which handles DOM virtualization.
        """
        posts: list[dict[str, Any]] = []

        # Method 1: try embedded JSON
        json_posts = self._extract_from_json(html, url)
        if json_posts:
            posts.extend(json_posts)
            logger.info("Extracted %d posts from embedded JSON", len(json_posts))

        # Method 2: DOM-based extraction
        dom_posts = self._extract_from_dom(html, url)
        if dom_posts:
            logger.info("Extracted %d posts from DOM", len(dom_posts))
            posts.extend(dom_posts)

        # Deduplicate by post_id
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for post in posts:
            pid = post.get("post_id", "")
            if pid and pid in seen:
                continue
            if pid:
                seen.add(pid)
            unique.append(post)

        logger.info("Total unique posts: %d", len(unique))
        return unique

    def _extract_from_json(
        self, html: str, url: str,
    ) -> list[dict[str, Any]]:
        """Try to extract posts from embedded ScheduledServerJS / Relay data."""
        posts: list[dict[str, Any]] = []

        # Look for ScheduledServerJS.handle payloads
        pattern = re.compile(
            r'(?:ScheduledServerJS|ServerJS)\.handle(?:WithCustomApplyEach)?\((\{.+?\})\);',
            re.DOTALL,
        )
        for match in pattern.finditer(html):
            try:
                data = json.loads(match.group(1))
            except json.JSONDecodeError:
                continue

            # Search for story/post nodes in the JSON tree
            stories = self._find_stories_in_json(data)
            for story in stories:
                post = self._parse_json_story(story, url)
                if post:
                    posts.append(post)

        # Also check RelayPrefetchedStreamCache
        relay_pattern = re.compile(
            r'__d\("RelayPrefetchedStreamCache"\).*?(\{.+?\})\s*\)',
            re.DOTALL,
        )
        relay_match = relay_pattern.search(html)
        if relay_match:
            try:
                relay_data = json.loads(relay_match.group(1))
                stories = self._find_stories_in_json(relay_data)
                for story in stories:
                    post = self._parse_json_story(story, url)
                    if post:
                        posts.append(post)
            except json.JSONDecodeError:
                pass

        return posts

    def _find_stories_in_json(self, data: Any, depth: int = 0) -> list[dict]:
        """Recursively locate story/post objects in Facebook's JSON blobs."""
        if depth > 20:
            return []
        results: list[dict] = []

        if isinstance(data, dict):
            # A node that looks like a group feed story
            if data.get("__typename") in (
                "Story", "GroupFeedStory", "GroupCommerceProductItem",
            ):
                results.append(data)
                return results

            # Check for creation_story or node.comet_sections
            if "creation_story" in data or "comet_sections" in data:
                results.append(data)
                return results

            for value in data.values():
                results.extend(self._find_stories_in_json(value, depth + 1))

        elif isinstance(data, list):
            for item in data:
                results.extend(self._find_stories_in_json(item, depth + 1))

        return results

    def _parse_json_story(
        self, story: dict, group_url: str,
    ) -> Optional[dict[str, Any]]:
        """Convert a JSON story node into a flat post dict."""
        post: dict[str, Any] = {}

        # Post ID
        post_id = (
            story.get("post_id")
            or story.get("id")
            or deep_get(story, "creation_story", "id")
            or deep_get(story, "node", "id")
        )
        if post_id:
            post["post_id"] = str(post_id)

        # Author
        actor = (
            story.get("actors", [{}])[0]
            if isinstance(story.get("actors"), list) and story.get("actors")
            else deep_get(story, "creation_story", "comet_sections", "actor_photo", "story", "actors", "0")
            or find_key_recursive(story, "actor")
            or {}
        )
        if isinstance(actor, dict):
            post["author_name"] = actor.get("name")
            actor_url = actor.get("url") or actor.get("uri")
            if actor_url:
                post["author_profile_url"] = actor_url

        # Text content
        message = (
            deep_get(story, "message", "text")
            or find_key_recursive(story, "text")
        )
        if message and isinstance(message, str):
            post["text"] = message

        # Timestamp
        created_time = (
            story.get("created_time")
            or deep_get(story, "creation_story", "creation_time")
            or find_key_recursive(story, "creation_time")
        )
        if created_time:
            post["timestamp"] = parse_timestamp(created_time)

        # Engagement
        feedback = find_key_recursive(story, "feedback") or {}
        if isinstance(feedback, dict):
            reaction_count = deep_get(feedback, "reaction_count", "count")
            if reaction_count is not None:
                post["like_count"] = reaction_count
            comment_count = deep_get(feedback, "comment_count", "total_count")
            if comment_count is not None:
                post["comment_count"] = comment_count
            share_count = deep_get(feedback, "share_count", "count")
            if share_count is not None:
                post["share_count"] = share_count

        # Images
        attachments = find_key_recursive(story, "attachments") or []
        image_urls: list[str] = []
        if isinstance(attachments, list):
            for att in attachments:
                if isinstance(att, dict):
                    media = deep_get(att, "media", "image", "uri")
                    if media:
                        image_urls.append(media)
                    photo_uri = deep_get(att, "media", "photo", "image", "uri")
                    if photo_uri:
                        image_urls.append(photo_uri)

        if image_urls:
            post["image_urls"] = image_urls
            post["image_count"] = len(image_urls)

        # Price / listing detection
        text = post.get("text", "")
        self._extract_price_info(post, text)
        self._extract_location(post, text)
        self._extract_condition(post, text)

        # Post type
        post["post_type"] = self._detect_post_type(post)

        # Post URL
        post_url = story.get("url") or story.get("uri")
        if post_url:
            post["post_url"] = post_url

        # Filter out empty values
        post = {k: v for k, v in post.items() if v is not None}
        return post if post.get("post_id") or post.get("text") else None

    def _extract_from_dom(
        self, html: str, url: str,
    ) -> list[dict[str, Any]]:
        """Extract posts from DOM using div[role="article"] elements."""
        soup = BeautifulSoup(html, "html.parser")
        articles = soup.find_all("div", attrs={"role": "article"})
        posts: list[dict[str, Any]] = []

        for article in articles:
            post = self._parse_article_element(article, url)
            if post:
                posts.append(post)

        return posts

    def _parse_article_element(
        self, article: Tag, group_url: str,
    ) -> Optional[dict[str, Any]]:
        """Parse a single div[role="article"] into a post dict."""
        post: dict[str, Any] = {}

        # Skip loading placeholders
        aria_label = article.get("aria-label", "")
        if aria_label and "loading" in aria_label.lower():
            return None

        # Post ID -- from data attributes or permalink
        post_id = self._extract_post_id_from_element(article)
        if post_id:
            post["post_id"] = post_id

        # Author name -- typically in strong, h4, or aria-label
        author_el = (
            article.find("strong")
            or article.find("h4")
        )
        if author_el:
            author_link = author_el.find("a") if author_el else None
            post["author_name"] = author_el.get_text(strip=True)
            if author_link and author_link.get("href"):
                href = author_link["href"]
                if href.startswith("/"):
                    href = "https://www.facebook.com" + href
                post["author_profile_url"] = href.split("?")[0]

        # If no author from strong/h4, try aria-label on the article
        if not post.get("author_name"):
            aria = article.get("aria-label", "")
            if aria and "loading" not in aria.lower():
                post["author_name"] = aria

        # Timestamp -- from <abbr> or datetime attributes or relative text
        timestamp_str = self._extract_timestamp_from_element(article)
        if timestamp_str:
            post["timestamp"] = timestamp_str

        # Post URL -- from timestamp/permalink link
        post_url = self._extract_post_url_from_element(article, group_url)
        if post_url:
            post["post_url"] = post_url
            # Try to extract post_id from the URL if not found elsewhere
            if not post.get("post_id"):
                url_id = self._extract_post_id_from_url(post_url)
                if url_id:
                    post["post_id"] = url_id

        # Text content -- main post body
        text = self._extract_text_from_element(article)
        if text:
            post["text"] = text

        # Images
        image_urls = self._extract_images_from_element(article)
        if image_urls:
            post["image_urls"] = image_urls
            post["image_count"] = len(image_urls)

        # Engagement counts
        self._extract_engagement_from_element(article, post)

        # Price / currency
        full_text = post.get("text", "")
        self._extract_price_info(post, full_text)

        # Location
        self._extract_location(post, full_text)

        # Condition
        self._extract_condition(post, full_text)

        # Post type
        post["post_type"] = self._detect_post_type(post)

        # Filter empty values
        post = {k: v for k, v in post.items() if v is not None}
        return post if post.get("post_id") or post.get("text") else None

    # ------------------------------------------------------------------
    # Element-level extraction helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_post_id_from_element(article: Tag) -> Optional[str]:
        """Try to find a post ID from data attributes or link patterns."""
        # data-ft attribute (older FB markup)
        data_ft = article.get("data-ft")
        if data_ft:
            try:
                ft = json.loads(data_ft) if isinstance(data_ft, str) else data_ft
                top_level_post_id = ft.get("top_level_post_id") or ft.get("tl_objid")
                if top_level_post_id:
                    return str(top_level_post_id)
            except (json.JSONDecodeError, TypeError):
                pass

        # aria-describedby or id attribute
        el_id = article.get("id", "")
        if el_id:
            id_match = re.search(r"(\d{10,})", el_id)
            if id_match:
                return id_match.group(1)

        # Search for permalink with post ID
        for link in article.find_all("a", href=True):
            href = link["href"]
            pid = _extract_id_from_href(href)
            if pid:
                return pid

        return None

    @staticmethod
    def _extract_post_id_from_url(url: str) -> Optional[str]:
        """Extract a numeric post ID from a Facebook URL."""
        return _extract_id_from_href(url)

    @staticmethod
    def _extract_timestamp_from_element(article: Tag) -> Optional[str]:
        """Extract timestamp from <abbr>, datetime attributes, or relative text."""
        # <abbr> with data-utime (older FB)
        abbr = article.find("abbr")
        if abbr:
            utime = abbr.get("data-utime")
            if utime:
                return parse_timestamp(int(utime))
            title = abbr.get("title")
            if title:
                return title

        # Elements with datetime attribute
        time_el = article.find("time")
        if time_el:
            dt = time_el.get("datetime")
            if dt:
                return dt

        # Look for relative time text in links (e.g. "2h", "3d", "Yesterday")
        for link in article.find_all("a", href=True):
            href = link.get("href", "")
            if "/posts/" in href or "/permalink/" in href or "story_fbid" in href:
                link_text = link.get_text(strip=True)
                if link_text and _is_relative_time(link_text):
                    return _parse_relative_time(link_text)

        # aria-label on timestamp spans
        for span in article.find_all("span"):
            aria = span.get("aria-label", "")
            if aria and _is_relative_time(aria):
                return _parse_relative_time(aria)

        return None

    @staticmethod
    def _extract_post_url_from_element(
        article: Tag, group_url: str,
    ) -> Optional[str]:
        """Extract the post permalink URL."""
        for link in article.find_all("a", href=True):
            href = link["href"]
            if any(p in href for p in ("/posts/", "/permalink/", "story_fbid")):
                if href.startswith("/"):
                    href = "https://www.facebook.com" + href
                return href.split("?")[0]
        return None

    @staticmethod
    def _extract_text_from_element(article: Tag) -> Optional[str]:
        """Extract the main post text content."""
        # Try data-ad-preview="message" (common in modern FB)
        text_div = article.find("div", attrs={"data-ad-preview": "message"})
        if text_div:
            return text_div.get_text(separator="\n", strip=True)

        # Try div[dir="auto"] which FB often uses for post text
        dir_auto_divs = article.find_all("div", attrs={"dir": "auto"})
        texts: list[str] = []
        for div in dir_auto_divs:
            t = div.get_text(strip=True)
            # Skip very short fragments that are likely UI labels
            if t and len(t) > 5:
                texts.append(t)
        if texts:
            return "\n".join(texts)

        return None

    @staticmethod
    def _extract_images_from_element(article: Tag) -> list[str]:
        """Extract image URLs from within a post article."""
        urls: list[str] = []
        for img in article.find_all("img"):
            src = img.get("src", "")
            # Skip tiny icons, emoji, and UI images
            if not src:
                continue
            if any(skip in src for skip in (
                "emoji", "rsrc.php", "static", "pixel", "tracking",
                "1x1", "blank.gif",
            )):
                continue
            # Prefer scontent images (actual Facebook content images)
            if "scontent" in src:
                urls.append(src)
                continue
            # Only include other images that look like content
            width = img.get("width")
            height = img.get("height")
            if width and height:
                try:
                    if int(width) < 50 or int(height) < 50:
                        continue
                except (ValueError, TypeError):
                    pass
            urls.append(src)
        return urls

    @staticmethod
    def _extract_engagement_from_element(
        article: Tag, post: dict[str, Any],
    ) -> None:
        """Extract like, comment, and share counts from the article element."""
        # aria-labels often contain counts like "42 reactions" or "5 comments"
        for span in article.find_all(["span", "div"]):
            aria = span.get("aria-label", "")
            if not aria:
                continue

            # Reactions / likes
            like_match = re.search(
                r"(\d[\d,\.]*[KMB]?)\s*(?:reactions?|likes?|people reacted)",
                aria, re.IGNORECASE,
            )
            if like_match and "like_count" not in post:
                post["like_count"] = parse_count(like_match.group(1))

            # Comments
            comment_match = re.search(
                r"(\d[\d,\.]*[KMB]?)\s*comments?", aria, re.IGNORECASE,
            )
            if comment_match and "comment_count" not in post:
                post["comment_count"] = parse_count(comment_match.group(1))

            # Shares
            share_match = re.search(
                r"(\d[\d,\.]*[KMB]?)\s*shares?", aria, re.IGNORECASE,
            )
            if share_match and "share_count" not in post:
                post["share_count"] = parse_count(share_match.group(1))

        # Also check visible text for count patterns like "42" near reaction icons
        text_spans = article.find_all("span", class_=True)
        for span in text_spans:
            text = span.get_text(strip=True)
            # Pattern: "X comments" or "X shares" in visible text
            cm = re.match(r"^(\d[\d,]*)\s+comments?$", text, re.IGNORECASE)
            if cm and "comment_count" not in post:
                post["comment_count"] = parse_count(cm.group(1))
            sm = re.match(r"^(\d[\d,]*)\s+shares?$", text, re.IGNORECASE)
            if sm and "share_count" not in post:
                post["share_count"] = parse_count(sm.group(1))

    # ------------------------------------------------------------------
    # Content analysis helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_price_info(post: dict[str, Any], text: str) -> None:
        """Extract price and currency from text content."""
        match = _PRICE_PATTERN.search(text)
        if match:
            price_str = match.group("price") or match.group("price2")
            symbol = match.group("currency") or match.group("currency2")
            if price_str:
                post["price"] = price_str.replace(",", "")
            if symbol:
                post["currency"] = _CURRENCY_SYMBOLS.get(symbol, symbol)

    @staticmethod
    def _extract_location(post: dict[str, Any], text: str) -> None:
        """Extract location from post text (e.g., all-caps city names)."""
        # Common pattern: "CITY, COUNTY" or "City, State" in listing posts
        loc_match = re.search(
            r"\b([A-Z][A-Za-z\s]+,\s*[A-Z][A-Za-z\s]+)\b", text,
        )
        if loc_match:
            location = loc_match.group(1).strip()
            # Avoid false positives from normal sentences
            if len(location) < 60:
                post["location"] = location

    @staticmethod
    def _extract_condition(post: dict[str, Any], text: str) -> None:
        """Extract item condition from text patterns."""
        match = _CONDITION_PATTERNS.search(text)
        if match:
            post["condition"] = match.group(1).strip().lower()

    @staticmethod
    def _detect_post_type(post: dict[str, Any]) -> str:
        """Auto-detect the post type based on content signals."""
        if post.get("price"):
            return "sale_listing"

        text = post.get("text", "")
        images = post.get("image_urls", [])

        # Check for video URLs in the text or attachments
        if re.search(r"(?:video|\.mp4|fb\.watch)", text, re.IGNORECASE):
            return "video"

        # Has images but minimal text
        if images and len(text) < 20:
            return "photo"

        # External link
        if re.search(r"https?://(?!facebook\.com|fb\.com)", text):
            return "link"

        return "discussion"

    # ------------------------------------------------------------------
    # Relative time parsing
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_relative_time(text: str) -> str:
        """Convert relative time text to ISO datetime string.

        Handles: "2h", "3d", "Yesterday", "Just now", "5m", "1w", etc.
        """
        return _parse_relative_time(text)

    # ------------------------------------------------------------------
    # Dynamic columns
    # ------------------------------------------------------------------

    @staticmethod
    def build_dynamic_columns(posts: list[dict[str, Any]]) -> list[str]:
        """Build column list from the union of all post keys.

        Columns with priority ordering come first; remaining keys are
        sorted alphabetically. Only includes columns that have at least
        one non-empty value across all posts.
        """
        # Gather all keys that have at least one non-empty value
        active_keys: set[str] = set()
        for post in posts:
            for key, value in post.items():
                if value is not None and value != "" and value != []:
                    active_keys.add(key)

        # Build ordered column list
        columns: list[str] = []
        for col in _COLUMN_PRIORITY:
            if col in active_keys:
                columns.append(col)
                active_keys.discard(col)

        # Remaining keys sorted alphabetically
        columns.extend(sorted(active_keys))
        return columns

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    @staticmethod
    def export_to_excel(
        posts: list[dict[str, Any]],
        output_path: str,
    ) -> str:
        """Export posts to an Excel file with formatted headers and auto-width.

        Args:
            posts: List of post dicts.
            output_path: Destination .xlsx path.

        Returns:
            The output path.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Group Posts"

        if not posts:
            wb.save(output_path)
            logger.info("Exported empty workbook to %s", output_path)
            return output_path

        columns = FacebookGroupScraper.build_dynamic_columns(posts)

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid",
        )

        # Write headers
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        # Freeze header row
        ws.freeze_panes = "A2"

        # Write data rows
        for row_idx, post in enumerate(posts, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                value = post.get(col_name, "")
                # Join lists with separator
                if isinstance(value, list):
                    value = " | ".join(str(v) for v in value)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns (capped at 60)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = len(col_name)
            for row_idx in range(2, len(posts) + 2):
                cell_value = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(cell_value))
            adjusted = min(max_len + 2, 60)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

        wb.save(output_path)
        logger.info("Exported %d posts to %s", len(posts), output_path)
        return output_path


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _extract_id_from_href(href: str) -> Optional[str]:
    """Extract a numeric post ID from a Facebook URL/href."""
    # /posts/123456
    match = re.search(r"/posts/(\d+)", href)
    if match:
        return match.group(1)
    # /permalink/123456
    match = re.search(r"/permalink/(\d+)", href)
    if match:
        return match.group(1)
    # story_fbid=123456
    match = re.search(r"story_fbid=(\d+)", href)
    if match:
        return match.group(1)
    # pfbid pattern
    match = re.search(r"pfbid\w+", href)
    if match:
        return match.group(0)
    return None


def _is_relative_time(text: str) -> bool:
    """Check if text looks like a relative time string."""
    text = text.strip().lower()
    if text in ("just now", "now", "yesterday"):
        return True
    if re.match(r"^\d+\s*[smhdw]$", text):
        return True
    if re.match(r"^\d+\s*(min|mins|minute|minutes|hour|hours|day|days|week|weeks)\s*ago$", text):
        return True
    return False


def _parse_relative_time(text: str) -> str:
    """Convert relative time text to ISO datetime string."""
    now = datetime.now(tz=timezone.utc)
    text = text.strip().lower()

    if text in ("just now", "now"):
        return now.isoformat()

    if text == "yesterday":
        return (now - timedelta(days=1)).isoformat()

    # "2h", "5m", "3d", "1w"
    short_match = re.match(r"^(\d+)\s*([smhdw])$", text)
    if short_match:
        amount = int(short_match.group(1))
        unit = short_match.group(2)
        deltas = {"s": timedelta(seconds=amount), "m": timedelta(minutes=amount),
                  "h": timedelta(hours=amount), "d": timedelta(days=amount),
                  "w": timedelta(weeks=amount)}
        return (now - deltas.get(unit, timedelta())).isoformat()

    # "5 minutes ago", "2 hours ago", etc.
    long_match = re.match(
        r"^(\d+)\s*(min|mins|minute|minutes|hour|hours|day|days|week|weeks)\s*ago$",
        text,
    )
    if long_match:
        amount = int(long_match.group(1))
        unit = long_match.group(2)
        if "min" in unit:
            delta = timedelta(minutes=amount)
        elif "hour" in unit:
            delta = timedelta(hours=amount)
        elif "day" in unit:
            delta = timedelta(days=amount)
        elif "week" in unit:
            delta = timedelta(weeks=amount)
        else:
            delta = timedelta()
        return (now - delta).isoformat()

    # Fallback: return as-is
    return text
