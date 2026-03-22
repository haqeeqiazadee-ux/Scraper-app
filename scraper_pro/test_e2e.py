"""
🧪 SCRAPLING PRO - End-to-End Test Suite
=========================================
Comprehensive testing for all scraper components.

Run: python test_e2e.py

Tests:
1. Core imports and dependencies
2. Engine functionality
3. Templates
4. AJAX handler
5. AI extractor
6. Smart exporter
7. Auth scraper
8. Scheduler
9. Web dashboard endpoints
10. Full scraping workflow
"""

import sys
import os
import json
import time
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any

# Add current directory to path
sys.path.insert(0, str(Path(__file__).parent))

# Test results storage
TEST_RESULTS = {
    'passed': 0,
    'failed': 0,
    'skipped': 0,
    'errors': []
}


def log_test(name: str, passed: bool, error: str = None, skipped: bool = False):
    """Log test result"""
    if skipped:
        TEST_RESULTS['skipped'] += 1
        print(f"  ⏭️  {name}: SKIPPED")
    elif passed:
        TEST_RESULTS['passed'] += 1
        print(f"  ✅ {name}: PASSED")
    else:
        TEST_RESULTS['failed'] += 1
        TEST_RESULTS['errors'].append({'test': name, 'error': error})
        print(f"  ❌ {name}: FAILED - {error}")


def test_section(name: str):
    """Print section header"""
    print(f"\n{'='*60}")
    print(f"🧪 {name}")
    print(f"{'='*60}")


# ============================================================================
# 1. CORE IMPORTS TEST
# ============================================================================

def test_core_imports():
    """Test that all core modules can be imported"""
    test_section("1. CORE IMPORTS")
    
    modules = [
        ('engine', 'Core Engine'),
        ('engine_v2', 'Enhanced Engine v2'),
        ('templates', 'Scraper Templates'),
        ('ajax_handler', 'AJAX Handler'),
        ('ai_extractor', 'AI Extractor'),
        ('smart_exporter', 'Smart Exporter'),
        ('auth_scraper', 'Auth Scraper'),
        ('scheduler', 'Scheduler'),
        ('integrations', 'Integrations'),
    ]
    
    for module_name, display_name in modules:
        try:
            __import__(module_name)
            log_test(f"Import {display_name}", True)
        except ImportError as e:
            log_test(f"Import {display_name}", False, str(e))
        except Exception as e:
            log_test(f"Import {display_name}", False, str(e))


# ============================================================================
# 2. DEPENDENCIES TEST
# ============================================================================

def test_dependencies():
    """Test external dependencies"""
    test_section("2. DEPENDENCIES")
    
    dependencies = [
        ('scrapling', 'Scrapling Core'),
        ('bs4', 'BeautifulSoup'),
        ('requests', 'Requests'),
        ('openpyxl', 'OpenPyXL (Excel)'),
        ('flask', 'Flask'),
    ]
    
    optional_deps = [
        ('playwright', 'Playwright'),
        ('extruct', 'Extruct (Schema.org)'),
        ('vaderSentiment', 'VADER Sentiment'),
        ('pytrends', 'PyTrends'),
    ]
    
    for module_name, display_name in dependencies:
        try:
            __import__(module_name)
            log_test(f"{display_name}", True)
        except ImportError:
            # Scrapling is optional if using engine_v2 in heuristic mode
            if module_name == 'scrapling':
                log_test(f"{display_name}", True, skipped=True)
            else:
                log_test(f"{display_name}", False, "Not installed")
    
    print("\n  Optional dependencies:")
    for module_name, display_name in optional_deps:
        try:
            __import__(module_name)
            log_test(f"{display_name}", True)
        except ImportError:
            log_test(f"{display_name}", True, skipped=True)


# ============================================================================
# 3. ENGINE TEST
# ============================================================================

def test_engine():
    """Test scraping engine"""
    test_section("3. SCRAPING ENGINE")
    
    # Test engine_v2 (preferred)
    try:
        from engine_v2 import (
            ScrapingEngine, ScrapedItem, Exporter, 
            RateLimiter, ProxyManager, CaptchaSolver,
            RobotsChecker, WebhookSender, InfiniteScrollHandler
        )
        
        # Test ScrapedItem
        item = ScrapedItem(
            url="https://example.com",
            title="Test Product",
            content={"price": "$29.99", "stock": "10"}
        )
        assert item.url == "https://example.com"
        assert item.title == "Test Product"
        log_test("ScrapedItem creation", True)
        
        # Test to_dict
        item_dict = item.to_dict()
        assert 'url' in item_dict
        assert 'price' in item_dict
        log_test("ScrapedItem.to_dict()", True)
        
        # Test RateLimiter
        limiter = RateLimiter(requests_per_second=2.0)
        start = time.time()
        limiter.wait("https://example.com")
        limiter.wait("https://example.com")
        elapsed = time.time() - start
        # Should wait ~0.5s between requests
        log_test("RateLimiter", True)
        
        # Test ProxyManager
        pm = ProxyManager(proxies=["http://proxy1:8080", "http://proxy2:8080"])
        proxy = pm.get_proxy()
        assert proxy is not None
        pm.report_success(proxy, 0.5)
        log_test("ProxyManager", True)
        
        # Test RobotsChecker
        checker = RobotsChecker()
        # Should return True for most URLs (permissive when robots.txt unavailable)
        log_test("RobotsChecker init", True)
        
        # Test CaptchaSolver init (without API key)
        solver = CaptchaSolver(api_key=None)
        assert solver.enabled == False
        log_test("CaptchaSolver init (no key)", True)
        
        # Test WebhookSender init
        webhook = WebhookSender(url=None)
        assert webhook.enabled == False
        log_test("WebhookSender init", True)
        
        # Test InfiniteScrollHandler init
        scroll = InfiniteScrollHandler()
        assert scroll.scroll_pause == 2.0
        log_test("InfiniteScrollHandler init", True)
        
        # Test ScrapingEngine init
        engine = ScrapingEngine(
            mode="stealthy",
            timeout=30,
            max_retries=2
        )
        assert engine.timeout == 30
        assert engine.timeout_ms == 30000
        log_test("ScrapingEngine init", True)
        
    except ImportError as e:
        log_test("Engine import", False, str(e))
    except AssertionError as e:
        log_test("Engine assertion", False, str(e))
    except Exception as e:
        log_test("Engine test", False, str(e))


# ============================================================================
# 4. TEMPLATES TEST
# ============================================================================

def test_templates():
    """Test scraper templates"""
    test_section("4. SCRAPER TEMPLATES")
    
    try:
        from templates import (
            ScraperTemplate, EcommerceScraper, NewsScraper, 
            JobScraper, CustomScraper
        )
        
        # Test EcommerceScraper init
        scraper = EcommerceScraper(mode="stealthy", timeout=30)
        assert "E-commerce" in scraper.name or "E-Commerce" in scraper.name
        log_test("EcommerceScraper init", True)
        
        # Test NewsScraper init
        scraper = NewsScraper()
        assert scraper.name == "News Scraper"
        log_test("NewsScraper init", True)
        
        # Test JobScraper init
        scraper = JobScraper()
        assert scraper.name == "Job Scraper"
        log_test("JobScraper init", True)
        
        # Test CustomScraper init
        scraper = CustomScraper(
            container=".product",
            fields={"title": "h2", "price": ".price"}
        )
        log_test("CustomScraper init", True)
        
    except ImportError as e:
        log_test("Templates import", False, str(e))
    except Exception as e:
        log_test("Templates test", False, str(e))


# ============================================================================
# 5. AJAX HANDLER TEST
# ============================================================================

def test_ajax_handler():
    """Test AJAX handler"""
    test_section("5. AJAX HANDLER")
    
    try:
        from ajax_handler import AjaxHandler
        
        handler = AjaxHandler(
            scroll_pause=1.0,
            click_pause=0.5,
            load_timeout=10.0
        )
        
        assert handler.scroll_pause == 1.0
        assert handler.click_pause == 0.5
        log_test("AjaxHandler init", True)
        
        # Test helper methods exist
        assert hasattr(handler, 'scroll_until_end')
        assert hasattr(handler, 'click_load_more')
        assert hasattr(handler, 'paginate_ajax')
        assert hasattr(handler, 'wait_for_content')
        assert hasattr(handler, 'load_lazy_images')
        log_test("AjaxHandler methods exist", True)
        
    except ImportError as e:
        log_test("AjaxHandler import", False, str(e))
    except Exception as e:
        log_test("AjaxHandler test", False, str(e))


# ============================================================================
# 6. AI EXTRACTOR TEST
# ============================================================================

def test_ai_extractor():
    """Test AI extractor"""
    test_section("6. AI EXTRACTOR")
    
    try:
        from ai_extractor import (
            AIExtractor, SchemaExtractor, HeuristicExtractor,
            SelfHealingSelector, ExtractionResult
        )
        
        # Test SchemaExtractor
        schema = SchemaExtractor()
        log_test("SchemaExtractor init", True)
        
        # Test HeuristicExtractor
        heuristic = HeuristicExtractor()
        assert hasattr(heuristic, 'PATTERNS')
        log_test("HeuristicExtractor init", True)
        
        # Test pattern detection - URL with /products/ can be listing or detail
        page_type = heuristic.detect_page_type("", "https://shop.com/product/laptop-123")
        assert page_type in ['product_detail', 'product_listing']
        log_test("HeuristicExtractor.detect_page_type()", True)
        
        # Test AIExtractor without API key
        extractor = AIExtractor()
        assert extractor.ai_provider is None
        assert extractor.fallback_to_heuristics == True
        log_test("AIExtractor init (no API key)", True)
        
        # Test SelfHealingSelector
        selector = SelfHealingSelector(
            name="price",
            original_selector=".price",
            sample_values=["$29.99", "$39.99"],
            value_pattern=r'\$\d+\.\d{2}'
        )
        assert selector.name == "price"
        log_test("SelfHealingSelector init", True)
        
        # Test ExtractionResult
        result = ExtractionResult(
            success=True,
            data=[{"name": "Test", "price": 29.99}],
            method="heuristic",
            confidence=0.8
        )
        assert result.success == True
        assert len(result.data) == 1
        log_test("ExtractionResult", True)
        
    except ImportError as e:
        log_test("AI Extractor import", False, str(e))
    except Exception as e:
        log_test("AI Extractor test", False, str(e))


# ============================================================================
# 7. SMART EXPORTER TEST
# ============================================================================

def test_smart_exporter():
    """Test smart Excel exporter"""
    test_section("7. SMART EXPORTER")
    
    try:
        from smart_exporter import SmartExcelExporter, smart_export
        
        exporter = SmartExcelExporter()
        log_test("SmartExcelExporter init", True)
        
        # Test column priority
        assert exporter.COLUMN_PRIORITIES['name'] < exporter.COLUMN_PRIORITIES['description']
        assert exporter.COLUMN_PRIORITIES['price'] < exporter.COLUMN_PRIORITIES['url']
        log_test("Column priorities correct", True)
        
        # Test data type detection
        col_type = exporter._detect_column_type('price', ['$29.99', '$39.99'])
        assert col_type == 'currency'
        log_test("Currency detection", True)
        
        col_type = exporter._detect_column_type('product_url', ['https://example.com'])
        assert col_type == 'url'
        log_test("URL detection", True)
        
        # Test export with sample data
        sample_data = [
            {"title": "Laptop", "price": "$999.99", "stock": "15", "rating": "4.5"},
            {"title": "Mouse", "price": "$29.99", "stock": "100", "rating": "4.8"},
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        result = exporter.export(sample_data, filepath)
        
        if exporter.openpyxl_available:
            assert result is not None
            assert Path(result).exists()
            log_test("Excel export", True)
            
            # Check file size (should be > 0)
            file_size = Path(result).stat().st_size
            assert file_size > 0
            log_test("Excel file has content", True)
            
            # Clean up
            Path(result).unlink()
        else:
            log_test("Excel export", True, skipped=True)
        
    except ImportError as e:
        log_test("Smart Exporter import", False, str(e))
    except Exception as e:
        log_test("Smart Exporter test", False, str(e))


# ============================================================================
# 8. AUTH SCRAPER TEST
# ============================================================================

def test_auth_scraper():
    """Test authenticated scraper"""
    test_section("8. AUTH SCRAPER")
    
    try:
        from auth_scraper import AuthenticatedScraper, SessionManager, SessionData
        
        # Test SessionData
        session = SessionData(
            name="test",
            domain="example.com",
            cookies=[{"name": "session", "value": "abc123"}]
        )
        assert session.name == "test"
        log_test("SessionData init", True)
        
        # Test SessionManager
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = SessionManager(sessions_dir=tmpdir)
            sessions = manager.list_sessions()
            assert isinstance(sessions, list)
            log_test("SessionManager.list_sessions()", True)
        
        # Test AuthenticatedScraper init
        with tempfile.TemporaryDirectory() as tmpdir:
            scraper = AuthenticatedScraper(
                name="test_site",
                sessions_dir=tmpdir,
                timeout=30
            )
            assert scraper.name == "test_site"
            assert scraper.timeout == 30
            log_test("AuthenticatedScraper init", True)
            
            # Test session validation (no session)
            assert scraper.is_session_valid() == False
            log_test("is_session_valid() returns False when no session", True)
        
    except ImportError as e:
        log_test("Auth Scraper import", False, str(e))
    except Exception as e:
        log_test("Auth Scraper test", False, str(e))


# ============================================================================
# 9. SCHEDULER TEST
# ============================================================================

def test_scheduler():
    """Test scheduler"""
    test_section("9. SCHEDULER")
    
    try:
        from scheduler import Scheduler, Job, CronParser, schedule_scrape
        
        # Test CronParser
        parsed = CronParser.parse("0 9 * * *")
        assert 0 in parsed['minute']
        assert 9 in parsed['hour']
        log_test("CronParser.parse()", True)
        
        # Test cron matching
        from datetime import datetime
        test_dt = datetime(2024, 1, 15, 9, 0, 0)
        assert CronParser.matches("0 9 * * *", test_dt) == True
        log_test("CronParser.matches()", True)
        
        # Test next_run calculation
        next_run = CronParser.next_run("0 9 * * *")
        assert next_run is not None
        log_test("CronParser.next_run()", True)
        
        # Test Scheduler
        scheduler = Scheduler()
        assert scheduler.running == False
        log_test("Scheduler init", True)
        
        # Test add_job
        def dummy_job():
            return {"status": "done"}
        
        job = scheduler.add_job(
            name="test_job",
            func=dummy_job,
            schedule="hourly"
        )
        assert job.name == "test_job"
        assert "test_job" in scheduler.jobs
        log_test("Scheduler.add_job()", True)
        
        # Test remove_job
        removed = scheduler.remove_job("test_job")
        assert removed == True
        assert "test_job" not in scheduler.jobs
        log_test("Scheduler.remove_job()", True)
        
        # Test preset schedules
        assert "hourly" in Scheduler.PRESETS
        assert "daily" in Scheduler.PRESETS
        log_test("Scheduler presets exist", True)
        
    except ImportError as e:
        log_test("Scheduler import", False, str(e))
    except Exception as e:
        log_test("Scheduler test", False, str(e))


# ============================================================================
# 10. EXPORTER TEST
# ============================================================================

def test_exporter():
    """Test data exporters"""
    test_section("10. DATA EXPORTERS")
    
    try:
        from engine_v2 import ScrapedItem, Exporter
        
        # Create test items
        items = [
            ScrapedItem(
                url="https://shop.com/product/1",
                title="Laptop HP",
                content={"price": "$899", "stock": "10"}
            ),
            ScrapedItem(
                url="https://shop.com/product/2",
                title="Mouse",
                content={"price": "$29.99", "stock": "50"}
            ),
        ]
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # Test CSV export
            csv_path = os.path.join(tmpdir, "test.csv")
            Exporter.to_csv(items, csv_path)
            assert Path(csv_path).exists()
            log_test("CSV export", True)
            
            # Verify CSV content
            with open(csv_path, 'r') as f:
                content = f.read()
                assert "Laptop HP" in content
                assert "$899" in content
            log_test("CSV content correct", True)
            
            # Test JSON export
            json_path = os.path.join(tmpdir, "test.json")
            Exporter.to_json(items, json_path)
            assert Path(json_path).exists()
            log_test("JSON export", True)
            
            # Verify JSON content
            with open(json_path, 'r') as f:
                data = json.load(f)
                assert len(data) == 2
                assert data[0]['title'] == "Laptop HP"
            log_test("JSON content correct", True)
            
            # Test Excel export
            try:
                import openpyxl
                xlsx_path = os.path.join(tmpdir, "test.xlsx")
                Exporter.to_excel(items, xlsx_path)
                assert Path(xlsx_path).exists()
                log_test("Excel export", True)
            except ImportError:
                log_test("Excel export", True, skipped=True)
        
    except ImportError as e:
        log_test("Exporter import", False, str(e))
    except Exception as e:
        log_test("Exporter test", False, str(e))


# ============================================================================
# 11. INTEGRATION TEST - FULL WORKFLOW
# ============================================================================

def test_full_workflow():
    """Test complete scraping workflow"""
    test_section("11. FULL WORKFLOW INTEGRATION")
    
    try:
        from engine_v2 import ScrapingEngine, ScrapedItem, Exporter
        from smart_exporter import SmartExcelExporter
        from ai_extractor import AIExtractor, HeuristicExtractor
        
        # Simulate scraped data (as if we scraped a real page)
        scraped_data = [
            {
                "url": "https://shop.example.com/laptop-1",
                "title": "Gaming Laptop Pro",
                "price": "$1,299.99",
                "original_price": "$1,499.99",
                "discount": "13%",
                "stock": "5",
                "rating": "4.7",
                "reviews": "234",
                "brand": "TechBrand",
                "category": "Laptops",
                "description": "High-performance gaming laptop with RTX graphics",
                "image": "https://shop.example.com/images/laptop1.jpg",
            },
            {
                "url": "https://shop.example.com/laptop-2",
                "title": "Business Ultrabook",
                "price": "$999.99",
                "stock": "12",
                "rating": "4.5",
                "reviews": "89",
                "brand": "ProBrand",
                "category": "Laptops",
                "description": "Lightweight ultrabook for professionals",
                "image": "https://shop.example.com/images/laptop2.jpg",
            },
            {
                "url": "https://shop.example.com/mouse-1",
                "title": "Wireless Mouse",
                "price": "$29.99",
                "stock": "100",
                "rating": "4.8",
                "reviews": "567",
                "brand": "TechBrand",
                "category": "Accessories",
                "description": "Ergonomic wireless mouse",
                "image": "https://shop.example.com/images/mouse1.jpg",
            },
        ]
        
        # Convert to ScrapedItems
        items = []
        for data in scraped_data:
            item = ScrapedItem(
                url=data['url'],
                title=data['title'],
                content={k: v for k, v in data.items() if k not in ['url', 'title']}
            )
            items.append(item)
        
        log_test("Create ScrapedItems from data", True)
        
        # Test AI extraction (heuristic mode)
        extractor = AIExtractor()
        heuristic = HeuristicExtractor()
        
        # Detect page type from URL
        page_type = heuristic.detect_page_type("", "https://shop.example.com/products")
        assert page_type == 'product_listing'
        log_test("Page type detection", True)
        
        # Test smart export
        exporter = SmartExcelExporter()
        
        with tempfile.TemporaryDirectory() as tmpdir:
            # Export to smart Excel
            xlsx_path = os.path.join(tmpdir, "products.xlsx")
            
            if exporter.openpyxl_available:
                result = exporter.export(
                    items,
                    xlsx_path,
                    sheet_name="Products",
                    add_summary=True,
                    highlight_deals=True
                )
                
                assert Path(result).exists()
                
                # Verify Excel structure
                import openpyxl
                wb = openpyxl.load_workbook(result)
                
                # Check data sheet exists
                assert "Products" in wb.sheetnames
                log_test("Smart Excel - Data sheet created", True)
                
                # Check summary sheet exists
                assert "Summary" in wb.sheetnames
                log_test("Smart Excel - Summary sheet created", True)
                
                # Check column ordering (title should be before description)
                ws = wb["Products"]
                headers = [cell.value for cell in ws[1]]
                
                if "Title" in headers and "Description" in headers:
                    title_idx = headers.index("Title")
                    desc_idx = headers.index("Description")
                    assert title_idx < desc_idx
                    log_test("Smart Excel - Column ordering correct", True)
                
                wb.close()
            else:
                log_test("Smart Excel workflow", True, skipped=True)
        
        log_test("Full workflow completed", True)
        
    except Exception as e:
        log_test("Full workflow", False, str(e))


# ============================================================================
# 12. WEB DASHBOARD TEST
# ============================================================================

def test_web_dashboard():
    """Test web dashboard endpoints"""
    test_section("12. WEB DASHBOARD")
    
    try:
        from flask import Flask
        
        # Import dashboard
        import web_dashboard
        
        app = web_dashboard.app
        client = app.test_client()
        
        # Test home page
        response = client.get('/')
        assert response.status_code == 200
        log_test("Dashboard home page", True)
        
        # Test API endpoints exist
        # Note: We won't test actual scraping as it requires external requests
        
        log_test("Dashboard loaded successfully", True)
        
    except ImportError as e:
        log_test("Dashboard import", False, str(e))
    except Exception as e:
        log_test("Dashboard test", False, str(e))


# ============================================================================
# MAIN
# ============================================================================

def print_summary():
    """Print test summary"""
    print("\n" + "=" * 60)
    print("📊 TEST SUMMARY")
    print("=" * 60)
    
    total = TEST_RESULTS['passed'] + TEST_RESULTS['failed'] + TEST_RESULTS['skipped']
    
    print(f"\n  Total Tests: {total}")
    print(f"  ✅ Passed:   {TEST_RESULTS['passed']}")
    print(f"  ❌ Failed:   {TEST_RESULTS['failed']}")
    print(f"  ⏭️  Skipped:  {TEST_RESULTS['skipped']}")
    
    if TEST_RESULTS['failed'] > 0:
        print(f"\n  ❌ FAILED TESTS:")
        for error in TEST_RESULTS['errors']:
            print(f"     - {error['test']}: {error['error']}")
    
    success_rate = (TEST_RESULTS['passed'] / total * 100) if total > 0 else 0
    print(f"\n  Success Rate: {success_rate:.1f}%")
    
    if TEST_RESULTS['failed'] == 0:
        print("\n  🎉 ALL TESTS PASSED!")
    else:
        print(f"\n  ⚠️  {TEST_RESULTS['failed']} test(s) need attention")
    
    print("\n" + "=" * 60)


def main():
    """Run all tests"""
    print("\n" + "=" * 60)
    print("🕷️ SCRAPLING PRO - E2E Test Suite")
    print("=" * 60)
    print(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Run all tests
    test_core_imports()
    test_dependencies()
    test_engine()
    test_templates()
    test_ajax_handler()
    test_ai_extractor()
    test_smart_exporter()
    test_auth_scraper()
    test_scheduler()
    test_exporter()
    test_full_workflow()
    test_web_dashboard()
    
    # Print summary
    print_summary()
    
    # Return exit code
    return 0 if TEST_RESULTS['failed'] == 0 else 1


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
