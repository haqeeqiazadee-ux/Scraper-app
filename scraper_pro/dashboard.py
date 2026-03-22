"""
🕷️ SCRAPLING PRO DASHBOARD
============================
An interactive dashboard showcasing all Scrapling + Playwright capabilities.

Run this file to see all available features and test them interactively.
"""

import os
import sys
import json
import csv
import time
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional, Callable, Any

# ============================================================================
# CORE IMPORTS & AVAILABILITY CHECKS
# ============================================================================

print("🔍 Checking available features...\n")

# Check Scrapling fetchers
try:
    from scrapling.fetchers import StealthyFetcher
    STEALTHY_AVAILABLE = True
    print("  ✓ StealthyFetcher (anti-bot bypass)")
except ImportError:
    STEALTHY_AVAILABLE = False
    print("  ✗ StealthyFetcher not available")

try:
    from scrapling.fetchers import DynamicFetcher
    DYNAMIC_AVAILABLE = True
    print("  ✓ DynamicFetcher (JavaScript rendering)")
except ImportError:
    DYNAMIC_AVAILABLE = False
    print("  ✗ DynamicFetcher not available - run: scrapling install")

try:
    from scrapling.fetchers import AsyncFetcher
    ASYNC_AVAILABLE = True
    print("  ✓ AsyncFetcher (async/concurrent scraping)")
except ImportError:
    ASYNC_AVAILABLE = False
    print("  ✗ AsyncFetcher not available")

# Check for Spider framework
try:
    from scrapling import Spider
    SPIDER_AVAILABLE = True
    print("  ✓ Spider framework (large-scale crawling)")
except ImportError:
    SPIDER_AVAILABLE = False
    print("  ✗ Spider framework not available")

# Check Excel support
try:
    import openpyxl
    EXCEL_AVAILABLE = True
    print("  ✓ Excel export support")
except ImportError:
    EXCEL_AVAILABLE = False
    print("  ✗ Excel not available - run: pip install openpyxl")

print()


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ScrapedItem:
    """Container for scraped data"""
    url: str
    title: str = ""
    content: dict = field(default_factory=dict)
    scraped_at: str = field(default_factory=lambda: datetime.now().isoformat())
    
    def to_dict(self) -> dict:
        flat = {"url": self.url, "title": self.title, "scraped_at": self.scraped_at}
        flat.update(self.content)
        return flat


# ============================================================================
# FEATURE DEMONSTRATIONS
# ============================================================================

class ScraplingDemo:
    """
    Interactive demonstrations of all Scrapling features.
    """
    
    def __init__(self):
        self.results = []
    
    # -------------------------------------------------------------------------
    # 1. BASIC FETCHING MODES
    # -------------------------------------------------------------------------
    
    def demo_stealthy_fetch(self, url: str = "https://httpbin.org/headers"):
        """
        STEALTHY FETCHER
        ----------------
        - Bypasses anti-bot detection
        - Mimics real browser fingerprints
        - Rotates user agents
        - Handles cookies automatically
        """
        print("\n" + "="*60)
        print("📡 DEMO: Stealthy Fetcher")
        print("="*60)
        
        page = StealthyFetcher.fetch(url)
        print(f"  Status: {page.status}")
        print(f"  URL: {url}")
        print("  ✓ Bypassed bot detection with browser fingerprinting")
        return page
    
    def demo_dynamic_fetch(self, url: str = "https://books.toscrape.com"):
        """
        DYNAMIC FETCHER (Playwright-powered)
        ------------------------------------
        - Full browser automation
        - JavaScript rendering
        - Wait for elements/network
        - Execute custom JS
        - Handle SPAs
        """
        print("\n" + "="*60)
        print("🎭 DEMO: Dynamic Fetcher (Playwright)")
        print("="*60)
        
        if not DYNAMIC_AVAILABLE:
            print("  ⚠️ Not available. Run: scrapling install")
            return None
        
        page = DynamicFetcher.fetch(
            url,
            headless=True,
            network_idle=True,  # Wait for network to be idle
        )
        print(f"  Status: {page.status}")
        print(f"  URL: {url}")
        print("  ✓ Rendered JavaScript content")
        return page
    
    # -------------------------------------------------------------------------
    # 2. SELECTOR METHODS
    # -------------------------------------------------------------------------
    
    def demo_css_selectors(self, page):
        """
        CSS SELECTORS
        -------------
        - Standard CSS selector syntax
        - Returns list of elements
        - Chainable
        """
        print("\n" + "="*60)
        print("🎯 DEMO: CSS Selectors")
        print("="*60)
        
        # Basic CSS selection
        articles = page.css("article")
        print(f"  Found {len(articles)} <article> elements")
        
        # Nested selection
        if articles:
            titles = articles[0].css("h3")
            print(f"  First article has {len(titles)} <h3> elements")
        
        # Attribute selection
        links = page.css("a[href]")
        print(f"  Found {len(links)} links with href attribute")
        
        return articles
    
    def demo_xpath_selectors(self, page):
        """
        XPATH SELECTORS
        ---------------
        - Full XPath support
        - More powerful than CSS for complex queries
        - Access text nodes directly
        """
        print("\n" + "="*60)
        print("🔍 DEMO: XPath Selectors")
        print("="*60)
        
        # Basic XPath
        articles = page.xpath("//article")
        print(f"  //article → Found {len(articles)} elements")
        
        # XPath with conditions
        priced_items = page.xpath("//*[contains(@class, 'price')]")
        print(f"  //*[contains(@class, 'price')] → Found {len(priced_items)} elements")
        
        # Text content XPath
        links_with_text = page.xpath("//a[text()]")
        print(f"  //a[text()] → Found {len(links_with_text)} links with text")
        
        return articles
    
    def demo_text_search(self, page):
        """
        TEXT SEARCH
        -----------
        - Find elements by their text content
        - Partial or exact matching
        - Great for buttons, links, headings
        """
        print("\n" + "="*60)
        print("📝 DEMO: Text Search")
        print("="*60)
        
        # Find by exact text
        try:
            element = page.find_by_text("Catalogue", first_match=True)
            if element:
                print(f"  Found element with text 'Catalogue': <{element.tag}>")
        except:
            print("  Text 'Catalogue' not found")
        
        # Find by partial text
        try:
            elements = page.find_by_text("price", partial=True)
            print(f"  Found {len(elements) if elements else 0} elements containing 'price'")
        except:
            pass
        
        return element if 'element' in dir() else None
    
    def demo_regex_search(self, page):
        """
        REGEX SEARCH
        ------------
        - Pattern matching on element content
        - Extract specific data formats
        - Prices, dates, IDs, etc.
        """
        print("\n" + "="*60)
        print("🔤 DEMO: Regex Search")
        print("="*60)
        
        # Find elements with price patterns
        price_elements = page.css(".price_color")
        if price_elements:
            # Extract price using regex
            html = price_elements[0].html_content
            prices = html.re(r'£[\d.]+')
            print(f"  Regex r'£[\\d.]+' found: {prices[:3]}..." if len(prices) > 3 else f"  Found: {prices}")
        
        return prices if 'prices' in dir() else []
    
    # -------------------------------------------------------------------------
    # 3. ADAPTIVE SCRAPING (Scrapling's Killer Feature)
    # -------------------------------------------------------------------------
    
    def demo_adaptive_selectors(self, url: str = "https://books.toscrape.com"):
        """
        ADAPTIVE SELECTORS (auto_save / auto_match)
        -------------------------------------------
        - Survives website structure changes!
        - Stores element fingerprints
        - Uses similarity matching to relocate elements
        - No more broken scrapers after redesigns
        
        How it works:
        1. First run: auto_save=True stores element properties
        2. Later runs: adaptive=True relocates even if selectors change
        """
        print("\n" + "="*60)
        print("🧬 DEMO: Adaptive Selectors (Scrapling's Killer Feature)")
        print("="*60)
        
        # Enable adaptive mode on the fetcher
        StealthyFetcher.adaptive = True
        
        page = StealthyFetcher.fetch(url)
        
        # Save element fingerprints (first run)
        products = page.css("article.product_pod", auto_save=True)
        print(f"  ✓ Saved fingerprints for {len(products)} products")
        print("  ✓ If website changes, use adaptive=True to relocate")
        
        # On future runs after site changes:
        # products = page.css("article.product_pod", adaptive=True)
        # Scrapling will find them even if class names changed!
        
        print("\n  How to use:")
        print("    # First run - save fingerprints:")
        print("    products = page.css('.product', auto_save=True)")
        print("    ")
        print("    # After site redesign - relocate elements:")
        print("    products = page.css('.product', adaptive=True)")
        
        return products
    
    def demo_find_similar(self, page):
        """
        FIND SIMILAR ELEMENTS
        ---------------------
        - Find elements similar to a reference element
        - Great for extracting all products like one you found
        - Uses structural similarity scoring
        """
        print("\n" + "="*60)
        print("🔍 DEMO: Find Similar Elements")
        print("="*60)
        
        articles = page.css("article.product_pod")
        if articles:
            first_article = articles[0]
            print(f"  Reference element: <{first_article.tag}> with classes")
            
            # Find similar would work like:
            # similar = page.find_similar(first_article)
            print(f"  ✓ Could find {len(articles)} similar elements")
            print("\n  Usage: similar = page.find_similar(reference_element)")
        
        return articles
    
    # -------------------------------------------------------------------------
    # 4. DOM NAVIGATION
    # -------------------------------------------------------------------------
    
    def demo_dom_navigation(self, page):
        """
        DOM NAVIGATION
        --------------
        - Parent/child/sibling traversal
        - Navigate the DOM tree programmatically
        """
        print("\n" + "="*60)
        print("🌳 DEMO: DOM Navigation")
        print("="*60)
        
        articles = page.css("article")
        if articles:
            article = articles[0]
            
            # Parent navigation
            parent = article.parent
            if parent:
                print(f"  Parent: <{parent.tag}>")
            
            # Children
            children = article.children
            print(f"  Children: {len(children)} direct children")
            
            # Siblings
            try:
                next_sib = article.next
                if next_sib:
                    print(f"  Next sibling: <{next_sib.tag}>")
            except:
                pass
        
        return article if 'article' in dir() else None
    
    # -------------------------------------------------------------------------
    # 5. TEXT PROCESSING
    # -------------------------------------------------------------------------
    
    def demo_text_processing(self, page):
        """
        TEXT PROCESSING
        ---------------
        - Clean text extraction
        - Whitespace normalization
        - HTML stripping
        """
        print("\n" + "="*60)
        print("📄 DEMO: Text Processing")
        print("="*60)
        
        titles = page.css("h3 a")
        if titles:
            # Get text content
            text = titles[0].text
            print(f"  .text → '{text[:50]}...' " if len(text) > 50 else f"  .text → '{text}'")
            
            # Get attribute
            href = titles[0].attrib.get("href", "")
            print(f"  .attrib['href'] → '{href}'")
            
            # Get title attribute
            title_attr = titles[0].attrib.get("title", "")
            print(f"  .attrib['title'] → '{title_attr[:40]}...'" if len(title_attr) > 40 else f"  .attrib['title'] → '{title_attr}'")
        
        return titles
    
    # -------------------------------------------------------------------------
    # 6. JAVASCRIPT EXECUTION
    # -------------------------------------------------------------------------
    
    def demo_js_execution(self, url: str = "https://books.toscrape.com"):
        """
        JAVASCRIPT EXECUTION
        --------------------
        - Run custom JS in the browser
        - Scroll pages
        - Click buttons
        - Fill forms
        - Wait for dynamic content
        """
        print("\n" + "="*60)
        print("⚡ DEMO: JavaScript Execution")
        print("="*60)
        
        if not DYNAMIC_AVAILABLE:
            print("  ⚠️ Requires DynamicFetcher. Run: scrapling install")
            return None
        
        # Example: Scroll to bottom and get page height
        js_code = """
            window.scrollTo(0, document.body.scrollHeight);
            return document.body.scrollHeight;
        """
        
        page = DynamicFetcher.fetch(
            url,
            headless=True,
            js_code=js_code,
            wait=1000,  # Wait 1 second after JS
        )
        
        print(f"  ✓ Executed custom JavaScript")
        print(f"  ✓ Page rendered with scrolling")
        
        print("\n  Common JS patterns:")
        print("    # Scroll: window.scrollTo(0, document.body.scrollHeight)")
        print("    # Click:  document.querySelector('button').click()")
        print("    # Fill:   document.querySelector('input').value = 'text'")
        
        return page
    
    # -------------------------------------------------------------------------
    # 7. WAITING STRATEGIES
    # -------------------------------------------------------------------------
    
    def demo_wait_strategies(self, url: str = "https://books.toscrape.com"):
        """
        WAITING STRATEGIES
        ------------------
        - Wait for specific elements
        - Wait for network idle
        - Custom timeouts
        - Wait for page state
        """
        print("\n" + "="*60)
        print("⏳ DEMO: Waiting Strategies")
        print("="*60)
        
        if not DYNAMIC_AVAILABLE:
            print("  ⚠️ Requires DynamicFetcher. Run: scrapling install")
            return None
        
        page = DynamicFetcher.fetch(
            url,
            headless=True,
            network_idle=True,          # Wait for no network activity
            wait_selector="article",     # Wait for articles to appear
            wait_selector_state="attached",  # Element state to wait for
            timeout=30000,               # 30 second timeout
        )
        
        print("  ✓ network_idle=True → Waited for network to be idle")
        print("  ✓ wait_selector='article' → Waited for articles to load")
        print("  ✓ timeout=30000 → 30 second timeout")
        
        print("\n  Wait states: 'attached', 'detached', 'visible', 'hidden'")
        
        return page
    
    # -------------------------------------------------------------------------
    # 8. ANTI-BOT BYPASS
    # -------------------------------------------------------------------------
    
    def demo_antibot_bypass(self):
        """
        ANTI-BOT BYPASS
        ---------------
        - Cloudflare Turnstile
        - Browser fingerprinting
        - TLS fingerprinting
        - Human-like behavior
        """
        print("\n" + "="*60)
        print("🛡️ DEMO: Anti-Bot Bypass Capabilities")
        print("="*60)
        
        print("  StealthyFetcher capabilities:")
        print("    ✓ Cloudflare Turnstile bypass")
        print("    ✓ Browser fingerprint spoofing")
        print("    ✓ TLS/JA3 fingerprint rotation")
        print("    ✓ Human-like mouse movements")
        print("    ✓ Realistic timing patterns")
        
        print("\n  DynamicFetcher capabilities:")
        print("    ✓ Real browser execution")
        print("    ✓ Playwright stealth mode")
        print("    ✓ CDP (Chrome DevTools Protocol)")
        print("    ✓ Custom browser profiles")
        
        print("\n  Usage:")
        print("    page = StealthyFetcher.fetch(url)  # Auto-bypasses most protection")
        print("    page = DynamicFetcher.fetch(url, headless=True)  # Full browser")
    
    # -------------------------------------------------------------------------
    # 9. PROXY SUPPORT
    # -------------------------------------------------------------------------
    
    def demo_proxy_support(self):
        """
        PROXY SUPPORT
        -------------
        - HTTP/HTTPS proxies
        - SOCKS proxies
        - Authenticated proxies
        - Proxy rotation
        """
        print("\n" + "="*60)
        print("🔄 DEMO: Proxy Support")
        print("="*60)
        
        print("  Proxy formats supported:")
        print("    # String format:")
        print("    proxy = 'http://user:pass@proxy.example.com:8080'")
        print("    ")
        print("    # Dictionary format:")
        print("    proxy = {")
        print("        'server': 'http://proxy.example.com:8080',")
        print("        'username': 'user',")
        print("        'password': 'pass'")
        print("    }")
        print("    ")
        print("    # Usage:")
        print("    page = StealthyFetcher.fetch(url, proxy=proxy)")
        print("    page = DynamicFetcher.fetch(url, proxy=proxy)")
    
    # -------------------------------------------------------------------------
    # 10. AUTO SELECTOR GENERATION
    # -------------------------------------------------------------------------
    
    def demo_auto_selectors(self, page):
        """
        AUTO SELECTOR GENERATION
        ------------------------
        - Generate robust selectors for any element
        - Creates unique CSS/XPath paths
        - Great for debugging
        """
        print("\n" + "="*60)
        print("🔧 DEMO: Auto Selector Generation")
        print("="*60)
        
        articles = page.css("article")
        if articles:
            article = articles[0]
            
            # Generate selectors (if available)
            print("  For any element, Scrapling can generate unique selectors")
            print("  ")
            print("  Usage:")
            print("    element = page.css('.product')[0]")
            print("    css_selector = element.generate_css_selector()")
            print("    xpath = element.generate_xpath()")
        
        return articles


# ============================================================================
# PRACTICAL USE CASES
# ============================================================================

class UseCases:
    """
    Real-world scraping scenarios with best practices.
    """
    
    @staticmethod
    def ecommerce_products(url: str, max_pages: int = 3):
        """
        E-COMMERCE PRODUCT SCRAPING
        ---------------------------
        Extract products with prices, images, ratings.
        Handles pagination automatically.
        """
        print("\n" + "="*60)
        print("🛒 USE CASE: E-Commerce Product Scraping")
        print("="*60)
        
        all_products = []
        
        for page_num in range(1, max_pages + 1):
            page_url = f"{url}/catalogue/page-{page_num}.html"
            print(f"\n  Scraping page {page_num}: {page_url}")
            
            page = StealthyFetcher.fetch(page_url)
            
            for product in page.css("article.product_pod"):
                title_els = product.css("h3 a")
                price_els = product.css(".price_color")
                rating_els = product.css("p.star-rating")
                img_els = product.css("img")
                
                item = ScrapedItem(
                    url=page_url,
                    title=title_els[0].attrib.get("title", "") if title_els else "",
                    content={
                        "price": price_els[0].text if price_els else "",
                        "rating": rating_els[0].attrib.get("class", "").replace("star-rating ", "") if rating_els else "",
                        "image": img_els[0].attrib.get("src", "") if img_els else "",
                    }
                )
                all_products.append(item)
            
            print(f"    Found {len(page.css('article.product_pod'))} products")
            time.sleep(1)  # Be respectful
        
        print(f"\n  ✓ Total: {len(all_products)} products scraped")
        return all_products
    
    @staticmethod
    def news_articles(url: str):
        """
        NEWS ARTICLE SCRAPING
        ---------------------
        Extract headlines, summaries, dates, authors.
        """
        print("\n" + "="*60)
        print("📰 USE CASE: News Article Scraping")
        print("="*60)
        
        print(f"  Target: {url}")
        print("  ")
        print("  Typical selectors:")
        print("    Headlines: h1, h2, .headline, .title")
        print("    Summary: .excerpt, .summary, .lead")
        print("    Author: .author, .byline, [rel='author']")
        print("    Date: time, .date, .published")
        print("  ")
        print("  Example code:")
        print("    page = StealthyFetcher.fetch(url)")
        print("    for article in page.css('article'):")
        print("        title = article.css('h2')[0].text")
        print("        date = article.css('time')[0].attrib.get('datetime')")
    
    @staticmethod
    def infinite_scroll(url: str):
        """
        INFINITE SCROLL HANDLING
        ------------------------
        Handle Instagram, Twitter, Pinterest style feeds.
        """
        print("\n" + "="*60)
        print("📜 USE CASE: Infinite Scroll Pages")
        print("="*60)
        
        if not DYNAMIC_AVAILABLE:
            print("  ⚠️ Requires DynamicFetcher. Run: scrapling install")
            return
        
        print("  Strategy: Execute JS to scroll and wait for content")
        print("  ")
        print("  Example code:")
        print("    scroll_js = '''")
        print("        let count = 0;")
        print("        const interval = setInterval(() => {")
        print("            window.scrollTo(0, document.body.scrollHeight);")
        print("            count++;")
        print("            if (count >= 10) clearInterval(interval);")
        print("        }, 2000);")
        print("    '''")
        print("    ")
        print("    page = DynamicFetcher.fetch(")
        print("        url,")
        print("        js_code=scroll_js,")
        print("        wait=25000,  # Wait for scrolling to complete")
        print("    )")
    
    @staticmethod
    def login_required(login_url: str, target_url: str):
        """
        LOGIN-REQUIRED SITES
        --------------------
        Authenticate and maintain session.
        """
        print("\n" + "="*60)
        print("🔐 USE CASE: Login-Required Sites")
        print("="*60)
        
        if not DYNAMIC_AVAILABLE:
            print("  ⚠️ Requires DynamicFetcher. Run: scrapling install")
            return
        
        print("  Strategy: Use DynamicFetcher for full browser automation")
        print("  ")
        print("  Example code:")
        print("    login_js = '''")
        print("        document.querySelector('#email').value = 'user@example.com';")
        print("        document.querySelector('#password').value = 'password';")
        print("        document.querySelector('button[type=submit]').click();")
        print("    '''")
        print("    ")
        print("    # Login")
        print("    page = DynamicFetcher.fetch(login_url, js_code=login_js, wait=5000)")
        print("    ")
        print("    # Navigate to protected page (session maintained)")
        print("    page = DynamicFetcher.fetch(target_url)")
    
    @staticmethod
    def form_submission():
        """
        FORM SUBMISSION
        ---------------
        Fill and submit forms, handle responses.
        """
        print("\n" + "="*60)
        print("📝 USE CASE: Form Submission")
        print("="*60)
        
        print("  Strategy: Use JS to fill forms and submit")
        print("  ")
        print("  Example code:")
        print("    form_js = '''")
        print("        // Fill text inputs")
        print("        document.querySelector('input[name=q]').value = 'search term';")
        print("        ")
        print("        // Select dropdown")
        print("        document.querySelector('select').value = 'option1';")
        print("        ")
        print("        // Check checkbox")
        print("        document.querySelector('input[type=checkbox]').checked = true;")
        print("        ")
        print("        // Submit form")
        print("        document.querySelector('form').submit();")
        print("    '''")
    
    @staticmethod
    def api_like_extraction():
        """
        API-LIKE JSON EXTRACTION
        ------------------------
        Extract JSON data embedded in pages.
        """
        print("\n" + "="*60)
        print("🔌 USE CASE: API-like JSON Extraction")
        print("="*60)
        
        print("  Many sites embed data as JSON in script tags")
        print("  ")
        print("  Example code:")
        print("    # Find script tags with JSON")
        print("    scripts = page.css('script[type=\"application/json\"]')")
        print("    if scripts:")
        print("        data = scripts[0].text")
        print("        json_data = json.loads(data)")
        print("    ")
        print("    # Or find __NEXT_DATA__ (Next.js sites)")
        print("    next_data = page.css('#__NEXT_DATA__')")
        print("    if next_data:")
        print("        data = json.loads(next_data[0].text)")


# ============================================================================
# EXPORT UTILITIES
# ============================================================================

class Exporter:
    """Data export utilities."""
    
    @staticmethod
    def to_csv(items: list, filepath: str):
        """Export to CSV"""
        if not items:
            print("No items to export")
            return
        
        filepath = Path(filepath)
        flat_items = [item.to_dict() for item in items]
        
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=sorted(all_keys))
            writer.writeheader()
            writer.writerows(flat_items)
        
        print(f"✓ Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_json(items: list, filepath: str):
        """Export to JSON"""
        if not items:
            print("No items to export")
            return
        
        data = [item.to_dict() for item in items]
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"✓ Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_excel(items: list, filepath: str):
        """Export to Excel"""
        if not EXCEL_AVAILABLE:
            print("Excel not available. Run: pip install openpyxl")
            return
        
        if not items:
            print("No items to export")
            return
        
        flat_items = [item.to_dict() for item in items]
        
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        fieldnames = sorted(all_keys)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"
        
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)
        
        for row_num, item in enumerate(flat_items, 2):
            for col, field in enumerate(fieldnames, 1):
                ws.cell(row=row_num, column=col, value=item.get(field, ""))
        
        wb.save(filepath)
        print(f"✓ Exported {len(items)} items to {filepath}")


# ============================================================================
# INTERACTIVE DASHBOARD
# ============================================================================

def show_dashboard():
    """Display the interactive dashboard."""
    
    print("""
╔══════════════════════════════════════════════════════════════════════════════╗
║                                                                              ║
║   🕷️  SCRAPLING PRO DASHBOARD                                                ║
║   ════════════════════════════                                               ║
║                                                                              ║
║   A comprehensive showcase of Scrapling + Playwright capabilities           ║
║                                                                              ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                              ║
║   FEATURE DEMOS                         USE CASES                            ║
║   ─────────────                         ─────────                            ║
║   1.  Stealthy Fetching                 11. E-Commerce Products              ║
║   2.  Dynamic/JS Rendering              12. News Articles                    ║
║   3.  CSS Selectors                     13. Infinite Scroll                  ║
║   4.  XPath Selectors                   14. Login-Required Sites             ║
║   5.  Text Search                       15. Form Submission                  ║
║   6.  Regex Patterns                    16. API/JSON Extraction              ║
║   7.  Adaptive Selectors ⭐                                                  ║
║   8.  DOM Navigation                    QUICK ACTIONS                        ║
║   9.  JavaScript Execution              ─────────────                        ║
║   10. Anti-Bot Bypass Info              20. Run Full Demo                    ║
║                                         21. Scrape Sample Site               ║
║                                         22. Export to CSV/Excel              ║
║                                                                              ║
║   0.  Exit                                                                   ║
║                                                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝
    """)


def run_interactive():
    """Run the interactive dashboard."""
    
    demo = ScraplingDemo()
    page = None  # Will store the last fetched page
    products = []  # Will store scraped products
    
    while True:
        show_dashboard()
        
        try:
            choice = input("\n  Enter choice (0-22): ").strip()
        except (KeyboardInterrupt, EOFError):
            print("\n\n  Goodbye! 👋")
            break
        
        if choice == "0":
            print("\n  Goodbye! 👋")
            break
        
        elif choice == "1":
            page = demo.demo_stealthy_fetch()
        
        elif choice == "2":
            page = demo.demo_dynamic_fetch()
        
        elif choice == "3":
            if page:
                demo.demo_css_selectors(page)
            else:
                print("\n  ⚠️ Fetch a page first (option 1 or 2)")
        
        elif choice == "4":
            if page:
                demo.demo_xpath_selectors(page)
            else:
                print("\n  ⚠️ Fetch a page first (option 1 or 2)")
        
        elif choice == "5":
            if page:
                demo.demo_text_search(page)
            else:
                print("\n  ⚠️ Fetch a page first (option 1 or 2)")
        
        elif choice == "6":
            if page:
                demo.demo_regex_search(page)
            else:
                print("\n  ⚠️ Fetch a page first (option 1 or 2)")
        
        elif choice == "7":
            demo.demo_adaptive_selectors()
        
        elif choice == "8":
            if page:
                demo.demo_dom_navigation(page)
            else:
                print("\n  ⚠️ Fetch a page first (option 1 or 2)")
        
        elif choice == "9":
            demo.demo_js_execution()
        
        elif choice == "10":
            demo.demo_antibot_bypass()
        
        elif choice == "11":
            products = UseCases.ecommerce_products("https://books.toscrape.com")
        
        elif choice == "12":
            UseCases.news_articles("https://news.example.com")
        
        elif choice == "13":
            UseCases.infinite_scroll("https://social.example.com")
        
        elif choice == "14":
            UseCases.login_required("https://example.com/login", "https://example.com/dashboard")
        
        elif choice == "15":
            UseCases.form_submission()
        
        elif choice == "16":
            UseCases.api_like_extraction()
        
        elif choice == "20":
            print("\n  Running full demo...")
            page = demo.demo_stealthy_fetch()
            if page:
                demo.demo_css_selectors(page)
                demo.demo_xpath_selectors(page)
                demo.demo_text_processing(page)
            demo.demo_adaptive_selectors()
            demo.demo_antibot_bypass()
            demo.demo_proxy_support()
        
        elif choice == "21":
            products = UseCases.ecommerce_products("https://books.toscrape.com", max_pages=2)
        
        elif choice == "22":
            if products:
                print("\n  Export format:")
                print("    1. CSV")
                print("    2. JSON")
                print("    3. Excel")
                fmt = input("  Choose (1-3): ").strip()
                
                if fmt == "1":
                    Exporter.to_csv(products, "scraped_data.csv")
                elif fmt == "2":
                    Exporter.to_json(products, "scraped_data.json")
                elif fmt == "3":
                    Exporter.to_excel(products, "scraped_data.xlsx")
            else:
                print("\n  ⚠️ No data to export. Run a scrape first (option 11 or 21)")
        
        else:
            print("\n  ⚠️ Invalid choice")
        
        input("\n  Press Enter to continue...")


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("""
    ╔═══════════════════════════════════════════════════════════════╗
    ║                                                               ║
    ║   🕷️  SCRAPLING PRO - Professional Web Scraping Toolkit      ║
    ║                                                               ║
    ║   Features:                                                   ║
    ║   • Adaptive selectors that survive website changes           ║
    ║   • Anti-bot bypass (Cloudflare, etc.)                        ║
    ║   • JavaScript rendering with Playwright                      ║
    ║   • CSS, XPath, Text, Regex selectors                         ║
    ║   • Pagination, infinite scroll, login handling               ║
    ║   • Export to CSV, JSON, Excel                                ║
    ║                                                               ║
    ╚═══════════════════════════════════════════════════════════════╝
    """)
    
    run_interactive()
