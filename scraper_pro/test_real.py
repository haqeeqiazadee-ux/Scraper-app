"""
🧪 SCRAPLING PRO - Real-World Test
===================================
Run this on YOUR machine to verify everything works.

Usage: python test_real.py
"""

import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent))

def print_header(text):
    print(f"\n{'='*60}")
    print(f"🧪 {text}")
    print(f"{'='*60}\n")

def test_step(name, func):
    """Run a test step"""
    try:
        result = func()
        print(f"  ✅ {name}: PASSED")
        return True, result
    except Exception as e:
        print(f"  ❌ {name}: FAILED - {e}")
        return False, None


def main():
    print("\n" + "="*60)
    print("🕷️ SCRAPLING PRO - Real-World Verification Test")
    print("="*60)
    
    results = {'passed': 0, 'failed': 0}
    
    # ========================================
    # TEST 1: Check Scrapling Installation
    # ========================================
    print_header("1. SCRAPLING INSTALLATION")
    
    def check_scrapling():
        from scrapling.fetchers import StealthyFetcher
        return True
    
    passed, _ = test_step("Scrapling StealthyFetcher", check_scrapling)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
        print("\n  ⚠️  FIX: Run these commands:")
        print("      pip install scrapling[all]")
        print("      scrapling install")
    
    def check_playwright():
        from scrapling.fetchers import PlayWrightFetcher
        return True
    
    passed, _ = test_step("Scrapling PlayWrightFetcher (for JS sites)", check_playwright)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
        print("\n  ⚠️  FIX: Run: scrapling install")
    
    # ========================================
    # TEST 2: Live Scraping Test (Simple Site)
    # ========================================
    print_header("2. LIVE SCRAPING TEST")
    
    def test_stealthy_fetch():
        from scrapling.fetchers import StealthyFetcher
        
        print("  📡 Fetching https://books.toscrape.com ...")
        page = StealthyFetcher.fetch(
            "https://books.toscrape.com",
            timeout=30000  # 30 seconds = 30000 MILLISECONDS (Scrapling uses ms!)
        )
        
        assert page.status == 200, f"Got status {page.status}"
        
        # Check we got content
        titles = page.css("article.product_pod h3 a")
        assert len(titles) > 0, "No products found"
        
        print(f"  📦 Found {len(titles)} products")
        print(f"  📦 First product: {titles[0].attrib.get('title', 'N/A')}")
        
        return len(titles)
    
    passed, product_count = test_step("Stealthy fetch (books.toscrape.com)", test_stealthy_fetch)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
    
    # ========================================
    # TEST 3: Engine v2 Test
    # ========================================
    print_header("3. ENGINE V2 TEST")
    
    def test_engine_v2():
        from engine_v2 import ScrapingEngine, ScrapedItem
        
        engine = ScrapingEngine(
            mode="stealthy",
            timeout=30,
            max_retries=2
        )
        
        # Verify timeout is set correctly
        assert engine.timeout == 30, f"Timeout is {engine.timeout}"
        assert engine.timeout_ms == 30000, f"Timeout ms is {engine.timeout_ms}"
        
        print(f"  ⚙️  Engine timeout: {engine.timeout}s ({engine.timeout_ms}ms)")
        
        return True
    
    passed, _ = test_step("Engine v2 timeout handling", test_engine_v2)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
    
    def test_engine_fetch():
        from engine_v2 import ScrapingEngine
        
        engine = ScrapingEngine(mode="stealthy", timeout=30)
        
        print("  📡 Fetching via engine...")
        page = engine.fetch("https://books.toscrape.com")
        
        assert page is not None, "Page is None"
        assert page.status == 200, f"Status {page.status}"
        
        return True
    
    passed, _ = test_step("Engine v2 live fetch", test_engine_fetch)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
    
    # ========================================
    # TEST 4: Smart Excel Export
    # ========================================
    print_header("4. SMART EXCEL EXPORT")
    
    def test_smart_export():
        from smart_exporter import SmartExcelExporter
        import tempfile
        import os
        
        # Create test data (simulating scraped products)
        test_data = [
            {
                "title": "Gaming Laptop",
                "price": "$1,299.99",
                "stock": "15",
                "rating": "4.7",
                "category": "Laptops",
                "description": "High-performance gaming laptop",
                "url": "https://shop.com/laptop-1"
            },
            {
                "title": "Wireless Mouse",
                "price": "$29.99",
                "stock": "100",
                "rating": "4.8",
                "category": "Accessories",
                "description": "Ergonomic wireless mouse",
                "url": "https://shop.com/mouse-1"
            },
        ]
        
        exporter = SmartExcelExporter()
        
        # Export to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        result = exporter.export(test_data, filepath, add_summary=True)
        
        assert result is not None, "Export returned None"
        assert os.path.exists(result), "File not created"
        
        file_size = os.path.getsize(result)
        assert file_size > 1000, f"File too small: {file_size} bytes"
        
        print(f"  📊 Created: {result}")
        print(f"  📊 Size: {file_size:,} bytes")
        
        # Verify structure
        import openpyxl
        wb = openpyxl.load_workbook(result)
        
        assert "Data" in wb.sheetnames or "Products" in wb.sheetnames, "No data sheet"
        assert "Summary" in wb.sheetnames, "No summary sheet"
        
        # Check column ordering
        ws = wb.active
        headers = [cell.value for cell in ws[1] if cell.value]
        print(f"  📊 Columns: {headers}")
        
        # Title should come before Description
        if "Title" in headers and "Description" in headers:
            assert headers.index("Title") < headers.index("Description"), "Column order wrong"
            print("  📊 Column ordering: CORRECT")
        
        wb.close()
        os.unlink(result)
        
        return True
    
    passed, _ = test_step("Smart Excel export with auto-ordering", test_smart_export)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
    
    # ========================================
    # TEST 5: Full Workflow
    # ========================================
    print_header("5. FULL SCRAPING WORKFLOW")
    
    def test_full_workflow():
        from engine_v2 import ScrapingEngine, ScrapedItem
        from smart_exporter import SmartExcelExporter
        import tempfile
        import os
        
        print("  📡 Step 1: Fetching books.toscrape.com...")
        engine = ScrapingEngine(mode="stealthy", timeout=30)
        page = engine.fetch("https://books.toscrape.com")
        
        assert page is not None, "Fetch failed"
        
        print("  🔍 Step 2: Extracting products...")
        products = []
        for article in page.css("article.product_pod"):
            title_el = article.css("h3 a")
            price_el = article.css(".price_color")
            
            title = title_el[0].attrib.get('title', '') if title_el else ''
            price = price_el[0].text if price_el else ''
            
            if title:
                products.append({
                    "title": title,
                    "price": price,
                    "category": "Books",
                    "source": "books.toscrape.com"
                })
        
        print(f"  📦 Found {len(products)} products")
        assert len(products) > 0, "No products extracted"
        
        print("  📊 Step 3: Exporting to Excel...")
        exporter = SmartExcelExporter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        result = exporter.export(products, filepath, add_summary=True)
        
        file_size = os.path.getsize(result)
        print(f"  📊 Created Excel: {file_size:,} bytes")
        
        os.unlink(result)
        
        print(f"\n  🎉 SUCCESS: Scraped {len(products)} products and exported to Excel!")
        
        return len(products)
    
    passed, count = test_step("Full scrape → extract → export workflow", test_full_workflow)
    if passed:
        results['passed'] += 1
    else:
        results['failed'] += 1
    
    # ========================================
    # SUMMARY
    # ========================================
    print_header("TEST SUMMARY")
    
    total = results['passed'] + results['failed']
    success_rate = (results['passed'] / total * 100) if total > 0 else 0
    
    print(f"  Total Tests: {total}")
    print(f"  ✅ Passed:   {results['passed']}")
    print(f"  ❌ Failed:   {results['failed']}")
    print(f"\n  Success Rate: {success_rate:.0f}%")
    
    if results['failed'] == 0:
        print("\n  🎉 ALL TESTS PASSED! Your installation is working correctly.")
        print("\n  You can now:")
        print("    1. Run: python web_dashboard.py")
        print("    2. Open: http://localhost:5000")
        print("    3. Start scraping!")
    else:
        print(f"\n  ⚠️  {results['failed']} test(s) failed. See errors above for fixes.")
    
    print("\n" + "="*60 + "\n")
    
    return results['failed'] == 0


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
