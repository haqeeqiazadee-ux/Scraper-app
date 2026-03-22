"""
Power Scraper - A Production-Ready Web Scraping Framework
Using Scrapling for maximum capability

Features:
- JavaScript rendering with DynamicFetcher
- Anti-bot protection bypass with StealthyFetcher
- Login/authentication support
- Pagination & infinite scroll handling
- CSV/Excel export
- Retry logic & error handling
- Rate limiting to be respectful
"""

import csv
import json
import time
import random
from pathlib import Path
from datetime import datetime
from typing import Optional, Callable, Any
from dataclasses import dataclass, field

# Scrapling imports - using new v0.4+ API
from scrapling.fetchers import Fetcher, StealthyFetcher

# DynamicFetcher for JS rendering - may not be available if browser not installed
try:
    from scrapling.fetchers import DynamicFetcher
    DYNAMIC_AVAILABLE = True
except Exception:
    DYNAMIC_AVAILABLE = False
    DynamicFetcher = None

# For Excel export
try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@dataclass
class ScrapedItem:
    """Generic container for scraped data"""
    url: str
    title: str = ""
    content: dict = field(default_factory=dict)
    scraped_at: str = field(default_factory=lambda: datetime.now().isoformat())
    
    def to_flat_dict(self) -> dict:
        """Flatten nested content for CSV export"""
        flat = {"url": self.url, "title": self.title, "scraped_at": self.scraped_at}
        for key, value in self.content.items():
            if isinstance(value, (list, dict)):
                flat[key] = json.dumps(value)
            else:
                flat[key] = value
        return flat


class PowerScraper:
    """
    A powerful, flexible web scraper built on Scrapling.
    
    Supports:
    - Multiple fetcher modes (basic, stealthy, dynamic)
    - Authentication flows
    - Pagination (click-based, URL-based, infinite scroll)
    - Anti-bot evasion
    - Rate limiting
    - Data export (CSV/Excel)
    """
    
    def __init__(
        self,
        mode: str = "stealthy",  # "basic", "stealthy", or "dynamic"
        headless: bool = True,
        rate_limit: float = 1.0,  # seconds between requests
        max_retries: int = 3,
        timeout: int = 30,  # seconds
    ):
        self.mode = mode
        self.headless = headless
        self.rate_limit = rate_limit
        self.max_retries = max_retries
        self.timeout = timeout
        self.results: list[ScrapedItem] = []
        self._last_request_time = 0
        
    def _respect_rate_limit(self):
        """Enforce rate limiting between requests"""
        elapsed = time.time() - self._last_request_time
        if elapsed < self.rate_limit:
            sleep_time = self.rate_limit - elapsed + random.uniform(0.1, 0.5)
            time.sleep(sleep_time)
        self._last_request_time = time.time()
    
    def fetch(self, url: str, **kwargs) -> Any:
        """
        Fetch a URL with retry logic and rate limiting.
        
        Returns a Scrapling page object with enhanced parsing.
        """
        for attempt in range(self.max_retries):
            try:
                self._respect_rate_limit()
                
                if self.mode == "basic":
                    page = Fetcher.fetch(url, timeout=self.timeout, **kwargs)
                elif self.mode == "stealthy":
                    page = StealthyFetcher.fetch(url, timeout=self.timeout, **kwargs)
                else:  # dynamic
                    if not DYNAMIC_AVAILABLE:
                        print("⚠️ DynamicFetcher not available, falling back to stealthy mode")
                        print("  To enable: scrapling install")
                        page = StealthyFetcher.fetch(url, timeout=self.timeout, **kwargs)
                    else:
                        page = DynamicFetcher.fetch(
                            url, 
                            headless=self.headless, 
                            timeout=self.timeout * 1000,  # ms for dynamic
                            **kwargs
                        )
                
                if page.status == 200:
                    return page
                elif page.status == 429:  # Rate limited
                    wait_time = (attempt + 1) * 30
                    print(f"Rate limited. Waiting {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    print(f"HTTP {page.status} for {url}")
                    
            except Exception as e:
                print(f"Attempt {attempt + 1} failed: {e}")
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
                    
        return None
    
    def fetch_with_login(
        self,
        login_url: str,
        target_url: str,
        username: str,
        password: str,
        username_selector: str = "input[name='username'], input[type='email'], #username, #email",
        password_selector: str = "input[name='password'], input[type='password'], #password",
        submit_selector: str = "button[type='submit'], input[type='submit'], .login-btn",
        wait_after_login: int = 3000,
    ) -> Any:
        """
        Perform login and then fetch target URL.
        
        Uses DynamicFetcher for full browser automation.
        """
        if not DYNAMIC_AVAILABLE:
            print("⚠️ DynamicFetcher required for login. Run: scrapling install")
            return None
            
        try:
            # Build JS code for login
            js_code = f"""
                // Fill username
                const usernameField = document.querySelector("{username_selector}");
                if (usernameField) {{
                    usernameField.value = "{username}";
                    usernameField.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
                
                // Fill password
                const passwordField = document.querySelector("{password_selector}");
                if (passwordField) {{
                    passwordField.value = "{password}";
                    passwordField.dispatchEvent(new Event('input', {{ bubbles: true }}));
                }}
                
                // Click submit
                const submitBtn = document.querySelector("{submit_selector}");
                if (submitBtn) {{
                    submitBtn.click();
                }}
            """
            
            # Navigate to login page and execute login
            page = DynamicFetcher.fetch(
                login_url,
                headless=self.headless,
                timeout=self.timeout * 1000,
                js_code=js_code,
                wait=wait_after_login,
            )
            
            # Now fetch the target URL
            time.sleep(wait_after_login / 1000)
            return DynamicFetcher.fetch(target_url, headless=self.headless)
            
        except Exception as e:
            print(f"Login failed: {e}")
            return None
    
    def fetch_with_pagination(
        self,
        start_url: str,
        parse_func: Callable,
        pagination_type: str = "url",  # "url", "click", or "scroll"
        next_page_selector: str = "a.next, .pagination a:last-child, [rel='next']",
        url_pattern: Optional[str] = None,  # e.g., "?page={}" for URL pagination
        max_pages: int = 10,
        scroll_pause: float = 2.0,  # seconds to wait for infinite scroll
        scroll_max: int = 20,  # max scroll attempts
    ) -> list[ScrapedItem]:
        """
        Fetch multiple pages with various pagination strategies.
        
        Args:
            start_url: Initial URL to start scraping
            parse_func: Function that takes a page and returns list[ScrapedItem]
            pagination_type: "url" (modify URL), "click" (click next), "scroll" (infinite)
            next_page_selector: CSS selector for next page button
            url_pattern: Pattern for URL pagination (e.g., "?page={}")
            max_pages: Maximum number of pages to scrape
            scroll_pause: Time to wait for content to load on scroll
            scroll_max: Maximum scroll attempts for infinite scroll
        """
        all_items = []
        
        if pagination_type == "url":
            # URL-based pagination (e.g., ?page=1, ?page=2)
            for page_num in range(1, max_pages + 1):
                if url_pattern:
                    url = start_url + url_pattern.format(page_num)
                else:
                    url = start_url if page_num == 1 else f"{start_url}?page={page_num}"
                
                print(f"Fetching page {page_num}: {url}")
                page = self.fetch(url)
                
                if not page:
                    print(f"Failed to fetch page {page_num}, stopping.")
                    break
                    
                items = parse_func(page)
                if not items:
                    print(f"No items found on page {page_num}, stopping.")
                    break
                    
                all_items.extend(items)
                print(f"Found {len(items)} items on page {page_num}")
                
        elif pagination_type == "click":
            # Click-based pagination (requires DynamicFetcher)
            if not DYNAMIC_AVAILABLE:
                print("⚠️ Click pagination requires DynamicFetcher. Run: scrapling install")
                print("  Falling back to URL pagination...")
                return self.fetch_with_pagination(
                    start_url, parse_func, "url", url_pattern=url_pattern, max_pages=max_pages
                )
            
            page = DynamicFetcher.fetch(start_url, headless=self.headless)
            page_num = 1
            
            while page_num <= max_pages:
                print(f"Processing page {page_num}")
                items = parse_func(page)
                all_items.extend(items)
                print(f"Found {len(items)} items on page {page_num}")
                
                # Try to find next button
                next_btns = page.css(next_page_selector)
                if not next_btns:
                    print("No next button found, stopping.")
                    break
                
                # Click next and get new page
                self._respect_rate_limit()
                page = DynamicFetcher.fetch(
                    start_url,
                    headless=self.headless,
                    js_code=f"""
                        const nextBtn = document.querySelector("{next_page_selector}");
                        if (nextBtn) nextBtn.click();
                    """,
                    wait=2000,
                )
                page_num += 1
                
        elif pagination_type == "scroll":
            # Infinite scroll pagination (requires DynamicFetcher)
            if not DYNAMIC_AVAILABLE:
                print("⚠️ Scroll pagination requires DynamicFetcher. Run: scrapling install")
                return all_items
            
            # Build JavaScript for infinite scroll
            scroll_js = f"""
                let previousHeight = 0;
                let scrollAttempts = 0;
                
                const scrollInterval = setInterval(() => {{
                    window.scrollTo(0, document.body.scrollHeight);
                    scrollAttempts++;
                    
                    if (document.body.scrollHeight === previousHeight || scrollAttempts >= {scroll_max}) {{
                        clearInterval(scrollInterval);
                    }}
                    previousHeight = document.body.scrollHeight;
                }}, {int(scroll_pause * 1000)});
            """
            
            page = DynamicFetcher.fetch(
                start_url,
                headless=self.headless,
                js_code=scroll_js,
                wait=int(scroll_pause * scroll_max * 1000) + 5000,
            )
            
            items = parse_func(page)
            all_items.extend(items)
            print(f"Found {len(items)} items after scrolling")
        
        self.results.extend(all_items)
        return all_items
    
    def extract_product_data(self, page) -> list[ScrapedItem]:
        """
        Generic product data extractor for e-commerce sites.
        
        Override this for site-specific extraction.
        """
        items = []
        
        # Common product card selectors
        product_selectors = [
            ".product", ".product-card", ".product-item",
            "[data-product]", ".item", ".listing-item"
        ]
        
        for selector in product_selectors:
            products = page.css(selector)
            if products:
                for product in products:
                    try:
                        # Get text content safely
                        title_el = product.css("h2, h3, .title, .name, .product-name")
                        price_el = product.css(".price, .cost, [data-price]")
                        img_el = product.css("img")
                        rating_el = product.css(".rating, .stars, [data-rating]")
                        desc_el = product.css(".description, .desc, .summary")
                        
                        item = ScrapedItem(
                            url=str(page.url) if hasattr(page, 'url') else "",
                            title=title_el[0].text if title_el else "",
                            content={
                                "price": price_el[0].text if price_el else "",
                                "image": img_el[0].attrib.get("src", "") if img_el else "",
                                "rating": rating_el[0].text if rating_el else "",
                                "description": desc_el[0].text if desc_el else "",
                            }
                        )
                        if item.title:  # Only add if we got a title
                            items.append(item)
                    except Exception as e:
                        print(f"Error extracting product: {e}")
                        
                if items:
                    break  # Found products with this selector
                    
        return items
    
    def extract_article_data(self, page) -> list[ScrapedItem]:
        """
        Generic article/news extractor.
        
        Override this for site-specific extraction.
        """
        items = []
        
        # Common article selectors
        article_selectors = [
            "article", ".article", ".post", ".story",
            ".news-item", ".entry", "[data-article]"
        ]
        
        for selector in article_selectors:
            articles = page.css(selector)
            if articles:
                for article in articles:
                    try:
                        title_el = article.css("h1, h2, h3, .title, .headline")
                        author_el = article.css(".author, .byline, [rel='author']")
                        date_el = article.css("time, .date, .published, .timestamp")
                        summary_el = article.css(".summary, .excerpt, .lead, p")
                        link_el = article.css("a")
                        
                        item = ScrapedItem(
                            url=str(page.url) if hasattr(page, 'url') else "",
                            title=title_el[0].text if title_el else "",
                            content={
                                "author": author_el[0].text if author_el else "",
                                "date": date_el[0].text if date_el else "",
                                "summary": summary_el[0].text if summary_el else "",
                                "link": link_el[0].attrib.get("href", "") if link_el else "",
                            }
                        )
                        if item.title:
                            items.append(item)
                    except Exception as e:
                        print(f"Error extracting article: {e}")
                        
                if items:
                    break
                    
        return items
    
    def export_csv(self, filepath: str, items: Optional[list[ScrapedItem]] = None):
        """Export scraped items to CSV"""
        items = items or self.results
        if not items:
            print("No items to export")
            return
            
        filepath = Path(filepath)
        flat_items = [item.to_flat_dict() for item in items]
        
        # Get all unique keys
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        fieldnames = sorted(all_keys)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(flat_items)
            
        print(f"Exported {len(items)} items to {filepath}")
    
    def export_excel(self, filepath: str, items: Optional[list[ScrapedItem]] = None):
        """Export scraped items to Excel"""
        if not EXCEL_AVAILABLE:
            print("openpyxl not installed. Run: pip install openpyxl")
            print("Falling back to CSV export...")
            csv_path = str(filepath).replace('.xlsx', '.csv')
            self.export_csv(csv_path, items)
            return
            
        items = items or self.results
        if not items:
            print("No items to export")
            return
            
        filepath = Path(filepath)
        flat_items = [item.to_flat_dict() for item in items]
        
        # Get all unique keys
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        fieldnames = sorted(all_keys)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"
        
        # Header row
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)
            
        # Data rows
        for row_num, item in enumerate(flat_items, 2):
            for col, field in enumerate(fieldnames, 1):
                ws.cell(row=row_num, column=col, value=item.get(field, ""))
        
        wb.save(filepath)
        print(f"Exported {len(items)} items to {filepath}")


# ============================================================================
# EXAMPLE USAGE TEMPLATES
# ============================================================================

def example_ecommerce_scraper():
    """Example: Scrape product data from an e-commerce site"""
    scraper = PowerScraper(mode="dynamic", headless=True)
    
    def parse_products(page):
        items = []
        for product in page.css(".product-card"):
            title_el = product.css(".product-title")
            price_el = product.css(".price")
            img_el = product.css("img")
            
            items.append(ScrapedItem(
                url=str(page.url) if hasattr(page, 'url') else "",
                title=title_el[0].text if title_el else "",
                content={
                    "price": price_el[0].text if price_el else "",
                    "image": img_el[0].attrib.get("src", "") if img_el else "",
                }
            ))
        return items
    
    items = scraper.fetch_with_pagination(
        start_url="https://example-shop.com/products",
        parse_func=parse_products,
        pagination_type="url",
        url_pattern="?page={}",
        max_pages=5,
    )
    
    scraper.export_excel("products.xlsx", items)


def example_news_scraper():
    """Example: Scrape news articles"""
    scraper = PowerScraper(mode="stealthy", rate_limit=2.0)
    
    def parse_articles(page):
        items = []
        for article in page.css("article"):
            title_el = article.css("h2")
            summary_el = article.css("p")
            link_el = article.css("a")
            
            items.append(ScrapedItem(
                url=str(page.url) if hasattr(page, 'url') else "",
                title=title_el[0].text if title_el else "",
                content={
                    "summary": summary_el[0].text if summary_el else "",
                    "link": link_el[0].attrib.get("href", "") if link_el else "",
                }
            ))
        return items
    
    items = scraper.fetch_with_pagination(
        start_url="https://example-news.com",
        parse_func=parse_articles,
        pagination_type="click",
        next_page_selector=".load-more",
        max_pages=10,
    )
    
    scraper.export_csv("articles.csv", items)


if __name__ == "__main__":
    # Quick demo
    print("Power Scraper Framework")
    print("=" * 50)
    print("\nAvailable modes:")
    print("  - basic: Fast, simple HTTP requests")
    print("  - stealthy: Anti-bot headers and fingerprinting")
    print("  - dynamic: Full browser for JS-heavy sites (requires: scrapling install)")
    print("\nPagination types:")
    print("  - url: Modify URL parameters (?page=1, ?page=2)")
    print("  - click: Click 'Next' button (requires dynamic mode)")
    print("  - scroll: Infinite scroll handling (requires dynamic mode)")
    print("\nSee example functions in this file for usage patterns!")
