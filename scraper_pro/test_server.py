"""
🧪 SCRAPLING PRO - Live Test Server
====================================
A Flask server that runs tests and provides real-time results via web interface.

Run: python test_server.py
Open: http://localhost:5555

This allows Claude to see test results in real-time!
"""

import sys
import os
import json
import time
import threading
import traceback
from pathlib import Path
from datetime import datetime
from flask import Flask, render_template_string, jsonify, request

sys.path.insert(0, str(Path(__file__).parent))

app = Flask(__name__)

# Store test results
test_results = {
    'status': 'idle',  # idle, running, complete
    'started_at': None,
    'completed_at': None,
    'tests': [],
    'summary': {'passed': 0, 'failed': 0, 'skipped': 0}
}

# HTML Template for test dashboard
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>🧪 Scrapling Pro - Live Test Dashboard</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <style>
        @keyframes pulse { 0%, 100% { opacity: 1; } 50% { opacity: 0.5; } }
        .pulse { animation: pulse 1s infinite; }
        body { background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%); }
    </style>
</head>
<body class="min-h-screen text-white p-8">
    <div class="max-w-4xl mx-auto">
        <div class="flex items-center justify-between mb-8">
            <div>
                <h1 class="text-3xl font-bold">🧪 Scrapling Pro - Live Test Dashboard</h1>
                <p class="text-gray-400 mt-1">Real-time test results</p>
            </div>
            <div id="status-badge" class="px-4 py-2 rounded-full text-sm font-medium bg-gray-700">
                Idle
            </div>
        </div>
        
        <!-- Control Panel -->
        <div class="bg-white/10 rounded-xl p-6 mb-6">
            <div class="flex gap-4">
                <button onclick="runTests()" id="run-btn"
                    class="px-6 py-3 bg-green-600 hover:bg-green-700 rounded-lg font-medium transition">
                    ▶️ Run All Tests
                </button>
                <button onclick="runTest('scraping')" 
                    class="px-4 py-3 bg-blue-600 hover:bg-blue-700 rounded-lg font-medium transition">
                    🌐 Test Scraping
                </button>
                <button onclick="runTest('export')"
                    class="px-4 py-3 bg-purple-600 hover:bg-purple-700 rounded-lg font-medium transition">
                    📊 Test Export
                </button>
                <button onclick="runTest('auth')"
                    class="px-4 py-3 bg-red-600 hover:bg-red-700 rounded-lg font-medium transition">
                    🔐 Test Auth
                </button>
            </div>
        </div>
        
        <!-- Summary Stats -->
        <div class="grid grid-cols-4 gap-4 mb-6">
            <div class="bg-white/10 rounded-xl p-4 text-center">
                <div id="stat-total" class="text-3xl font-bold">0</div>
                <div class="text-gray-400 text-sm">Total</div>
            </div>
            <div class="bg-green-500/20 rounded-xl p-4 text-center">
                <div id="stat-passed" class="text-3xl font-bold text-green-400">0</div>
                <div class="text-gray-400 text-sm">Passed</div>
            </div>
            <div class="bg-red-500/20 rounded-xl p-4 text-center">
                <div id="stat-failed" class="text-3xl font-bold text-red-400">0</div>
                <div class="text-gray-400 text-sm">Failed</div>
            </div>
            <div class="bg-yellow-500/20 rounded-xl p-4 text-center">
                <div id="stat-skipped" class="text-3xl font-bold text-yellow-400">0</div>
                <div class="text-gray-400 text-sm">Skipped</div>
            </div>
        </div>
        
        <!-- Test Results -->
        <div class="bg-white/10 rounded-xl p-6">
            <h2 class="text-xl font-semibold mb-4">Test Results</h2>
            <div id="test-results" class="space-y-2">
                <div class="text-gray-500 text-center py-8">
                    Click "Run All Tests" to start testing
                </div>
            </div>
        </div>
        
        <!-- JSON Output (for Claude to read) -->
        <div class="bg-white/10 rounded-xl p-6 mt-6">
            <h2 class="text-xl font-semibold mb-4">📋 Raw Results (Copy for Claude)</h2>
            <pre id="json-output" class="bg-black/50 rounded-lg p-4 text-xs text-green-400 overflow-x-auto">
{}</pre>
            <button onclick="copyResults()" class="mt-2 px-4 py-2 bg-white/10 hover:bg-white/20 rounded-lg text-sm">
                📋 Copy to Clipboard
            </button>
        </div>
    </div>
    
    <script>
        let pollInterval = null;
        
        async function runTests() {
            document.getElementById('run-btn').disabled = true;
            document.getElementById('run-btn').textContent = '⏳ Running...';
            
            await fetch('/api/run-tests', {method: 'POST'});
            startPolling();
        }
        
        async function runTest(category) {
            await fetch('/api/run-test/' + category, {method: 'POST'});
            startPolling();
        }
        
        function startPolling() {
            if (pollInterval) clearInterval(pollInterval);
            pollInterval = setInterval(updateResults, 500);
            updateResults();
        }
        
        async function updateResults() {
            const response = await fetch('/api/results');
            const data = await response.json();
            
            // Update status badge
            const badge = document.getElementById('status-badge');
            if (data.status === 'running') {
                badge.className = 'px-4 py-2 rounded-full text-sm font-medium bg-yellow-600 pulse';
                badge.textContent = '⏳ Running...';
            } else if (data.status === 'complete') {
                badge.className = 'px-4 py-2 rounded-full text-sm font-medium bg-green-600';
                badge.textContent = '✅ Complete';
                clearInterval(pollInterval);
                document.getElementById('run-btn').disabled = false;
                document.getElementById('run-btn').textContent = '▶️ Run All Tests';
            } else {
                badge.className = 'px-4 py-2 rounded-full text-sm font-medium bg-gray-700';
                badge.textContent = 'Idle';
            }
            
            // Update stats
            document.getElementById('stat-total').textContent = data.tests.length;
            document.getElementById('stat-passed').textContent = data.summary.passed;
            document.getElementById('stat-failed').textContent = data.summary.failed;
            document.getElementById('stat-skipped').textContent = data.summary.skipped;
            
            // Update test results
            const container = document.getElementById('test-results');
            if (data.tests.length > 0) {
                container.innerHTML = data.tests.map(test => {
                    let icon, color;
                    if (test.status === 'passed') { icon = '✅'; color = 'text-green-400'; }
                    else if (test.status === 'failed') { icon = '❌'; color = 'text-red-400'; }
                    else if (test.status === 'skipped') { icon = '⏭️'; color = 'text-yellow-400'; }
                    else { icon = '⏳'; color = 'text-gray-400'; }
                    
                    return `
                        <div class="flex items-start gap-3 p-3 bg-white/5 rounded-lg">
                            <span class="text-xl">${icon}</span>
                            <div class="flex-1">
                                <div class="font-medium ${color}">${test.name}</div>
                                <div class="text-sm text-gray-400">${test.category}</div>
                                ${test.error ? `<div class="text-sm text-red-400 mt-1">${test.error}</div>` : ''}
                                ${test.details ? `<div class="text-sm text-gray-500 mt-1">${test.details}</div>` : ''}
                            </div>
                            <div class="text-xs text-gray-500">${test.duration || ''}ms</div>
                        </div>
                    `;
                }).join('');
            }
            
            // Update JSON output
            document.getElementById('json-output').textContent = JSON.stringify(data, null, 2);
        }
        
        function copyResults() {
            const json = document.getElementById('json-output').textContent;
            navigator.clipboard.writeText(json);
            alert('Copied to clipboard!');
        }
        
        // Initial load
        updateResults();
    </script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)


@app.route('/api/results')
def get_results():
    return jsonify(test_results)


@app.route('/api/run-tests', methods=['POST'])
def run_all_tests():
    """Run all tests in background thread"""
    thread = threading.Thread(target=execute_all_tests)
    thread.start()
    return jsonify({'status': 'started'})


@app.route('/api/run-test/<category>', methods=['POST'])
def run_single_test(category):
    """Run single test category"""
    thread = threading.Thread(target=execute_tests, args=([category],))
    thread.start()
    return jsonify({'status': 'started'})


def add_test_result(name, category, status, error=None, details=None, duration=None):
    """Add a test result"""
    test_results['tests'].append({
        'name': name,
        'category': category,
        'status': status,
        'error': error,
        'details': details,
        'duration': duration,
        'timestamp': datetime.now().isoformat()
    })
    
    if status == 'passed':
        test_results['summary']['passed'] += 1
    elif status == 'failed':
        test_results['summary']['failed'] += 1
    elif status == 'skipped':
        test_results['summary']['skipped'] += 1


def execute_all_tests():
    """Execute all test categories"""
    execute_tests(['imports', 'dependencies', 'scraping', 'export', 'auth', 'workflow'])


def execute_tests(categories):
    """Execute specified test categories"""
    global test_results
    
    # Reset results
    test_results = {
        'status': 'running',
        'started_at': datetime.now().isoformat(),
        'completed_at': None,
        'tests': [],
        'summary': {'passed': 0, 'failed': 0, 'skipped': 0}
    }
    
    try:
        if 'imports' in categories:
            run_import_tests()
        
        if 'dependencies' in categories:
            run_dependency_tests()
        
        if 'scraping' in categories:
            run_scraping_tests()
        
        if 'export' in categories:
            run_export_tests()
        
        if 'auth' in categories:
            run_auth_tests()
        
        if 'workflow' in categories:
            run_workflow_tests()
    
    except Exception as e:
        add_test_result('Test Execution', 'system', 'failed', str(e))
    
    test_results['status'] = 'complete'
    test_results['completed_at'] = datetime.now().isoformat()


def run_import_tests():
    """Test all imports"""
    modules = [
        ('engine_v2', 'Core Engine v2'),
        ('templates', 'Scraper Templates'),
        ('ajax_handler', 'AJAX Handler'),
        ('ai_extractor', 'AI Extractor'),
        ('smart_exporter', 'Smart Exporter'),
        ('auth_scraper', 'Auth Scraper'),
        ('scheduler', 'Scheduler'),
    ]
    
    for module_name, display_name in modules:
        start = time.time()
        try:
            __import__(module_name)
            duration = int((time.time() - start) * 1000)
            add_test_result(f'Import {display_name}', 'imports', 'passed', duration=duration)
        except Exception as e:
            duration = int((time.time() - start) * 1000)
            add_test_result(f'Import {display_name}', 'imports', 'failed', str(e), duration=duration)


def run_dependency_tests():
    """Test dependencies"""
    deps = [
        ('scrapling', 'Scrapling Core'),
        ('scrapling.fetchers', 'Scrapling Fetchers'),
        ('bs4', 'BeautifulSoup'),
        ('openpyxl', 'OpenPyXL'),
        ('flask', 'Flask'),
        ('requests', 'Requests'),
    ]
    
    for module_name, display_name in deps:
        start = time.time()
        try:
            __import__(module_name)
            duration = int((time.time() - start) * 1000)
            add_test_result(f'{display_name}', 'dependencies', 'passed', duration=duration)
        except ImportError as e:
            duration = int((time.time() - start) * 1000)
            add_test_result(f'{display_name}', 'dependencies', 'failed', f'Not installed: {e}', duration=duration)
    
    # Check PlayWrightFetcher specifically
    try:
        from scrapling.fetchers import PlayWrightFetcher
        add_test_result('PlayWrightFetcher (JS rendering)', 'dependencies', 'passed')
    except ImportError:
        add_test_result('PlayWrightFetcher (JS rendering)', 'dependencies', 'failed', 
                       'Run: scrapling install')


def run_scraping_tests():
    """Test actual scraping"""
    # Test 1: Stealthy fetch
    start = time.time()
    try:
        from scrapling.fetchers import StealthyFetcher
        
        # IMPORTANT: Scrapling timeout is in MILLISECONDS
        page = StealthyFetcher.fetch("https://books.toscrape.com", timeout=30000)  # 30 seconds = 30000ms
        duration = int((time.time() - start) * 1000)
        
        if page.status == 200:
            products = page.css("article.product_pod")
            add_test_result(
                'Stealthy Fetch (books.toscrape.com)', 
                'scraping', 
                'passed',
                details=f'Status 200, found {len(products)} products',
                duration=duration
            )
        else:
            add_test_result(
                'Stealthy Fetch', 
                'scraping', 
                'failed',
                f'Status {page.status}',
                duration=duration
            )
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Stealthy Fetch', 'scraping', 'failed', str(e)[:200], duration=duration)
    
    # Test 2: Engine v2 fetch
    start = time.time()
    try:
        from engine_v2 import ScrapingEngine
        
        engine = ScrapingEngine(mode="stealthy", timeout=30)
        
        # Verify timeout is correct
        if engine.timeout_ms != 30000:
            add_test_result(
                'Engine v2 Timeout', 
                'scraping', 
                'failed',
                f'Expected 30000ms, got {engine.timeout_ms}ms'
            )
        else:
            add_test_result(
                'Engine v2 Timeout Config', 
                'scraping', 
                'passed',
                details=f'Timeout: {engine.timeout}s ({engine.timeout_ms}ms)'
            )
        
        page = engine.fetch("https://httpbin.org/get")
        duration = int((time.time() - start) * 1000)
        
        if page and page.status == 200:
            add_test_result(
                'Engine v2 Fetch (httpbin.org)', 
                'scraping', 
                'passed',
                details='Status 200',
                duration=duration
            )
        else:
            add_test_result(
                'Engine v2 Fetch', 
                'scraping', 
                'failed',
                'Page is None or bad status',
                duration=duration
            )
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Engine v2 Fetch', 'scraping', 'failed', str(e)[:200], duration=duration)


def run_export_tests():
    """Test export functionality"""
    import tempfile
    
    # Test Smart Excel Export
    start = time.time()
    try:
        from smart_exporter import SmartExcelExporter
        
        test_data = [
            {"title": "Laptop", "price": "$999", "stock": "10", "rating": "4.5"},
            {"title": "Mouse", "price": "$29", "stock": "100", "rating": "4.8"},
        ]
        
        exporter = SmartExcelExporter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        result = exporter.export(test_data, filepath, add_summary=True)
        duration = int((time.time() - start) * 1000)
        
        if result and os.path.exists(result):
            file_size = os.path.getsize(result)
            
            # Verify structure
            import openpyxl
            wb = openpyxl.load_workbook(result)
            sheets = wb.sheetnames
            
            # Check column ordering
            ws = wb.active
            headers = [cell.value for cell in ws[1] if cell.value]
            
            wb.close()
            os.unlink(result)
            
            add_test_result(
                'Smart Excel Export',
                'export',
                'passed',
                details=f'Size: {file_size} bytes, Sheets: {sheets}, Headers: {headers}',
                duration=duration
            )
        else:
            add_test_result('Smart Excel Export', 'export', 'failed', 'File not created', duration=duration)
    
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Smart Excel Export', 'export', 'failed', str(e)[:200], duration=duration)
    
    # Test CSV Export
    start = time.time()
    try:
        from engine_v2 import ScrapedItem, Exporter
        
        items = [
            ScrapedItem(url="http://test.com", title="Test", content={"price": "$10"}),
        ]
        
        with tempfile.NamedTemporaryFile(suffix='.csv', delete=False) as f:
            filepath = f.name
        
        Exporter.to_csv(items, filepath)
        duration = int((time.time() - start) * 1000)
        
        if os.path.exists(filepath):
            with open(filepath, 'r') as f:
                content = f.read()
            os.unlink(filepath)
            add_test_result('CSV Export', 'export', 'passed', details=f'{len(content)} chars', duration=duration)
        else:
            add_test_result('CSV Export', 'export', 'failed', 'File not created', duration=duration)
    
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('CSV Export', 'export', 'failed', str(e)[:200], duration=duration)


def run_auth_tests():
    """Test auth scraper components"""
    import tempfile
    
    # Test session management
    start = time.time()
    try:
        from auth_scraper import AuthenticatedScraper, SessionManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            scraper = AuthenticatedScraper(name="test_site", sessions_dir=tmpdir)
            
            # Should have no session
            is_valid = scraper.is_session_valid()
            
            duration = int((time.time() - start) * 1000)
            
            if is_valid == False:
                add_test_result(
                    'Auth Scraper Init',
                    'auth',
                    'passed',
                    details='Session correctly reports as invalid when empty',
                    duration=duration
                )
            else:
                add_test_result('Auth Scraper Init', 'auth', 'failed', 'Session should be invalid')
    
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Auth Scraper Init', 'auth', 'failed', str(e)[:200], duration=duration)
    
    # Test session manager
    start = time.time()
    try:
        from auth_scraper import SessionManager
        
        with tempfile.TemporaryDirectory() as tmpdir:
            manager = SessionManager(sessions_dir=tmpdir)
            sessions = manager.list_sessions()
            
            duration = int((time.time() - start) * 1000)
            
            add_test_result(
                'Session Manager',
                'auth',
                'passed',
                details=f'Listed {len(sessions)} sessions',
                duration=duration
            )
    
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Session Manager', 'auth', 'failed', str(e)[:200], duration=duration)


def run_workflow_tests():
    """Test complete workflow"""
    import tempfile
    
    start = time.time()
    try:
        from engine_v2 import ScrapingEngine, ScrapedItem
        from smart_exporter import SmartExcelExporter
        
        # Step 1: Fetch
        engine = ScrapingEngine(mode="stealthy", timeout=30)
        page = engine.fetch("https://books.toscrape.com")
        
        if not page:
            raise Exception("Fetch failed - page is None")
        
        # Step 2: Extract
        products = []
        for article in page.css("article.product_pod")[:5]:  # Limit to 5
            title_el = article.css("h3 a")
            price_el = article.css(".price_color")
            
            title = title_el[0].attrib.get('title', '') if title_el else 'Unknown'
            price = price_el[0].text if price_el else '$0'
            
            products.append({
                "title": title,
                "price": price,
                "category": "Books"
            })
        
        if len(products) == 0:
            raise Exception("No products extracted")
        
        # Step 3: Export
        exporter = SmartExcelExporter()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            filepath = f.name
        
        result = exporter.export(products, filepath)
        
        duration = int((time.time() - start) * 1000)
        
        if result and os.path.exists(result):
            file_size = os.path.getsize(result)
            os.unlink(result)
            
            add_test_result(
                'Full Workflow (Fetch → Extract → Export)',
                'workflow',
                'passed',
                details=f'Scraped {len(products)} products, exported {file_size} bytes',
                duration=duration
            )
        else:
            add_test_result('Full Workflow', 'workflow', 'failed', 'Export failed')
    
    except Exception as e:
        duration = int((time.time() - start) * 1000)
        add_test_result('Full Workflow', 'workflow', 'failed', str(e)[:200], duration=duration)


if __name__ == '__main__':
    print("""
    ╔═══════════════════════════════════════════════════════════════╗
    ║                                                               ║
    ║   🧪 SCRAPLING PRO - Live Test Dashboard                      ║
    ║                                                               ║
    ║   Open: http://localhost:5555                                 ║
    ║                                                               ║
    ║   Click "Run All Tests" to start                              ║
    ║   Copy the JSON results and share with Claude                 ║
    ║                                                               ║
    ╚═══════════════════════════════════════════════════════════════╝
    """)
    app.run(host='0.0.0.0', port=5555, debug=False)
