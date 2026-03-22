"""
🕷️ SCRAPLING PRO - Core Engine
================================
Professional-grade web scraping engine with:
- Concurrent/async scraping
- Proxy rotation
- Rate limiting
- Retry logic
- Session management
- Export utilities
"""

import asyncio
import csv
import json
import time
import random
import logging
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Callable, Any, List, Dict
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urljoin
import threading

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger("ScraplingPro")

# ============================================================================
# IMPORTS & AVAILABILITY
# ============================================================================

try:
    from scrapling.fetchers import StealthyFetcher
    STEALTHY_AVAILABLE = True
except ImportError:
    STEALTHY_AVAILABLE = False
    logger.warning("StealthyFetcher not available")

try:
    from scrapling.fetchers import DynamicFetcher
    DYNAMIC_AVAILABLE = True
except ImportError:
    DYNAMIC_AVAILABLE = False
    logger.warning("DynamicFetcher not available - run: scrapling install")

try:
    from scrapling.fetchers import AsyncFetcher
    ASYNC_AVAILABLE = True
except ImportError:
    ASYNC_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ScrapedItem:
    """Container for scraped data"""
    url: str
    title: str = ""
    content: Dict = field(default_factory=dict)
    metadata: Dict = field(default_factory=dict)
    scraped_at: str = field(default_factory=lambda: datetime.now().isoformat())
    success: bool = True
    error: str = ""
    
    def to_dict(self) -> Dict:
        flat = {
            "url": self.url, 
            "title": self.title, 
            "scraped_at": self.scraped_at,
            "success": self.success,
            "error": self.error,
        }
        flat.update(self.content)
        flat.update({f"meta_{k}": v for k, v in self.metadata.items()})
        return flat


@dataclass
class ProxyConfig:
    """Proxy configuration"""
    server: str
    username: str = ""
    password: str = ""
    protocol: str = "http"
    
    def to_url(self) -> str:
        if self.username and self.password:
            return f"{self.protocol}://{self.username}:{self.password}@{self.server}"
        return f"{self.protocol}://{self.server}"
    
    def to_dict(self) -> Dict:
        d = {"server": f"{self.protocol}://{self.server}"}
        if self.username:
            d["username"] = self.username
            d["password"] = self.password
        return d


# ============================================================================
# PROXY MANAGER
# ============================================================================

class ProxyManager:
    """
    Manages proxy rotation with health checking.
    
    Features:
    - Round-robin rotation
    - Random rotation
    - Weighted rotation (prefer faster proxies)
    - Health checking & auto-removal of dead proxies
    - Proxy cooldown after failures
    """
    
    def __init__(self, proxies: List[str] = None, rotation_strategy: str = "round_robin"):
        """
        Args:
            proxies: List of proxy URLs like "http://user:pass@host:port"
            rotation_strategy: "round_robin", "random", or "weighted"
        """
        self.proxies = []
        self.proxy_stats = {}  # {proxy: {"success": 0, "fail": 0, "last_used": time}}
        self.rotation_strategy = rotation_strategy
        self.current_index = 0
        self.lock = threading.Lock()
        
        if proxies:
            for p in proxies:
                self.add_proxy(p)
    
    def add_proxy(self, proxy: str):
        """Add a proxy to the pool"""
        with self.lock:
            if proxy not in self.proxies:
                self.proxies.append(proxy)
                self.proxy_stats[proxy] = {
                    "success": 0, 
                    "fail": 0, 
                    "last_used": 0,
                    "cooldown_until": 0,
                    "avg_response_time": 0,
                }
    
    def remove_proxy(self, proxy: str):
        """Remove a proxy from the pool"""
        with self.lock:
            if proxy in self.proxies:
                self.proxies.remove(proxy)
                del self.proxy_stats[proxy]
    
    def get_proxy(self) -> Optional[str]:
        """Get next proxy based on rotation strategy"""
        with self.lock:
            if not self.proxies:
                return None
            
            now = time.time()
            available = [p for p in self.proxies 
                        if self.proxy_stats[p]["cooldown_until"] < now]
            
            if not available:
                # All proxies on cooldown, return the one with shortest wait
                return min(self.proxies, 
                          key=lambda p: self.proxy_stats[p]["cooldown_until"])
            
            if self.rotation_strategy == "round_robin":
                proxy = available[self.current_index % len(available)]
                self.current_index += 1
            
            elif self.rotation_strategy == "random":
                proxy = random.choice(available)
            
            elif self.rotation_strategy == "weighted":
                # Prefer proxies with better success rates
                weights = []
                for p in available:
                    stats = self.proxy_stats[p]
                    total = stats["success"] + stats["fail"]
                    if total == 0:
                        weights.append(1.0)  # New proxy, give it a chance
                    else:
                        weights.append(stats["success"] / total)
                
                proxy = random.choices(available, weights=weights, k=1)[0]
            
            else:
                proxy = available[0]
            
            self.proxy_stats[proxy]["last_used"] = now
            return proxy
    
    def report_success(self, proxy: str, response_time: float = 0):
        """Report successful request with this proxy"""
        with self.lock:
            if proxy in self.proxy_stats:
                stats = self.proxy_stats[proxy]
                stats["success"] += 1
                if response_time > 0:
                    # Running average
                    old_avg = stats["avg_response_time"]
                    stats["avg_response_time"] = (old_avg * 0.9) + (response_time * 0.1)
    
    def report_failure(self, proxy: str, cooldown_seconds: int = 60):
        """Report failed request, put proxy on cooldown"""
        with self.lock:
            if proxy in self.proxy_stats:
                stats = self.proxy_stats[proxy]
                stats["fail"] += 1
                stats["cooldown_until"] = time.time() + cooldown_seconds
                
                # Remove proxy if too many failures
                total = stats["success"] + stats["fail"]
                if total >= 10 and (stats["fail"] / total) > 0.8:
                    logger.warning(f"Removing bad proxy: {proxy[:30]}...")
                    self.proxies.remove(proxy)
    
    def get_stats(self) -> Dict:
        """Get statistics for all proxies"""
        with self.lock:
            return {
                "total_proxies": len(self.proxies),
                "proxies": {
                    p: {
                        "success_rate": s["success"] / max(1, s["success"] + s["fail"]),
                        "total_requests": s["success"] + s["fail"],
                        "avg_response_time": s["avg_response_time"],
                    }
                    for p, s in self.proxy_stats.items()
                }
            }


# ============================================================================
# RATE LIMITER
# ============================================================================

class RateLimiter:
    """
    Token bucket rate limiter with per-domain support.
    
    Features:
    - Global rate limiting
    - Per-domain rate limiting
    - Burst allowance
    - Adaptive rate limiting based on responses
    """
    
    def __init__(
        self, 
        requests_per_second: float = 1.0,
        burst_size: int = 5,
        per_domain: bool = True,
    ):
        self.rate = requests_per_second
        self.burst_size = burst_size
        self.per_domain = per_domain
        
        self.tokens = burst_size
        self.last_update = time.time()
        self.domain_limiters = {}
        self.lock = threading.Lock()
    
    def _get_domain(self, url: str) -> str:
        """Extract domain from URL"""
        try:
            return urlparse(url).netloc
        except:
            return "default"
    
    def _refill_tokens(self, limiter_state: Dict) -> float:
        """Refill tokens based on elapsed time"""
        now = time.time()
        elapsed = now - limiter_state["last_update"]
        limiter_state["tokens"] = min(
            self.burst_size,
            limiter_state["tokens"] + elapsed * self.rate
        )
        limiter_state["last_update"] = now
        return limiter_state["tokens"]
    
    def acquire(self, url: str = "") -> float:
        """
        Acquire permission to make a request.
        Returns the time to wait (0 if immediate).
        """
        with self.lock:
            if self.per_domain and url:
                domain = self._get_domain(url)
                if domain not in self.domain_limiters:
                    self.domain_limiters[domain] = {
                        "tokens": self.burst_size,
                        "last_update": time.time(),
                    }
                state = self.domain_limiters[domain]
            else:
                state = {"tokens": self.tokens, "last_update": self.last_update}
            
            tokens = self._refill_tokens(state)
            
            if tokens >= 1:
                state["tokens"] -= 1
                if not self.per_domain:
                    self.tokens = state["tokens"]
                    self.last_update = state["last_update"]
                return 0
            else:
                wait_time = (1 - tokens) / self.rate
                return wait_time
    
    def wait(self, url: str = ""):
        """Wait until we can make a request"""
        wait_time = self.acquire(url)
        if wait_time > 0:
            time.sleep(wait_time)


# ============================================================================
# SCRAPING ENGINE
# ============================================================================

class ScrapingEngine:
    """
    Professional scraping engine with all features.
    
    Features:
    - Multiple fetcher modes
    - Proxy rotation
    - Rate limiting
    - Retry logic with exponential backoff
    - Concurrent scraping
    - Session management
    - Adaptive selectors
    """
    
    def __init__(
        self,
        mode: str = "stealthy",
        proxies: List[str] = None,
        proxy_rotation: str = "round_robin",
        rate_limit: float = 1.0,
        max_retries: int = 3,
        timeout: int = 30,
        headless: bool = True,
        adaptive: bool = False,
        max_workers: int = 5,
    ):
        self.mode = mode
        self.headless = headless
        self.timeout = timeout
        self.max_retries = max_retries
        self.adaptive = adaptive
        self.max_workers = max_workers
        
        # Managers
        self.proxy_manager = ProxyManager(proxies, proxy_rotation) if proxies else None
        self.rate_limiter = RateLimiter(rate_limit)
        
        # Results storage
        self.results: List[ScrapedItem] = []
        self.stats = {
            "total_requests": 0,
            "successful": 0,
            "failed": 0,
            "start_time": None,
            "end_time": None,
        }
        
        # Enable adaptive mode if requested
        if adaptive and STEALTHY_AVAILABLE:
            StealthyFetcher.adaptive = True
    
    def _get_fetcher_kwargs(self) -> Dict:
        """Build kwargs for fetcher"""
        kwargs = {"timeout": self.timeout}
        
        if self.proxy_manager:
            proxy = self.proxy_manager.get_proxy()
            if proxy:
                kwargs["proxy"] = proxy
        
        return kwargs
    
    def fetch(self, url: str, **extra_kwargs) -> Any:
        """
        Fetch a single URL with retry logic.
        """
        self.stats["total_requests"] += 1
        
        for attempt in range(self.max_retries):
            try:
                # Rate limiting
                self.rate_limiter.wait(url)
                
                # Get proxy and build kwargs
                kwargs = self._get_fetcher_kwargs()
                kwargs.update(extra_kwargs)
                proxy_used = kwargs.get("proxy")
                
                # Time the request
                start_time = time.time()
                
                # Fetch based on mode
                if self.mode == "dynamic" and DYNAMIC_AVAILABLE:
                    page = DynamicFetcher.fetch(
                        url,
                        headless=self.headless,
                        timeout=self.timeout * 1000,
                        **{k: v for k, v in kwargs.items() if k not in ["timeout"]}
                    )
                else:
                    page = StealthyFetcher.fetch(url, **kwargs)
                
                response_time = time.time() - start_time
                
                # Check success
                if page.status == 200:
                    self.stats["successful"] += 1
                    if proxy_used and self.proxy_manager:
                        self.proxy_manager.report_success(proxy_used, response_time)
                    return page
                
                elif page.status == 429:  # Rate limited
                    wait_time = (attempt + 1) * 30
                    logger.warning(f"Rate limited on {url}, waiting {wait_time}s")
                    time.sleep(wait_time)
                
                else:
                    logger.warning(f"HTTP {page.status} for {url}")
                    if proxy_used and self.proxy_manager:
                        self.proxy_manager.report_failure(proxy_used)
                        
            except Exception as e:
                logger.error(f"Attempt {attempt + 1} failed for {url}: {e}")
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
        
        self.stats["failed"] += 1
        return None
    
    def scrape_url(
        self, 
        url: str, 
        parser: Callable,
        **fetch_kwargs
    ) -> List[ScrapedItem]:
        """
        Scrape a single URL using the provided parser function.
        
        Args:
            url: URL to scrape
            parser: Function that takes a page and returns List[ScrapedItem]
            **fetch_kwargs: Additional kwargs for the fetcher
        """
        page = self.fetch(url, **fetch_kwargs)
        
        if not page:
            return [ScrapedItem(url=url, success=False, error="Failed to fetch")]
        
        try:
            items = parser(page)
            self.results.extend(items)
            return items
        except Exception as e:
            logger.error(f"Parser error for {url}: {e}")
            return [ScrapedItem(url=url, success=False, error=str(e))]
    
    def scrape_urls(
        self,
        urls: List[str],
        parser: Callable,
        concurrent: bool = True,
        **fetch_kwargs
    ) -> List[ScrapedItem]:
        """
        Scrape multiple URLs, optionally in parallel.
        
        Args:
            urls: List of URLs to scrape
            parser: Function that takes a page and returns List[ScrapedItem]
            concurrent: Whether to use thread pool
            **fetch_kwargs: Additional kwargs for the fetcher
        """
        self.stats["start_time"] = datetime.now()
        all_items = []
        
        if concurrent and len(urls) > 1:
            logger.info(f"Scraping {len(urls)} URLs with {self.max_workers} workers")
            
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {
                    executor.submit(self.scrape_url, url, parser, **fetch_kwargs): url
                    for url in urls
                }
                
                for future in as_completed(futures):
                    url = futures[future]
                    try:
                        items = future.result()
                        all_items.extend(items)
                        logger.info(f"✓ {url} - {len(items)} items")
                    except Exception as e:
                        logger.error(f"✗ {url} - {e}")
                        all_items.append(ScrapedItem(url=url, success=False, error=str(e)))
        else:
            for url in urls:
                logger.info(f"Scraping: {url}")
                items = self.scrape_url(url, parser, **fetch_kwargs)
                all_items.extend(items)
        
        self.stats["end_time"] = datetime.now()
        return all_items
    
    def scrape_with_pagination(
        self,
        start_url: str,
        parser: Callable,
        pagination_type: str = "url",
        url_pattern: str = "?page={}",
        next_selector: str = "a.next",
        max_pages: int = 10,
        **fetch_kwargs
    ) -> List[ScrapedItem]:
        """
        Scrape with automatic pagination handling.
        
        Args:
            start_url: Starting URL
            parser: Parser function
            pagination_type: "url", "click", or "scroll"
            url_pattern: Pattern for URL pagination
            next_selector: CSS selector for next button
            max_pages: Maximum pages to scrape
        """
        self.stats["start_time"] = datetime.now()
        all_items = []
        
        if pagination_type == "url":
            for page_num in range(1, max_pages + 1):
                url = start_url + url_pattern.format(page_num)
                logger.info(f"Page {page_num}: {url}")
                
                items = self.scrape_url(url, parser, **fetch_kwargs)
                if not items or (len(items) == 1 and not items[0].success):
                    logger.info("No more items, stopping pagination")
                    break
                
                all_items.extend(items)
        
        elif pagination_type == "scroll" and DYNAMIC_AVAILABLE:
            scroll_js = f"""
                let count = 0;
                const maxScrolls = {max_pages};
                const interval = setInterval(() => {{
                    window.scrollTo(0, document.body.scrollHeight);
                    count++;
                    if (count >= maxScrolls) clearInterval(interval);
                }}, 2000);
            """
            
            page = self.fetch(
                start_url,
                js_code=scroll_js,
                wait=max_pages * 2500,
                **fetch_kwargs
            )
            
            if page:
                all_items = parser(page)
                self.results.extend(all_items)
        
        self.stats["end_time"] = datetime.now()
        return all_items
    
    def get_stats(self) -> Dict:
        """Get scraping statistics"""
        stats = self.stats.copy()
        
        if stats["start_time"] and stats["end_time"]:
            duration = (stats["end_time"] - stats["start_time"]).total_seconds()
            stats["duration_seconds"] = duration
            stats["requests_per_second"] = stats["total_requests"] / max(1, duration)
        
        if self.proxy_manager:
            stats["proxy_stats"] = self.proxy_manager.get_stats()
        
        return stats


# ============================================================================
# EXPORT UTILITIES
# ============================================================================

class Exporter:
    """Multi-format data exporter"""
    
    @staticmethod
    def to_csv(items: List[ScrapedItem], filepath: str):
        if not items:
            logger.warning("No items to export")
            return
        
        flat_items = [item.to_dict() for item in items]
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=sorted(all_keys))
            writer.writeheader()
            writer.writerows(flat_items)
        
        logger.info(f"Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_json(items: List[ScrapedItem], filepath: str, indent: int = 2):
        if not items:
            logger.warning("No items to export")
            return
        
        data = [item.to_dict() for item in items]
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, ensure_ascii=False, default=str)
        
        logger.info(f"Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_excel(items: List[ScrapedItem], filepath: str):
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return
        
        if not items:
            logger.warning("No items to export")
            return
        
        flat_items = [item.to_dict() for item in items]
        all_keys = set()
        for item in flat_items:
            all_keys.update(item.keys())
        fieldnames = sorted(all_keys)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"
        
        # Header styling
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Data rows
        for row_num, item in enumerate(flat_items, 2):
            for col, field in enumerate(fieldnames, 1):
                value = item.get(field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(50, max_length + 2)
        
        wb.save(filepath)
        logger.info(f"Exported {len(items)} items to {filepath}")


# ============================================================================
# TEMPLATE PARSERS
# ============================================================================

class Parsers:
    """Ready-to-use parser templates"""
    
    @staticmethod
    def ecommerce_products(page) -> List[ScrapedItem]:
        """Parse e-commerce product listings"""
        items = []
        
        selectors = [
            "article.product_pod", ".product-card", ".product-item",
            ".product", "[data-product]", ".item"
        ]
        
        for selector in selectors:
            products = page.css(selector)
            if products:
                for product in products:
                    title_el = product.css("h2, h3, .title, .name, h3 a")
                    price_el = product.css(".price, .price_color, .cost")
                    img_el = product.css("img")
                    rating_el = product.css(".rating, .star-rating, [class*='star']")
                    link_el = product.css("a")
                    
                    title = ""
                    if title_el:
                        title = title_el[0].attrib.get("title") or title_el[0].text or ""
                    
                    items.append(ScrapedItem(
                        url=str(page.url) if hasattr(page, 'url') else "",
                        title=title.strip(),
                        content={
                            "price": price_el[0].text.strip() if price_el else "",
                            "image": img_el[0].attrib.get("src", "") if img_el else "",
                            "rating": rating_el[0].attrib.get("class", "").split()[-1] if rating_el else "",
                            "link": link_el[0].attrib.get("href", "") if link_el else "",
                        }
                    ))
                break
        
        return items
    
    @staticmethod
    def news_articles(page) -> List[ScrapedItem]:
        """Parse news article listings"""
        items = []
        
        for article in page.css("article, .article, .post, .story, .news-item"):
            title_el = article.css("h1, h2, h3, .title, .headline")
            summary_el = article.css("p, .summary, .excerpt, .description")
            author_el = article.css(".author, .byline, [rel='author']")
            date_el = article.css("time, .date, .published")
            link_el = article.css("a")
            
            items.append(ScrapedItem(
                url=str(page.url) if hasattr(page, 'url') else "",
                title=title_el[0].text.strip() if title_el else "",
                content={
                    "summary": summary_el[0].text.strip()[:200] if summary_el else "",
                    "author": author_el[0].text.strip() if author_el else "",
                    "date": date_el[0].attrib.get("datetime") or (date_el[0].text.strip() if date_el else ""),
                    "link": link_el[0].attrib.get("href", "") if link_el else "",
                }
            ))
        
        return items
    
    @staticmethod
    def job_listings(page) -> List[ScrapedItem]:
        """Parse job listings"""
        items = []
        
        for job in page.css(".job, .job-card, .job-listing, .position, [data-job]"):
            title_el = job.css("h2, h3, .title, .job-title")
            company_el = job.css(".company, .employer, .organization")
            location_el = job.css(".location, .job-location")
            salary_el = job.css(".salary, .compensation, .pay")
            link_el = job.css("a")
            
            items.append(ScrapedItem(
                url=str(page.url) if hasattr(page, 'url') else "",
                title=title_el[0].text.strip() if title_el else "",
                content={
                    "company": company_el[0].text.strip() if company_el else "",
                    "location": location_el[0].text.strip() if location_el else "",
                    "salary": salary_el[0].text.strip() if salary_el else "",
                    "link": link_el[0].attrib.get("href", "") if link_el else "",
                }
            ))
        
        return items
    
    @staticmethod
    def generic_links(page) -> List[ScrapedItem]:
        """Extract all links from a page"""
        items = []
        
        for link in page.css("a[href]"):
            href = link.attrib.get("href", "")
            text = link.text or ""
            
            if href and not href.startswith(("#", "javascript:", "mailto:")):
                items.append(ScrapedItem(
                    url=str(page.url) if hasattr(page, 'url') else "",
                    title=text.strip(),
                    content={"href": href}
                ))
        
        return items


# ============================================================================
# QUICK API
# ============================================================================

def quick_scrape(
    url: str,
    item_selector: str,
    fields: Dict[str, str],
    max_pages: int = 1,
    url_pattern: str = "",
    **engine_kwargs
) -> List[ScrapedItem]:
    """
    Quick one-liner scraping API.
    
    Example:
        items = quick_scrape(
            url="https://books.toscrape.com",
            item_selector="article.product_pod",
            fields={
                "title": "h3 a @title",
                "price": ".price_color",
                "image": "img @src",
            }
        )
    
    Field syntax:
        - "selector" → get text content
        - "selector @attr" → get attribute value
    """
    
    def parser(page):
        items = []
        for element in page.css(item_selector):
            content = {}
            title = ""
            
            for field_name, selector in fields.items():
                # Check for attribute extraction
                if " @" in selector:
                    sel, attr = selector.rsplit(" @", 1)
                    els = element.css(sel)
                    value = els[0].attrib.get(attr, "") if els else ""
                else:
                    els = element.css(selector)
                    value = els[0].text.strip() if els else ""
                
                if field_name == "title":
                    title = value
                else:
                    content[field_name] = value
            
            items.append(ScrapedItem(
                url=str(page.url) if hasattr(page, 'url') else url,
                title=title,
                content=content
            ))
        
        return items
    
    engine = ScrapingEngine(**engine_kwargs)
    
    if max_pages > 1 and url_pattern:
        return engine.scrape_with_pagination(
            url, parser, "url", url_pattern, max_pages=max_pages
        )
    else:
        return engine.scrape_url(url, parser)


# ============================================================================
# MAIN - DEMO
# ============================================================================

if __name__ == "__main__":
    print("🕷️ Scrapling Pro - Core Engine")
    print("=" * 50)
    
    # Quick demo
    items = quick_scrape(
        url="https://books.toscrape.com",
        item_selector="article.product_pod",
        fields={
            "title": "h3 a @title",
            "price": ".price_color",
            "rating": ".star-rating @class",
        }
    )
    
    print(f"\nScraped {len(items)} items:")
    for item in items[:3]:
        print(f"  - {item.title}: {item.content.get('price', 'N/A')}")
    
    # Export
    Exporter.to_json(items, "demo_output.json")
