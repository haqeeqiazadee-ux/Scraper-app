"""
📊 SCRAPLING PRO - Intelligent Excel Exporter
==============================================
AI-powered Excel export that:
1. Automatically detects data types
2. Organizes columns intelligently
3. Groups related data together
4. Formats cells appropriately (currency, dates, etc.)
5. Creates summary sheets with analytics

Features:
- Smart column ordering (Title → Price → Quantity → Details)
- Auto-formatting (currency, percentages, dates)
- Data validation and cleaning
- Multiple sheets for different data types
- Summary statistics and charts
- Conditional formatting (highlight deals, low stock, etc.)

Usage:
    from smart_exporter import SmartExcelExporter
    
    exporter = SmartExcelExporter()
    exporter.export(scraped_items, "products.xlsx")
"""

import re
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass
from collections import defaultdict

logger = logging.getLogger("SmartExporter")


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ColumnConfig:
    """Configuration for a column"""
    name: str
    display_name: str
    data_type: str  # "text", "currency", "number", "percentage", "date", "url", "image"
    width: int = 15
    priority: int = 50  # Lower = appears first
    format_code: str = None
    

# ============================================================================
# SMART EXCEL EXPORTER
# ============================================================================

class SmartExcelExporter:
    """
    Intelligent Excel exporter that understands your data.
    
    - Automatically detects what each column contains
    - Orders columns logically (product name first, then price, etc.)
    - Formats currencies, dates, percentages correctly
    - Groups related items together
    - Adds summary statistics
    
    Usage:
        exporter = SmartExcelExporter()
        
        # Basic export
        exporter.export(items, "output.xlsx")
        
        # With options
        exporter.export(
            items,
            "output.xlsx",
            sheet_name="Products",
            add_summary=True,
            add_charts=True,
            group_by="category"
        )
    """
    
    # Column priority (lower = appears first)
    COLUMN_PRIORITIES = {
        # Identifiers (leftmost)
        'id': 1, 'sku': 2, 'product_id': 3, 'item_id': 4,
        
        # Core product info
        'name': 10, 'title': 10, 'product_name': 10, 'product_title': 10,
        'brand': 15, 'manufacturer': 16,
        'category': 20, 'type': 21, 'product_type': 22,
        
        # Pricing (important, near left)
        'price': 30, 'cost': 31, 'sale_price': 32, 'original_price': 33,
        'regular_price': 34, 'discount': 35, 'discount_percent': 36,
        'currency': 37,
        
        # Availability
        'stock': 40, 'quantity': 41, 'inventory': 42, 'in_stock': 43,
        'availability': 44, 'stock_status': 45,
        
        # Ratings & Reviews
        'rating': 50, 'stars': 51, 'score': 52,
        'reviews': 55, 'review_count': 56, 'ratings_count': 57,
        
        # Description (longer text, after key data)
        'description': 60, 'short_description': 61, 'summary': 62,
        'features': 65, 'specifications': 66, 'specs': 67,
        
        # Media
        'image': 70, 'image_url': 71, 'thumbnail': 72, 'images': 73,
        'video': 75, 'video_url': 76,
        
        # Links
        'url': 80, 'link': 81, 'product_url': 82,
        
        # Metadata (rightmost)
        'scraped_at': 90, 'created_at': 91, 'updated_at': 92,
        'source': 95, 'source_url': 96,
    }
    
    # Data type detection patterns
    TYPE_PATTERNS = {
        'currency': [
            r'^\$[\d,]+\.?\d*$',  # $29.99
            r'^£[\d,]+\.?\d*$',   # £29.99
            r'^€[\d,]+\.?\d*$',   # €29.99
            r'^[\d,]+\.?\d*\s*(USD|GBP|EUR|INR|PKR)$',  # 29.99 USD
            r'^\d+\.?\d*$',  # Plain number (for price columns)
        ],
        'percentage': [
            r'^\d+\.?\d*\s*%$',  # 25%
            r'^\d+\.?\d*\s*percent$',  # 25 percent
        ],
        'date': [
            r'^\d{4}-\d{2}-\d{2}',  # 2024-01-15
            r'^\d{2}/\d{2}/\d{4}',  # 01/15/2024
            r'^\d{2}-\d{2}-\d{4}',  # 15-01-2024
        ],
        'url': [
            r'^https?://',  # http:// or https://
            r'^www\.',  # www.
        ],
        'rating': [
            r'^\d\.?\d?\s*/\s*\d+$',  # 4.5/5
            r'^\d\.?\d?\s*stars?$',  # 4.5 stars
        ],
        'number': [
            r'^[\d,]+$',  # 1,234
            r'^[\d,]+\.?\d*$',  # 1,234.56
        ],
    }
    
    # Currency symbols by locale
    CURRENCY_FORMATS = {
        '$': '"$"#,##0.00',
        '£': '"£"#,##0.00',
        '€': '"€"#,##0.00',
        '₹': '"₹"#,##0.00',
        'PKR': '"PKR "#,##0.00',
        'default': '#,##0.00',
    }
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if openpyxl is available"""
        try:
            import openpyxl
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            logger.warning("openpyxl not installed. Run: pip install openpyxl")
    
    def export(
        self,
        items: List[Any],
        filepath: str,
        sheet_name: str = "Data",
        add_summary: bool = True,
        add_charts: bool = False,
        group_by: str = None,
        highlight_deals: bool = True,
        auto_filter: bool = True,
        freeze_header: bool = True,
    ) -> str:
        """
        Export items to a smart Excel file.
        
        Args:
            items: List of scraped items (dicts or ScrapedItem objects)
            filepath: Output file path (.xlsx)
            sheet_name: Name for the data sheet
            add_summary: Add a summary sheet with statistics
            add_charts: Add charts for numeric data
            group_by: Column name to group data by
            highlight_deals: Highlight items with discounts
            auto_filter: Add filter dropdowns to headers
            freeze_header: Freeze the header row
            
        Returns:
            Path to the created file
        """
        if not self.openpyxl_available:
            # Fallback to CSV
            return self._export_csv_fallback(items, filepath)
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
        
        # Convert items to dicts
        data = self._normalize_items(items)
        
        if not data:
            logger.warning("No data to export")
            return None
        
        # Analyze and prepare columns
        columns = self._analyze_columns(data)
        
        # Sort columns by priority
        columns = sorted(columns, key=lambda c: c.priority)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        currency_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        url_font = Font(color="0563C1", underline="single")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_idx, col_config in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_config.display_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
            # Set column width
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config.width
        
        # Write data
        for row_idx, item in enumerate(data, 2):
            for col_idx, col_config in enumerate(columns, 1):
                value = item.get(col_config.name, "")
                
                # Clean and convert value
                value, cell_format = self._process_value(value, col_config)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Apply formatting based on data type
                if col_config.data_type == "currency":
                    cell.number_format = col_config.format_code or self.CURRENCY_FORMATS['default']
                    cell.alignment = Alignment(horizontal="right")
                
                elif col_config.data_type == "percentage":
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                
                elif col_config.data_type == "number":
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                
                elif col_config.data_type == "url":
                    cell.font = url_font
                    if value and isinstance(value, str) and value.startswith("http"):
                        cell.hyperlink = value
                        cell.value = value  # Show actual URL, not "View Link"
                
                elif col_config.data_type == "date":
                    cell.number_format = "YYYY-MM-DD"
                
                else:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Freeze header row
        if freeze_header:
            ws.freeze_panes = "A2"
        
        # Add auto-filter
        if auto_filter and data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(data) + 1}"
        
        # Add conditional formatting for deals
        if highlight_deals:
            self._add_deal_highlighting(ws, columns, len(data))
        
        # Add summary sheet
        if add_summary:
            self._add_summary_sheet(wb, data, columns)
        
        # Save
        filepath = str(filepath)
        if not filepath.endswith('.xlsx'):
            filepath += '.xlsx'
        
        wb.save(filepath)
        logger.info(f"Exported {len(data)} items to {filepath}")
        
        return filepath
    
    def export_grouped(
        self,
        items: List[Any],
        filepath: str,
        group_by: str,
    ) -> str:
        """
        Export with separate sheets for each group.
        
        Example: Group products by category, each category gets its own sheet.
        """
        if not self.openpyxl_available:
            return self._export_csv_fallback(items, filepath)
        
        import openpyxl
        
        data = self._normalize_items(items)
        
        # Group data
        groups = defaultdict(list)
        for item in data:
            group_value = item.get(group_by, "Other")
            groups[str(group_value)[:30]].append(item)  # Sheet names max 31 chars
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        for group_name, group_items in groups.items():
            ws = wb.create_sheet(title=group_name)
            self._write_sheet(ws, group_items)
        
        # Add summary sheet
        summary = wb.create_sheet(title="Summary", index=0)
        self._write_group_summary(summary, groups)
        
        wb.save(filepath)
        return filepath
    
    def _normalize_items(self, items: List[Any]) -> List[Dict]:
        """Convert items to list of dicts"""
        data = []
        
        for item in items:
            if hasattr(item, 'to_dict'):
                d = item.to_dict()
            elif isinstance(item, dict):
                d = item
            else:
                d = {'value': str(item)}
            
            # Flatten nested dicts
            flat = {}
            for key, value in d.items():
                if isinstance(value, dict):
                    for k, v in value.items():
                        flat[f"{key}_{k}"] = v
                else:
                    flat[key] = value
            
            data.append(flat)
        
        return data
    
    def _analyze_columns(self, data: List[Dict]) -> List[ColumnConfig]:
        """Analyze data and create column configurations"""
        columns = {}
        
        for item in data:
            for key, value in item.items():
                if key not in columns:
                    columns[key] = {
                        'values': [],
                        'non_empty': 0,
                    }
                
                columns[key]['values'].append(value)
                if value not in [None, '', 'null', 'None']:
                    columns[key]['non_empty'] += 1
        
        # Create column configs
        configs = []
        
        for name, info in columns.items():
            # Skip mostly empty columns
            if info['non_empty'] < len(data) * 0.1:
                continue
            
            # Detect data type
            data_type = self._detect_column_type(name, info['values'])
            
            # Get priority
            name_lower = name.lower().replace(' ', '_')
            priority = self.COLUMN_PRIORITIES.get(name_lower, 50)
            
            # Detect currency format
            format_code = None
            if data_type == 'currency':
                format_code = self._detect_currency_format(info['values'])
            
            # Calculate width
            width = self._calculate_column_width(name, info['values'], data_type)
            
            # Create display name
            display_name = self._create_display_name(name)
            
            configs.append(ColumnConfig(
                name=name,
                display_name=display_name,
                data_type=data_type,
                width=width,
                priority=priority,
                format_code=format_code,
            ))
        
        return configs
    
    def _detect_column_type(self, name: str, values: List) -> str:
        """Detect the data type of a column"""
        name_lower = name.lower()
        
        # Check column name hints
        if any(x in name_lower for x in ['price', 'cost', 'amount', 'total', 'fee']):
            return 'currency'
        if any(x in name_lower for x in ['percent', 'rate', 'ratio', 'discount']):
            return 'percentage'
        if any(x in name_lower for x in ['date', 'time', 'created', 'updated', 'scraped']):
            return 'date'
        if any(x in name_lower for x in ['url', 'link', 'href']):
            return 'url'
        if any(x in name_lower for x in ['image', 'img', 'photo', 'picture', 'thumbnail']):
            return 'url'
        if any(x in name_lower for x in ['rating', 'stars', 'score']):
            return 'number'
        if any(x in name_lower for x in ['count', 'quantity', 'stock', 'inventory', 'reviews']):
            return 'number'
        
        # Check value patterns
        sample_values = [str(v) for v in values if v not in [None, '', 'null']][:10]
        
        for data_type, patterns in self.TYPE_PATTERNS.items():
            matches = 0
            for value in sample_values:
                for pattern in patterns:
                    if re.match(pattern, value.strip(), re.IGNORECASE):
                        matches += 1
                        break
            
            if matches >= len(sample_values) * 0.7:  # 70% match
                return data_type
        
        return 'text'
    
    def _detect_currency_format(self, values: List) -> str:
        """Detect currency symbol from values"""
        for value in values:
            if value is None:
                continue
            value_str = str(value).strip()
            
            for symbol, format_code in self.CURRENCY_FORMATS.items():
                if symbol in value_str:
                    return format_code
        
        return self.CURRENCY_FORMATS['default']
    
    def _calculate_column_width(self, name: str, values: List, data_type: str) -> int:
        """Calculate appropriate column width"""
        # Base width from column name
        base_width = len(name) + 2
        
        # Check value lengths
        max_value_width = 0
        for value in values[:50]:  # Sample first 50
            if value:
                max_value_width = max(max_value_width, len(str(value)))
        
        # Apply limits based on data type
        if data_type == 'url':
            return 15  # Show "View Link" instead
        elif data_type == 'currency':
            return max(12, base_width)
        elif data_type in ['description', 'text']:
            return min(50, max(base_width, max_value_width // 2))
        else:
            return min(30, max(base_width, max_value_width + 2))
    
    def _create_display_name(self, name: str) -> str:
        """Create human-readable column header"""
        # Convert snake_case to Title Case
        display = name.replace('_', ' ').replace('-', ' ')
        display = ' '.join(word.capitalize() for word in display.split())
        return display
    
    def _process_value(self, value: Any, col_config: ColumnConfig) -> Tuple[Any, str]:
        """Process and clean a value for the cell"""
        if value is None or value == '' or value == 'null' or value == 'None':
            return '', None
        
        value_str = str(value).strip()
        
        if col_config.data_type == 'currency':
            # Extract numeric value
            clean = re.sub(r'[^\d.,]', '', value_str)
            clean = clean.replace(',', '')
            try:
                return float(clean), col_config.format_code
            except:
                return value_str, None
        
        elif col_config.data_type == 'percentage':
            # Extract percentage value
            clean = re.sub(r'[^\d.]', '', value_str)
            try:
                return float(clean) / 100, None
            except:
                return value_str, None
        
        elif col_config.data_type == 'number':
            clean = re.sub(r'[^\d.,]', '', value_str)
            clean = clean.replace(',', '')
            try:
                if '.' in clean:
                    return float(clean), None
                return int(clean), None
            except:
                return value_str, None
        
        return value_str, None
    
    def _add_deal_highlighting(self, ws, columns: List[ColumnConfig], data_count: int):
        """Add conditional formatting to highlight deals"""
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import PatternFill
        
        # Find discount column
        discount_col = None
        for idx, col in enumerate(columns, 1):
            if 'discount' in col.name.lower() or 'sale' in col.name.lower():
                discount_col = idx
                break
        
        if discount_col:
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(discount_col)
            
            # Highlight rows with discounts
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            ws.conditional_formatting.add(
                f"A2:Z{data_count + 1}",
                FormulaRule(
                    formula=[f'${col_letter}2>0'],
                    fill=green_fill
                )
            )
    
    def _add_summary_sheet(self, wb, data: List[Dict], columns: List[ColumnConfig]):
        """Add a summary sheet with statistics"""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        ws = wb.create_sheet(title="Summary")
        
        # Title
        ws['A1'] = "📊 Data Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Basic stats
        ws['A3'] = "Total Items:"
        ws['B3'] = len(data)
        ws['B3'].font = Font(bold=True)
        
        ws['A4'] = "Export Date:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Column statistics
        row = 6
        ws[f'A{row}'] = "Column Statistics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Headers
        headers = ['Column', 'Type', 'Non-Empty', 'Min', 'Max', 'Avg']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
        
        # Stats for each column
        for col_config in columns:
            values = [item.get(col_config.name) for item in data]
            non_empty = sum(1 for v in values if v not in [None, '', 'null'])
            
            ws.cell(row=row, column=1, value=col_config.display_name)
            ws.cell(row=row, column=2, value=col_config.data_type)
            ws.cell(row=row, column=3, value=f"{non_empty}/{len(data)}")
            
            # Numeric stats
            if col_config.data_type in ['currency', 'number', 'percentage']:
                numeric_values = []
                for v in values:
                    try:
                        clean = re.sub(r'[^\d.,]', '', str(v))
                        clean = clean.replace(',', '')
                        if clean:
                            numeric_values.append(float(clean))
                    except:
                        pass
                
                if numeric_values:
                    ws.cell(row=row, column=4, value=min(numeric_values))
                    ws.cell(row=row, column=5, value=max(numeric_values))
                    ws.cell(row=row, column=6, value=round(sum(numeric_values) / len(numeric_values), 2))
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
    
    def _write_sheet(self, ws, data: List[Dict]):
        """Write data to a worksheet"""
        if not data:
            return
        
        from openpyxl.styles import Font, Alignment, PatternFill
        
        columns = self._analyze_columns(data)
        columns = sorted(columns, key=lambda c: c.priority)
        
        # Headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col.display_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        
        # Data
        for row_idx, item in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                value = item.get(col.name, "")
                value, _ = self._process_value(value, col)
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _write_group_summary(self, ws, groups: Dict[str, List]):
        """Write group summary"""
        from openpyxl.styles import Font, PatternFill
        
        ws['A1'] = "Group Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        headers = ['Group', 'Count']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for row, (group_name, items) in enumerate(groups.items(), 4):
            ws.cell(row=row, column=1, value=group_name)
            ws.cell(row=row, column=2, value=len(items))
    
    def _export_csv_fallback(self, items: List[Any], filepath: str) -> str:
        """Fallback to CSV if openpyxl not available"""
        import csv
        
        data = self._normalize_items(items)
        
        if not data:
            return None
        
        filepath = filepath.replace('.xlsx', '.csv')
        
        # Get all columns
        all_columns = set()
        for item in data:
            all_columns.update(item.keys())
        
        columns = sorted(all_columns, key=lambda c: self.COLUMN_PRIORITIES.get(c.lower(), 50))
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"Exported {len(data)} items to {filepath} (CSV fallback)")
        return filepath


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def smart_export(items: List[Any], filepath: str, **kwargs) -> str:
    """Quick export with smart formatting"""
    exporter = SmartExcelExporter()
    return exporter.export(items, filepath, **kwargs)


def export_with_summary(items: List[Any], filepath: str) -> str:
    """Export with summary statistics"""
    exporter = SmartExcelExporter()
    return exporter.export(items, filepath, add_summary=True, highlight_deals=True)


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("📊 Scrapling Pro - Intelligent Excel Exporter")
    print("=" * 50)
    print("""
Features:
    ✅ Auto-detects data types (currency, dates, URLs, etc.)
    ✅ Smart column ordering (Title → Price → Quantity → Details)
    ✅ Proper formatting (currency symbols, percentages)
    ✅ Summary sheet with statistics
    ✅ Conditional formatting (highlights deals)
    ✅ Auto-filter enabled
    ✅ Frozen header row

Usage:

    from smart_exporter import SmartExcelExporter, smart_export
    
    # Quick export
    smart_export(scraped_items, "products.xlsx")
    
    # With options
    exporter = SmartExcelExporter()
    exporter.export(
        items=scraped_items,
        filepath="products.xlsx",
        sheet_name="Products",
        add_summary=True,
        highlight_deals=True
    )
    
    # Group by category (separate sheets)
    exporter.export_grouped(
        items=scraped_items,
        filepath="products_by_category.xlsx",
        group_by="category"
    )

Column Order (automatic):
    1. ID / SKU
    2. Name / Title
    3. Brand
    4. Category
    5. Price (with currency formatting)
    6. Stock / Quantity
    7. Rating / Reviews
    8. Description
    9. Image URL
    10. Product URL
    11. Metadata (scraped_at, source, etc.)
    """)
