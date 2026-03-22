"""
🕷️ SCRAPLING PRO - Enhanced Engine v2
======================================
Fixed and enhanced scraping engine with:
- Proper timeout handling (seconds → milliseconds conversion)
- Infinite scroll support
- CAPTCHA solving integration
- robots.txt compliance
- Webhook support
- Scheduled execution
"""

import asyncio
import csv
import json
import time
import random
import logging
import re
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Callable, Any, List, Dict, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
from urllib.parse import urlparse, urljoin
from urllib.robotparser import RobotFileParser
import threading

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s | %(levelname)s | %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger("ScraplingPro")

# ============================================================================
# IMPORTS & AVAILABILITY
# ============================================================================

STEALTHY_AVAILABLE = False
DYNAMIC_AVAILABLE = False
ASYNC_AVAILABLE = False
EXCEL_AVAILABLE = False

try:
    from scrapling.fetchers import StealthyFetcher
    STEALTHY_AVAILABLE = True
except ImportError:
    logger.warning("StealthyFetcher not available - run: pip install scrapling[all]")

try:
    from scrapling.fetchers import PlayWrightFetcher
    DYNAMIC_AVAILABLE = True
    DynamicFetcher = PlayWrightFetcher
except ImportError:
    try:
        from scrapling import PlayWrightFetcher
        DYNAMIC_AVAILABLE = True
        DynamicFetcher = PlayWrightFetcher
    except ImportError:
        logger.warning("PlayWrightFetcher not available - run: scrapling install")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    pass


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class ScrapedItem:
    """Container for scraped data"""
    url: str
    title: str = ""
    content: Dict = field(default_factory=dict)
    metadata: Dict = field(default_factory=dict)
    scraped_at: str = field(default_factory=lambda: datetime.now().isoformat())
    success: bool = True
    error: str = ""
    
    def to_dict(self) -> Dict:
        flat = {
            "url": self.url, 
            "title": self.title, 
            "scraped_at": self.scraped_at,
            "success": self.success,
            "error": self.error,
        }
        flat.update(self.content)
        flat.update({f"meta_{k}": v for k, v in self.metadata.items()})
        return flat


@dataclass
class ProxyConfig:
    """Proxy configuration"""
    server: str
    username: str = ""
    password: str = ""
    protocol: str = "http"
    
    def to_url(self) -> str:
        if self.username and self.password:
            return f"{self.protocol}://{self.username}:{self.password}@{self.server}"
        return f"{self.protocol}://{self.server}"
    
    def to_dict(self) -> Dict:
        d = {"server": f"{self.protocol}://{self.server}"}
        if self.username:
            d["username"] = self.username
            d["password"] = self.password
        return d


# ============================================================================
# CAPTCHA SOLVER INTEGRATION
# ============================================================================

class CaptchaSolver:
    """
    CAPTCHA solving integration.
    Supports: 2Captcha, Anti-Captcha, CapMonster
    
    Usage:
        solver = CaptchaSolver(api_key="your_key", service="2captcha")
        solution = solver.solve_recaptcha(site_key, page_url)
    """
    
    def __init__(self, api_key: str = None, service: str = "2captcha"):
        self.api_key = api_key
        self.service = service
        self.enabled = api_key is not None
        
        if self.enabled:
            logger.info(f"CAPTCHA solver enabled: {service}")
    
    def solve_recaptcha_v2(self, site_key: str, page_url: str) -> Optional[str]:
        """Solve reCAPTCHA v2"""
        if not self.enabled:
            logger.warning("CAPTCHA solver not configured")
            return None
        
        try:
            import requests
            
            if self.service == "2captcha":
                # Submit CAPTCHA
                submit_url = "http://2captcha.com/in.php"
                submit_data = {
                    "key": self.api_key,
                    "method": "userrecaptcha",
                    "googlekey": site_key,
                    "pageurl": page_url,
                    "json": 1
                }
                
                response = requests.post(submit_url, data=submit_data)
                result = response.json()
                
                if result.get("status") != 1:
                    logger.error(f"CAPTCHA submit failed: {result}")
                    return None
                
                captcha_id = result.get("request")
                
                # Poll for solution
                result_url = f"http://2captcha.com/res.php?key={self.api_key}&action=get&id={captcha_id}&json=1"
                
                for _ in range(30):  # Max 30 attempts (2.5 minutes)
                    time.sleep(5)
                    response = requests.get(result_url)
                    result = response.json()
                    
                    if result.get("status") == 1:
                        logger.info("CAPTCHA solved successfully")
                        return result.get("request")
                    elif result.get("request") != "CAPCHA_NOT_READY":
                        logger.error(f"CAPTCHA solve failed: {result}")
                        return None
                
                logger.error("CAPTCHA solve timeout")
                return None
                
            elif self.service == "anti-captcha":
                # Anti-Captcha implementation
                submit_url = "https://api.anti-captcha.com/createTask"
                submit_data = {
                    "clientKey": self.api_key,
                    "task": {
                        "type": "RecaptchaV2TaskProxyless",
                        "websiteURL": page_url,
                        "websiteKey": site_key
                    }
                }
                
                response = requests.post(submit_url, json=submit_data)
                result = response.json()
                
                if result.get("errorId") != 0:
                    logger.error(f"CAPTCHA submit failed: {result}")
                    return None
                
                task_id = result.get("taskId")
                
                # Poll for solution
                result_url = "https://api.anti-captcha.com/getTaskResult"
                
                for _ in range(30):
                    time.sleep(5)
                    response = requests.post(result_url, json={
                        "clientKey": self.api_key,
                        "taskId": task_id
                    })
                    result = response.json()
                    
                    if result.get("status") == "ready":
                        logger.info("CAPTCHA solved successfully")
                        return result.get("solution", {}).get("gRecaptchaResponse")
                    elif result.get("errorId") != 0:
                        logger.error(f"CAPTCHA solve failed: {result}")
                        return None
                
                return None
        
        except ImportError:
            logger.error("requests library required for CAPTCHA solving")
            return None
        except Exception as e:
            logger.error(f"CAPTCHA solve error: {e}")
            return None
    
    def solve_hcaptcha(self, site_key: str, page_url: str) -> Optional[str]:
        """Solve hCaptcha"""
        if not self.enabled:
            return None
        
        try:
            import requests
            
            if self.service == "2captcha":
                submit_url = "http://2captcha.com/in.php"
                submit_data = {
                    "key": self.api_key,
                    "method": "hcaptcha",
                    "sitekey": site_key,
                    "pageurl": page_url,
                    "json": 1
                }
                
                response = requests.post(submit_url, data=submit_data)
                result = response.json()
                
                if result.get("status") != 1:
                    return None
                
                captcha_id = result.get("request")
                result_url = f"http://2captcha.com/res.php?key={self.api_key}&action=get&id={captcha_id}&json=1"
                
                for _ in range(30):
                    time.sleep(5)
                    response = requests.get(result_url)
                    result = response.json()
                    
                    if result.get("status") == 1:
                        return result.get("request")
                    elif result.get("request") != "CAPCHA_NOT_READY":
                        return None
                
                return None
                
        except Exception as e:
            logger.error(f"hCaptcha solve error: {e}")
            return None


# ============================================================================
# ROBOTS.TXT COMPLIANCE
# ============================================================================

class RobotsChecker:
    """
    Check robots.txt compliance before scraping.
    
    Usage:
        checker = RobotsChecker()
        if checker.can_fetch("https://example.com/page"):
            # OK to scrape
    """
    
    def __init__(self, user_agent: str = "ScraplingPro/2.0"):
        self.user_agent = user_agent
        self._cache: Dict[str, RobotFileParser] = {}
        self._cache_time: Dict[str, float] = {}
        self.cache_duration = 3600  # 1 hour
    
    def _get_parser(self, url: str) -> RobotFileParser:
        """Get or create robots.txt parser for domain"""
        parsed = urlparse(url)
        domain = f"{parsed.scheme}://{parsed.netloc}"
        
        # Check cache
        now = time.time()
        if domain in self._cache:
            if now - self._cache_time.get(domain, 0) < self.cache_duration:
                return self._cache[domain]
        
        # Fetch robots.txt
        rp = RobotFileParser()
        robots_url = f"{domain}/robots.txt"
        
        try:
            rp.set_url(robots_url)
            rp.read()
            self._cache[domain] = rp
            self._cache_time[domain] = now
        except Exception as e:
            logger.debug(f"Could not fetch robots.txt for {domain}: {e}")
            # Return permissive parser if robots.txt unavailable
            rp = RobotFileParser()
            self._cache[domain] = rp
            self._cache_time[domain] = now
        
        return rp
    
    def can_fetch(self, url: str) -> bool:
        """Check if URL can be fetched according to robots.txt"""
        try:
            rp = self._get_parser(url)
            return rp.can_fetch(self.user_agent, url)
        except Exception:
            return True  # Allow if check fails
    
    def get_crawl_delay(self, url: str) -> Optional[float]:
        """Get crawl delay for domain"""
        try:
            rp = self._get_parser(url)
            delay = rp.crawl_delay(self.user_agent)
            return float(delay) if delay else None
        except Exception:
            return None


# ============================================================================
# WEBHOOK SUPPORT
# ============================================================================

class WebhookSender:
    """
    Send scraped data to webhooks in real-time.
    
    Usage:
        webhook = WebhookSender("https://your-endpoint.com/webhook")
        webhook.send(items)
    """
    
    def __init__(self, url: str, headers: Dict = None, batch_size: int = 100):
        self.url = url
        self.headers = headers or {"Content-Type": "application/json"}
        self.batch_size = batch_size
        self.enabled = url is not None
    
    def send(self, items: List[ScrapedItem]) -> bool:
        """Send items to webhook"""
        if not self.enabled:
            return False
        
        try:
            import requests
            
            # Convert to dicts
            data = [item.to_dict() for item in items]
            
            # Send in batches
            for i in range(0, len(data), self.batch_size):
                batch = data[i:i + self.batch_size]
                
                response = requests.post(
                    self.url,
                    json={"items": batch, "count": len(batch), "timestamp": datetime.now().isoformat()},
                    headers=self.headers,
                    timeout=30
                )
                
                if response.status_code >= 400:
                    logger.error(f"Webhook failed: {response.status_code}")
                    return False
            
            logger.info(f"Sent {len(data)} items to webhook")
            return True
            
        except Exception as e:
            logger.error(f"Webhook error: {e}")
            return False
    
    def send_single(self, item: ScrapedItem) -> bool:
        """Send single item immediately"""
        return self.send([item])


# ============================================================================
# RATE LIMITER
# ============================================================================

class RateLimiter:
    """Per-domain rate limiting"""
    
    def __init__(self, requests_per_second: float = 1.0):
        self.min_interval = 1.0 / requests_per_second if requests_per_second > 0 else 0
        self.last_request: Dict[str, float] = {}
        self.lock = threading.Lock()
    
    def wait(self, url: str):
        """Wait if needed to respect rate limit"""
        domain = urlparse(url).netloc
        
        with self.lock:
            now = time.time()
            last = self.last_request.get(domain, 0)
            elapsed = now - last
            
            if elapsed < self.min_interval:
                sleep_time = self.min_interval - elapsed
                time.sleep(sleep_time)
            
            self.last_request[domain] = time.time()


# ============================================================================
# PROXY MANAGER
# ============================================================================

class ProxyManager:
    """Advanced proxy rotation with health checking"""
    
    def __init__(self, proxies: List[str] = None, rotation_strategy: str = "round_robin"):
        self.proxies = []
        self.proxy_stats = {}
        self.rotation_strategy = rotation_strategy
        self.current_index = 0
        self.lock = threading.Lock()
        
        if proxies:
            for p in proxies:
                self.add_proxy(p)
    
    def add_proxy(self, proxy: str):
        """Add a proxy to the pool"""
        with self.lock:
            if proxy not in self.proxies:
                self.proxies.append(proxy)
                self.proxy_stats[proxy] = {
                    "success": 0, 
                    "fail": 0, 
                    "last_used": 0,
                    "cooldown_until": 0,
                    "avg_response_time": 0,
                }
    
    def remove_proxy(self, proxy: str):
        """Remove a proxy from the pool"""
        with self.lock:
            if proxy in self.proxies:
                self.proxies.remove(proxy)
                del self.proxy_stats[proxy]
    
    def get_proxy(self) -> Optional[str]:
        """Get next proxy based on rotation strategy"""
        with self.lock:
            if not self.proxies:
                return None
            
            now = time.time()
            available = [p for p in self.proxies 
                        if self.proxy_stats[p]["cooldown_until"] < now]
            
            if not available:
                return min(self.proxies, 
                          key=lambda p: self.proxy_stats[p]["cooldown_until"])
            
            if self.rotation_strategy == "round_robin":
                proxy = available[self.current_index % len(available)]
                self.current_index += 1
            elif self.rotation_strategy == "random":
                proxy = random.choice(available)
            elif self.rotation_strategy == "weighted":
                weights = []
                for p in available:
                    stats = self.proxy_stats[p]
                    total = stats["success"] + stats["fail"]
                    weights.append(stats["success"] / total if total > 0 else 1.0)
                proxy = random.choices(available, weights=weights, k=1)[0]
            else:
                proxy = available[0]
            
            self.proxy_stats[proxy]["last_used"] = now
            return proxy
    
    def report_success(self, proxy: str, response_time: float = 0):
        """Report successful request"""
        with self.lock:
            if proxy in self.proxy_stats:
                self.proxy_stats[proxy]["success"] += 1
                if response_time > 0:
                    old_avg = self.proxy_stats[proxy]["avg_response_time"]
                    self.proxy_stats[proxy]["avg_response_time"] = (old_avg * 0.9) + (response_time * 0.1)
    
    def report_failure(self, proxy: str, cooldown_seconds: int = 60):
        """Report failed request"""
        with self.lock:
            if proxy in self.proxy_stats:
                self.proxy_stats[proxy]["fail"] += 1
                self.proxy_stats[proxy]["cooldown_until"] = time.time() + cooldown_seconds


# ============================================================================
# INFINITE SCROLL HANDLER
# ============================================================================

class InfiniteScrollHandler:
    """
    Handle infinite scroll pages.
    
    Usage:
        handler = InfiniteScrollHandler()
        all_items = handler.scroll_and_extract(page, ".item", max_scrolls=10)
    """
    
    def __init__(self, scroll_pause: float = 2.0, scroll_increment: int = 800):
        self.scroll_pause = scroll_pause
        self.scroll_increment = scroll_increment
    
    def scroll_to_bottom(self, page, max_scrolls: int = 10) -> int:
        """
        Scroll page to load all content.
        Returns number of scrolls performed.
        """
        scrolls = 0
        last_height = 0
        
        for _ in range(max_scrolls):
            # Get current height
            current_height = self._execute_js(page, "return document.body.scrollHeight")
            
            if current_height == last_height:
                break  # No more content to load
            
            # Scroll down
            self._execute_js(page, f"window.scrollBy(0, {self.scroll_increment})")
            time.sleep(self.scroll_pause)
            
            last_height = current_height
            scrolls += 1
        
        return scrolls
    
    def scroll_and_collect(
        self, 
        page, 
        item_selector: str, 
        max_scrolls: int = 10,
        max_items: int = None
    ) -> List[Any]:
        """
        Scroll and collect items as they load.
        """
        collected_items = []
        seen_content = set()
        
        for scroll in range(max_scrolls):
            # Get current items
            items = self._safe_css(page, item_selector)
            
            for item in items:
                # Use content hash to avoid duplicates
                content = str(item.html) if hasattr(item, 'html') else str(item)
                content_hash = hash(content)
                
                if content_hash not in seen_content:
                    seen_content.add(content_hash)
                    collected_items.append(item)
                    
                    if max_items and len(collected_items) >= max_items:
                        return collected_items
            
            # Scroll down
            self._execute_js(page, f"window.scrollBy(0, {self.scroll_increment})")
            time.sleep(self.scroll_pause)
            
            # Check if we've reached the bottom
            at_bottom = self._execute_js(
                page,
                "return (window.innerHeight + window.scrollY) >= document.body.scrollHeight - 100"
            )
            if at_bottom:
                break
        
        return collected_items
    
    def click_load_more(
        self,
        page,
        button_selector: str,
        item_selector: str,
        max_clicks: int = 20,
    ) -> List[Any]:
        """
        Click 'Load More' button repeatedly and collect items.
        """
        collected_items = []
        seen_content = set()
        
        for click in range(max_clicks):
            # Collect current items
            items = self._safe_css(page, item_selector)
            
            for item in items:
                content = str(item.html) if hasattr(item, 'html') else str(item)
                content_hash = hash(content)
                
                if content_hash not in seen_content:
                    seen_content.add(content_hash)
                    collected_items.append(item)
            
            # Try to click load more button
            try:
                button = self._safe_css(page, button_selector)
                if not button:
                    break
                
                self._execute_js(page, f"document.querySelector('{button_selector}').click()")
                time.sleep(self.scroll_pause)
                
            except Exception as e:
                logger.debug(f"Load more click failed: {e}")
                break
        
        return collected_items
    
    def _execute_js(self, page, script: str) -> Any:
        """Execute JavaScript on page"""
        try:
            if hasattr(page, 'execute_script'):
                return page.execute_script(script)
            elif hasattr(page, 'evaluate'):
                # Playwright style
                if script.startswith('return '):
                    script = script[7:]  # Remove 'return '
                return page.evaluate(f"() => {script}")
        except Exception as e:
            logger.debug(f"JS execution failed: {e}")
            return None
    
    def _safe_css(self, page, selector: str) -> List:
        """Safely get elements by CSS selector"""
        try:
            if hasattr(page, 'css'):
                return page.css(selector) or []
            elif hasattr(page, 'query_selector_all'):
                return page.query_selector_all(selector) or []
            return []
        except Exception:
            return []


# ============================================================================
# ENHANCED SCRAPING ENGINE
# ============================================================================

class ScrapingEngine:
    """
    Enhanced scraping engine with all features.
    
    Features:
    - Proper timeout handling (seconds)
    - CAPTCHA solving
    - robots.txt compliance
    - Webhook support
    - Infinite scroll
    - Proxy rotation
    - Rate limiting
    """
    
    def __init__(
        self,
        mode: str = "stealthy",
        proxies: List[str] = None,
        proxy_rotation: str = "round_robin",
        rate_limit: float = 1.0,
        max_retries: int = 3,
        timeout: int = 30,  # In SECONDS
        headless: bool = True,
        adaptive: bool = False,
        max_workers: int = 5,
        respect_robots: bool = True,
        captcha_api_key: str = None,
        captcha_service: str = "2captcha",
        webhook_url: str = None,
    ):
        self.mode = mode
        self.max_retries = max_retries
        self.timeout = timeout  # Seconds
        self.timeout_ms = timeout * 1000  # Milliseconds for Playwright
        self.headless = headless
        self.adaptive = adaptive
        self.max_workers = max_workers
        
        # Components
        self.proxy_manager = ProxyManager(proxies, proxy_rotation) if proxies else None
        self.rate_limiter = RateLimiter(rate_limit)
        self.robots_checker = RobotsChecker() if respect_robots else None
        self.captcha_solver = CaptchaSolver(captcha_api_key, captcha_service)
        self.webhook = WebhookSender(webhook_url) if webhook_url else None
        self.scroll_handler = InfiniteScrollHandler()
        
        # Stats
        self.stats = {
            "total_requests": 0,
            "successful": 0,
            "failed": 0,
            "captchas_solved": 0,
            "robots_blocked": 0,
        }
        
        # Enable adaptive mode
        if adaptive and STEALTHY_AVAILABLE:
            StealthyFetcher.adaptive = True
        
        logger.info(f"Engine initialized: mode={mode}, timeout={timeout}s, retries={max_retries}")
    
    def _check_robots(self, url: str) -> bool:
        """Check if URL is allowed by robots.txt"""
        if not self.robots_checker:
            return True
        
        allowed = self.robots_checker.can_fetch(url)
        if not allowed:
            logger.warning(f"Blocked by robots.txt: {url}")
            self.stats["robots_blocked"] += 1
        return allowed
    
    def _get_fetcher_kwargs(self) -> Dict:
        """Build kwargs for fetcher"""
        kwargs = {}
        
        if self.proxy_manager:
            proxy = self.proxy_manager.get_proxy()
            if proxy:
                kwargs["proxy"] = proxy
        
        return kwargs
    
    def fetch(self, url: str, **extra_kwargs) -> Any:
        """
        Fetch a single URL with retry logic.
        
        Args:
            url: URL to fetch
            **extra_kwargs: Additional kwargs for fetcher
            
        Returns:
            Page object or None on failure
        """
        # Check robots.txt
        if not self._check_robots(url):
            return None
        
        self.stats["total_requests"] += 1
        
        for attempt in range(self.max_retries):
            try:
                # Rate limiting
                self.rate_limiter.wait(url)
                
                # Apply crawl delay from robots.txt if available
                if self.robots_checker:
                    crawl_delay = self.robots_checker.get_crawl_delay(url)
                    if crawl_delay:
                        time.sleep(crawl_delay)
                
                # Get proxy and build kwargs
                kwargs = self._get_fetcher_kwargs()
                kwargs.update(extra_kwargs)
                proxy_used = kwargs.get("proxy")
                
                # Time the request
                start_time = time.time()
                
                # Fetch based on mode
                if self.mode == "dynamic" and DYNAMIC_AVAILABLE:
                    # Use PlayWrightFetcher with proper timeout (milliseconds)
                    page = DynamicFetcher.fetch(
                        url,
                        headless=self.headless,
                        timeout=self.timeout_ms,  # Already in milliseconds
                        network_idle=True,  # Wait for network to be idle
                        **{k: v for k, v in kwargs.items() if k not in ["timeout"]}
                    )
                elif STEALTHY_AVAILABLE:
                    # StealthyFetcher ALSO expects milliseconds!
                    page = StealthyFetcher.fetch(
                        url, 
                        timeout=self.timeout_ms,  # MUST be milliseconds
                        network_idle=True,
                        **{k: v for k, v in kwargs.items() if k not in ["timeout"]}
                    )
                else:
                    raise RuntimeError("No fetcher available. Install scrapling: pip install scrapling[all]")
                
                response_time = time.time() - start_time
                
                # Check success
                if page.status == 200:
                    self.stats["successful"] += 1
                    if proxy_used and self.proxy_manager:
                        self.proxy_manager.report_success(proxy_used, response_time)
                    logger.info(f"Fetched ({page.status}) {url} in {response_time:.2f}s")
                    return page
                
                elif page.status == 429:  # Rate limited
                    wait_time = (attempt + 1) * 30
                    logger.warning(f"Rate limited on {url}, waiting {wait_time}s")
                    time.sleep(wait_time)
                
                elif page.status == 403:  # Possibly blocked
                    logger.warning(f"HTTP 403 for {url} - might be blocked")
                    if proxy_used and self.proxy_manager:
                        self.proxy_manager.report_failure(proxy_used)
                
                else:
                    logger.warning(f"HTTP {page.status} for {url}")
                    if proxy_used and self.proxy_manager:
                        self.proxy_manager.report_failure(proxy_used)
                        
            except Exception as e:
                error_msg = str(e)
                logger.warning(f"Attempt {attempt + 1} failed: {error_msg[:100]}. Retrying in {2 ** attempt}s...")
                
                if proxy_used and self.proxy_manager:
                    self.proxy_manager.report_failure(proxy_used)
                
                if attempt < self.max_retries - 1:
                    time.sleep(2 ** attempt)  # Exponential backoff
        
        logger.error(f"Failed after {self.max_retries} attempts: {url}")
        self.stats["failed"] += 1
        return None
    
    def fetch_with_scroll(
        self, 
        url: str, 
        item_selector: str,
        max_scrolls: int = 10,
        max_items: int = None,
        **extra_kwargs
    ) -> List[Any]:
        """
        Fetch a page with infinite scroll support.
        
        Returns list of elements matching item_selector.
        """
        page = self.fetch(url, **extra_kwargs)
        
        if not page:
            return []
        
        return self.scroll_handler.scroll_and_collect(
            page, item_selector, max_scrolls, max_items
        )
    
    def fetch_many(self, urls: List[str], concurrent: bool = True) -> List[Any]:
        """Fetch multiple URLs, optionally concurrently"""
        if concurrent:
            pages = []
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {executor.submit(self.fetch, url): url for url in urls}
                for future in as_completed(futures):
                    pages.append(future.result())
            return pages
        else:
            return [self.fetch(url) for url in urls]
    
    def scrape_url(
        self, 
        url: str, 
        parser: Callable,
        **fetch_kwargs
    ) -> List[ScrapedItem]:
        """
        Scrape a single URL using the provided parser function.
        """
        page = self.fetch(url, **fetch_kwargs)
        
        if not page:
            return [ScrapedItem(url=url, success=False, error="Failed to fetch")]
        
        try:
            items = parser(page)
            
            # Send to webhook if configured
            if self.webhook and items:
                self.webhook.send(items)
            
            return items
        except Exception as e:
            logger.error(f"Parser error for {url}: {e}")
            return [ScrapedItem(url=url, success=False, error=str(e))]
    
    def scrape_urls(
        self, 
        urls: List[str], 
        parser: Callable,
        concurrent: bool = True
    ) -> List[ScrapedItem]:
        """Scrape multiple URLs"""
        all_items = []
        
        if concurrent:
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {executor.submit(self.scrape_url, url, parser): url for url in urls}
                for future in as_completed(futures):
                    all_items.extend(future.result())
        else:
            for url in urls:
                all_items.extend(self.scrape_url(url, parser))
        
        return all_items
    
    def scrape_with_pagination(
        self,
        start_url: str,
        parser: Callable,
        pagination_type: str = "url",  # "url", "click", "scroll"
        url_pattern: str = "?page={}",
        max_pages: int = 10,
        **kwargs
    ) -> List[ScrapedItem]:
        """
        Scrape multiple pages with pagination support.
        
        pagination_type:
        - "url": URL-based pagination (page=1, page=2, etc.)
        - "scroll": Infinite scroll
        - "click": Click-based pagination (requires dynamic mode)
        """
        all_items = []
        
        if pagination_type == "url":
            # Generate page URLs
            urls = [start_url]
            for page_num in range(2, max_pages + 1):
                if "{}" in url_pattern:
                    page_url = start_url + url_pattern.format(page_num)
                else:
                    page_url = start_url + url_pattern + str(page_num)
                urls.append(page_url)
            
            # Scrape all pages
            for url in urls:
                items = self.scrape_url(url, parser, **kwargs)
                if not items or not any(item.success for item in items):
                    break  # Stop if page has no items
                all_items.extend(items)
                logger.info(f"Page scraped: {len(items)} items from {url}")
        
        elif pagination_type == "scroll":
            # Use infinite scroll
            page = self.fetch(start_url, **kwargs)
            if page:
                # Scroll and collect
                elements = self.scroll_handler.scroll_and_collect(
                    page, kwargs.get("item_selector", ".item"), max_pages * 10
                )
                # Parse elements
                for el in elements:
                    try:
                        item = parser(el)
                        if isinstance(item, list):
                            all_items.extend(item)
                        else:
                            all_items.append(item)
                    except Exception as e:
                        logger.error(f"Parse error: {e}")
        
        return all_items
    
    def get_stats(self) -> Dict:
        """Get scraping statistics"""
        return {
            **self.stats,
            "success_rate": self.stats["successful"] / max(self.stats["total_requests"], 1) * 100,
        }


# ============================================================================
# EXPORTER
# ============================================================================

class Exporter:
    """Export scraped data to various formats"""
    
    @staticmethod
    def to_csv(items: List[ScrapedItem], filepath: str):
        """Export to CSV"""
        if not items:
            logger.warning("No items to export")
            return
        
        flat_items = [item.to_dict() for item in items]
        fieldnames = list(flat_items[0].keys())
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(flat_items)
        
        logger.info(f"Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_json(items: List[ScrapedItem], filepath: str):
        """Export to JSON"""
        flat_items = [item.to_dict() for item in items]
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(flat_items, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Exported {len(items)} items to {filepath}")
    
    @staticmethod
    def to_excel(items: List[ScrapedItem], filepath: str):
        """Export to Excel"""
        if not EXCEL_AVAILABLE:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            Exporter.to_csv(items, filepath.replace('.xlsx', '.csv'))
            return
        
        flat_items = [item.to_dict() for item in items]
        fieldnames = list(flat_items[0].keys()) if flat_items else []
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scraped Data"
        
        # Header row
        for col, field in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Data rows
        for row_num, item in enumerate(flat_items, 2):
            for col, field in enumerate(fieldnames, 1):
                value = item.get(field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(50, max_length + 2)
        
        wb.save(filepath)
        logger.info(f"Exported {len(items)} items to {filepath}")


# ============================================================================
# QUICK API
# ============================================================================

def quick_scrape(
    url: str,
    item_selector: str,
    fields: Dict[str, str],
    max_pages: int = 1,
    url_pattern: str = "",
    mode: str = "stealthy",
    timeout: int = 30,
    **engine_kwargs
) -> List[ScrapedItem]:
    """
    Quick one-liner scraping API.
    
    Example:
        items = quick_scrape(
            url="https://books.toscrape.com",
            item_selector="article.product_pod",
            fields={
                "title": "h3 a @title",
                "price": ".price_color",
            }
        )
    """
    
    def parser(page):
        items = []
        for element in page.css(item_selector):
            content = {}
            title = ""
            
            for field_name, selector in fields.items():
                if " @" in selector:
                    sel, attr = selector.rsplit(" @", 1)
                    els = element.css(sel)
                    value = els[0].attrib.get(attr, "") if els else ""
                else:
                    els = element.css(selector)
                    value = els[0].text.strip() if els and hasattr(els[0], 'text') else ""
                
                if field_name == "title":
                    title = value
                else:
                    content[field_name] = value
            
            items.append(ScrapedItem(url=url, title=title, content=content))
        
        return items
    
    engine = ScrapingEngine(mode=mode, timeout=timeout, **engine_kwargs)
    
    if max_pages > 1 and url_pattern:
        return engine.scrape_with_pagination(url, parser, "url", url_pattern, max_pages=max_pages)
    else:
        return engine.scrape_url(url, parser)


# ============================================================================
# PARSERS
# ============================================================================

class Parsers:
    """Ready-to-use parser templates"""
    
    @staticmethod
    def ecommerce_products(page) -> List[ScrapedItem]:
        """Parse e-commerce product listings"""
        items = []
        
        selectors = [
            "article.product_pod", ".product-card", ".product-item",
            ".product", "[data-product]", ".item"
        ]
        
        for selector in selectors:
            products = page.css(selector)
            if products:
                for product in products:
                    title_el = product.css("h2, h3, .title, .name, h3 a")
                    price_el = product.css(".price, .price_color, .cost")
                    img_el = product.css("img")
                    link_el = product.css("a")
                    
                    title = ""
                    if title_el:
                        title = title_el[0].attrib.get("title") or getattr(title_el[0], 'text', '') or ""
                    
                    items.append(ScrapedItem(
                        url=str(page.url) if hasattr(page, 'url') else "",
                        title=str(title).strip(),
                        content={
                            "price": price_el[0].text.strip() if price_el and hasattr(price_el[0], 'text') else "",
                            "image": img_el[0].attrib.get("src", "") if img_el else "",
                            "link": link_el[0].attrib.get("href", "") if link_el else "",
                        }
                    ))
                break
        
        return items


# ============================================================================
# MAIN
# ============================================================================

if __name__ == "__main__":
    print("🕷️ Scrapling Pro - Enhanced Engine v2")
    print("=" * 50)
    print(f"StealthyFetcher: {'✅ Available' if STEALTHY_AVAILABLE else '❌ Not available'}")
    print(f"DynamicFetcher:  {'✅ Available' if DYNAMIC_AVAILABLE else '❌ Not available'}")
    print(f"Excel Export:    {'✅ Available' if EXCEL_AVAILABLE else '❌ Not available'}")
    
    print("\nNew Features:")
    print("  ✅ Proper timeout handling (seconds → ms)")
    print("  ✅ CAPTCHA solving (2Captcha, Anti-Captcha)")
    print("  ✅ robots.txt compliance")
    print("  ✅ Webhook support")
    print("  ✅ Infinite scroll handling")
    print("  ✅ Enhanced proxy rotation")
